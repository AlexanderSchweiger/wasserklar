import os
import re
from datetime import date, datetime, timedelta
from decimal import Decimal

from flask import (
    render_template, redirect, url_for, flash, request,
    current_app, send_file, abort,
)
from flask_login import login_required, current_user

from app.invoices import bp
from app.invoices.send_email_hooks import run_before_send, read_message_id
from app.invoices.render_hooks import build_pdf_context
from app.extensions import db
from app.models import Invoice, InvoiceItem, EmailEvent, Customer, WaterMeter, MeterReading, WaterTariff, Booking, Account, Property, OpenItem, Project, RealAccount, InvoiceCounter, AppSetting, BillingRun, BillingPeriod, ReadingCorrection
from app.meters.estimation import apply_corrections_to_invoice, cap_invoice_at_zero, reverse_corrections_for_invoice
from app.email_tracking import record_email_sent
from app.utils import next_invoice_number as _next_invoice_number
from app.settings_service import get_wg, send_mail, wg_settings, get_contact_info, get_contact_info_font_size, get_invoice_sender_address
from app.invoices.design import get_design
from app.pagination import paginate_query


def _current_design():
    """Liest das aktuell konfigurierte Rechnungsdesign aus den AppSettings."""
    return get_design(AppSetting.get("invoice.design", "classic"))


def _render_pdf_html(invoice, *, for_email=False):
    """Rendert die HTML-Vorlage für WeasyPrint mit aktuellem Design.

    ``for_email``: True, wenn das PDF als E-Mail-Anhang erzeugt wird. Damit
    koennen Provider Inhalte unterdruecken, die nur auf der gedruckten Rechnung
    Sinn ergeben (z.B. der „Rechnung per E-Mail?"-Block).

    Das Design kann ein eigenes Template ueber den Schluessel ``template``
    vorgeben (z.B. das SaaS-„wasserklar"-Design); sonst die OSS-Standardvorlage.
    """
    design = _current_design()
    template_name = design.get("template", "invoices/pdf_template.html")
    extra = build_pdf_context(invoice, for_email=for_email)
    return render_template(
        template_name,
        invoice=invoice,
        design=design,
        contact_info=get_contact_info(),
        contact_info_font_size=get_contact_info_font_size(),
        invoice_sender_address=get_invoice_sender_address(),
        for_email=for_email,
        **extra,
    )


def _resolve_open_item_account_id(invoice, form_account_id=None):
    """Ermittelt das Buchungskonto für einen aus einer Rechnung erzeugten Offenen Posten.

    Das Konto wird ausschließlich pro Offenem Posten gefuehrt (``OpenItem.account_id``);
    der Rechnungslauf traegt kein eigenes Konto mehr. Es zaehlt daher allein der
    Formularwert; ist keiner gesetzt, bleibt das Konto offen (``None``).
    """
    return form_account_id


# Nach app/invoices/services.py extrahiert (wiederverwendet von den
# Zaehlertausch-Touren); Alias-Rückbindung haelt alle bestehenden Aufrufer
# und Test-Monkeypatches auf die Underscore-Namen stabil.
from app.invoices.services import (  # noqa: E402
    invoice_period_year as _invoice_period_year,
    create_or_update_open_item as _create_or_update_open_item,
)


def _status_response(invoice):
    """Einheitliche Antwort für ``set_status``: Fragment bei HTMX, sonst Redirect."""
    if request.headers.get("HX-Request"):
        return render_template("invoices/_status_badge.html", invoice=invoice)
    return redirect(url_for("invoices.detail", invoice_id=invoice.id))


def _book_invoice_payment_if_needed(invoice):
    """Legt beim Markieren als 'Bezahlt' eine Zahlungsbuchung an – aber nur, wenn
    noch keine wirksame Buchung für die Rechnung existiert.

    Verhindert Doppelbuchungen, wenn die Rechnung bereits über die Zahlungs-
    erfassung (``pay``) oder einen früheren Bezahlt-Wechsel verbucht wurde.
    """
    from app.accounting import services as acc_svc
    existing = (
        Booking.query
        .filter(Booking.invoice_id == invoice.id)
        .filter(acc_svc.storno_filter())
        .first()
    )
    if existing is not None:
        # Schon verbucht – nur den Offenen Posten schließen, nicht erneut buchen.
        if invoice.open_item:
            invoice.open_item.status = OpenItem.STATUS_PAID
        return

    default_ra = RealAccount.query.filter_by(is_default=True, active=True).first() \
        or RealAccount.query.filter_by(active=True).first()
    real_account_id = default_ra.id if default_ra else None
    # Bei mehreren Dimensionen (Konto/Projekt/Steuersatz) entsteht automatisch
    # eine Sammelbuchung (ADR-002), sonst eine Einzelbuchung.
    acc_svc.booking_group_from_invoice_payment(
        invoice=invoice,
        amount=invoice.total_amount,
        payment_date=date.today(),
        real_account_id=real_account_id,
        created_by_id=current_user.id,
        open_item=invoice.open_item,
        reference=invoice.invoice_number,
    )
    if invoice.open_item:
        invoice.open_item.status = OpenItem.STATUS_PAID


def _billing_period_list():
    """Alle Abrechnungsperioden, neueste zuerst (fuer Auswahl-Dropdowns)."""
    return BillingPeriod.query.order_by(
        BillingPeriod.start_date.desc(), BillingPeriod.id.desc()
    ).all()


def _invoice_is_locked(invoice):
    """Gesperrt wenn Status nicht mehr Entwurf ist."""
    return invoice.status != Invoice.STATUS_DRAFT


def _get_document_format(override=None):
    """Gibt das konfigurierte Dokumentformat zurück ('pdf', 'docx' oder 'both').
    override kommt aus dem Request-Parameter ?fmt="""
    fmt = override or AppSetting.get("invoice.document_format", "pdf")
    return fmt if fmt in ("pdf", "docx", "both") else "pdf"


def _get_doc_dir(invoice):
    """Gibt den jahresspezifischen Unterordner für Rechnungsdokumente zurück und legt ihn an.

    Struktur: <PDF_DIR>/<Jahr>/ z.B. instance/pdfs/2024/
    """
    year = invoice.date.year if invoice.date else "misc"
    doc_dir = os.path.join(current_app.config["PDF_DIR"], str(year))
    os.makedirs(doc_dir, exist_ok=True)
    return doc_dir


def _versioned_path(doc_dir: str, invoice_number: str, ext: str) -> str:
    """Gibt einen eindeutigen Dateipfad zurück.

    Existiert bereits eine Datei mit dem Basisnamen, wird _V2, _V3, … angehängt,
    damit ältere Versionen erhalten bleiben.

    Beispiel: 2025-00042.pdf → 2025-00042_V2.pdf → 2025-00042_V3.pdf
    """
    base = os.path.join(doc_dir, f"{invoice_number}.{ext}")
    if not os.path.exists(base):
        return base
    v = 2
    while True:
        candidate = os.path.join(doc_dir, f"{invoice_number}_V{v}.{ext}")
        if not os.path.exists(candidate):
            return candidate
        v += 1


_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


def _sync_open_item(invoice):
    """Passt den Betrag des verknüpften Offenen Postens an den neuen Rechnungsbetrag an.
    Gibt True zurück wenn ein Offener Posten aktualisiert wurde."""
    if invoice.open_item:
        invoice.open_item.amount = invoice.total_amount
        return True
    return False


# Standard-Betreff fuer den Rechnungsversand. WG-Name zuerst, damit der Kunde
# den Absender erkennt — Mails mit nackter Rechnungsnummer im Betreff wirken
# unserioes. Jinja-Platzhalter wie in der Body-Vorlage.
DEFAULT_INVOICE_EMAIL_SUBJECT = "{{ wg_name }} - Rechnung {{ rechnungsnummer }}"


def _email_template_context(invoice):
    """Gemeinsame Platzhalter fuer Betreff- und Body-Vorlage des Rechnungsmails."""
    return dict(
        name=invoice.customer.letter_name,
        rechnungsnummer=invoice.invoice_number,
        buchungsjahr=(invoice.billing_period.name if invoice.billing_period else ""),
        betrag=f"{invoice.total_amount:.2f}",
        faelligkeitsdatum=(
            invoice.due_date.strftime("%d.%m.%Y") if invoice.due_date else "—"
        ),
        iban=get_wg('iban'),
        wg_name=get_wg('name'),
    )


def _render_email_subject(invoice):
    """Rendert den E-Mail-Betreff: DB-Vorlage wenn vorhanden, sonst Standard.

    Faellt bei kaputter/leerer Vorlage auf einen simplen Default zurueck, damit
    ein Tippfehler in der Betreff-Vorlage nicht den gesamten Versand blockiert.
    """
    from jinja2 import Environment
    tpl = AppSetting.get("email_subject_template") or DEFAULT_INVOICE_EMAIL_SUBJECT
    try:
        rendered = (
            Environment().from_string(tpl).render(**_email_template_context(invoice)).strip()
        )
    except Exception:
        rendered = ""
    return rendered or f"Rechnung {invoice.invoice_number}"


def _render_email_body(invoice):
    """Rendert den E-Mail-Text: DB-Vorlage wenn vorhanden, sonst statisches Template."""
    from jinja2 import Environment
    custom = AppSetting.get("email_body_template")
    if custom:
        return Environment().from_string(custom).render(**_email_template_context(invoice))
    return render_template("invoices/email_body.txt", invoice=invoice)


_SORT_KEYS = {"nr", "kunde", "objekt", "datum", "faellig", "betrag", "status"}
_DEFAULT_SORT = "nr"


@bp.route("/")
@login_required
def index():
    status_filter = request.args.get("status", "")
    period_filter = request.args.get("period_id", "", type=str)
    date_from = request.args.get("date_from", date(date.today().year, 1, 1).isoformat()).strip()
    date_to = request.args.get("date_to", "").strip()
    q = request.args.get("q", "").strip()
    project_id_filter = request.args.get("project_id", "", type=str)
    mail_filter = request.args.get("mail_filter", "")
    sort = request.args.get("sort", _DEFAULT_SORT)
    if sort not in _SORT_KEYS:
        sort = _DEFAULT_SORT
    sort_dir = request.args.get("dir", "asc")
    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"

    query = (
        Invoice.query
        .join(Customer, Invoice.customer_id == Customer.id)
        .outerjoin(Property, Invoice.property_id == Property.id)
    )
    if status_filter:
        query = query.filter(Invoice.status == status_filter)
    if period_filter:
        query = query.filter(Invoice.billing_period_id == int(period_filter))
    if date_from:
        query = query.filter(Invoice.date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(Invoice.date <= date.fromisoformat(date_to))
    if q:
        from sqlalchemy import or_
        query = query.filter(or_(
            Customer.name.ilike(f"%{q}%"),
            Invoice.invoice_number.ilike(f"%{q}%"),
            Property.object_number.ilike(f"%{q}%"),
            Property.strasse.ilike(f"%{q}%"),
        ))
    if project_id_filter:
        from sqlalchemy import exists
        query = query.filter(
            exists().where(
                (Booking.invoice_id == Invoice.id) &
                (Booking.project_id == int(project_id_filter))
            )
        )
    if mail_filter == "mail":
        query = query.filter(Customer.rechnung_per_email == True)
    elif mail_filter == "post":
        query = query.filter(Customer.rechnung_per_email != True)

    query = _apply_invoice_sort(query, sort, sort_dir)
    pagination = paginate_query(query, page_key="invoices")
    invoices = pagination.items
    projects_for_filter = Project.query.order_by(Project.name).all()
    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    doc_format = AppSetting.get("invoice.document_format", "pdf")
    sort_ctx = dict(
        sort=sort, dir=sort_dir,
        status_filter=status_filter, period_filter=period_filter,
        date_from=date_from, date_to=date_to,
        q=q, project_id_filter=project_id_filter, mail_filter=mail_filter,
    )
    if request.headers.get("HX-Request"):
        return render_template(
            "invoices/_table.html",
            invoices=invoices, doc_format=doc_format, pagination=pagination,
            **sort_ctx,
        )
    return render_template(
        "invoices/index.html",
        invoices=invoices,
        statuses=Invoice.ALL_STATUSES,
        periods=BillingPeriod.query.order_by(
            BillingPeriod.start_date.desc(), BillingPeriod.id.desc()
        ).all(),
        projects_for_filter=projects_for_filter,
        doc_format=doc_format,
        accounts=accounts,
        pagination=pagination,
        **sort_ctx,
    )


@bp.route("/corrections")
@login_required
def corrections():
    """Übersicht der Schätz-Korrekturposten (Gutschrift/Nachforderung).

    Offene Posten warten auf den naechsten Rechnungslauf des Kunden; verrechnete
    sind in einer Folgerechnung eingezogen. Reine Transparenz-Ansicht — der
    Einzug passiert automatisch im Rechnungslauf."""
    status = request.args.get("status", "open")
    customer_id = request.args.get("customer_id", type=int)
    query = (
        ReadingCorrection.query
        .join(Customer, ReadingCorrection.customer_id == Customer.id)
    )
    if customer_id:
        query = query.filter(ReadingCorrection.customer_id == customer_id)
    if status == "open":
        query = query.filter(ReadingCorrection.status.in_([
            ReadingCorrection.STATUS_OPEN, ReadingCorrection.STATUS_PARTIAL,
        ]))
    elif status == "applied":
        query = query.filter(
            ReadingCorrection.status == ReadingCorrection.STATUS_APPLIED)
    query = query.order_by(
        ReadingCorrection.created_at.desc(), ReadingCorrection.id.desc())
    pagination = paginate_query(query, page_key="corrections")

    open_q = ReadingCorrection.query.filter(
        ReadingCorrection.status.in_([
            ReadingCorrection.STATUS_OPEN, ReadingCorrection.STATUS_PARTIAL,
        ])
    )
    if customer_id:
        open_q = open_q.filter(ReadingCorrection.customer_id == customer_id)
    open_total = sum(
        (Decimal(str(c.remaining_amount or 0)) for c in open_q.all()), Decimal("0"))

    filter_customer = db.session.get(Customer, customer_id) if customer_id else None

    return render_template(
        "invoices/corrections.html",
        corrections=pagination.items, pagination=pagination,
        status=status, open_total=open_total, filter_customer=filter_customer,
    )


def _apply_invoice_sort(query, sort: str, direction: str):
    desc = direction == "desc"

    def order(col):
        return (col.desc(),) if desc else (col.asc(),)

    if sort == "kunde":
        return query.order_by(*order(Customer.name), Invoice.invoice_number.asc())
    if sort == "objekt":
        from sqlalchemy import case as sa_case
        null_last = sa_case((Property.object_number.is_(None), 1), else_=0).asc()
        return query.order_by(null_last, *order(Property.object_number))
    if sort == "datum":
        return query.order_by(*order(Invoice.date), Invoice.invoice_number.asc())
    if sort == "faellig":
        from sqlalchemy import case as sa_case
        null_last = sa_case((Invoice.due_date.is_(None), 1), else_=0).asc()
        return query.order_by(null_last, *order(Invoice.due_date))
    if sort == "betrag":
        return query.order_by(*order(Invoice.total_amount))
    if sort == "status":
        return query.order_by(*order(Invoice.status), Invoice.invoice_number.asc())
    # Default: sort == "nr"
    return query.order_by(*order(Invoice.invoice_number))


_HAUSNUMMER_RE = re.compile(r"(\d+)")


def _billing_run_sort_key(prop_readings, sort_order):
    """Sort-Key für einen property_readings-Eintrag im Rechnungslauf.

    Fehlende Werte (kein Besitzer, keine Nummer, kein Ort) kommen ans Ende.
    """
    prop = prop_readings[0].meter.property
    ownership = prop.current_owner()
    customer = db.session.get(Customer, ownership.customer_id) if ownership else None

    if sort_order == "customer_name":
        return ((customer.name.casefold() if customer else "\xff"),)

    if sort_order == "customer_number":
        cn = customer.customer_number if customer else None
        return (1 if cn is None else 0, cn if cn is not None else 999_999_999)

    if sort_order == "object_number":
        on = prop.object_number or ""
        m = _HAUSNUMMER_RE.match(on)
        num = int(m.group(1)) if m else 999_999_999
        rest = on[m.end():].casefold() if m else on.casefold()
        return (1 if not on else 0, num, rest)

    # "address": Straße alphabetisch, dann Hausnummer numerisch (Zusätze ignorieren)
    strasse = (prop.strasse or "\xff").casefold()
    hn = prop.hausnummer or ""
    m = _HAUSNUMMER_RE.search(hn)
    hn_num = int(m.group(1)) if m else 999_999_999
    return (strasse, hn_num)


@bp.route("/generate", methods=["GET", "POST"])
@login_required
def generate():
    """Massenrechnungslauf: Alle Kunden mit Ablesung für ein Jahr."""
    from app.accounting import services as acc_svc
    tariffs = WaterTariff.query.order_by(WaterTariff.valid_from.desc()).all()
    if request.method == "POST":
        period = db.session.get(
            BillingPeriod, request.form.get("billing_period_id", type=int)
        )
        if period is None:
            flash("Bitte eine Abrechnungsperiode für den Rechnungslauf auswählen.", "danger")
            return redirect(url_for("invoices.generate"))
        tariff_id = int(request.form["tariff_id"])
        tariff = db.get_or_404(WaterTariff, tariff_id)
        due_days = int(request.form.get("due_days", 30))
        valid_sort_orders = {k for k, _ in BillingRun.SORT_ORDER_CHOICES}
        sort_order = request.form.get("sort_order", "customer_name")
        if sort_order not in valid_sort_orders:
            sort_order = "customer_name"

        # Rechnungsdatum für den Rechnungslauf ist heute.
        invoice_date = date.today()
        fy_error = acc_svc.open_fiscal_year_error(invoice_date)
        if fy_error:
            flash(f"{fy_error} Rechnungslauf nicht möglich.", "danger")
            return redirect(url_for("invoices.generate"))

        # Standard-Wasser-Steuersatz nur in USt-pflichtigen Buchungsjahren anwenden
        water_tax = acc_svc.default_water_tax_rate(invoice_date.year)

        # Alle Ablesungen der Periode holen (inkl. ausgebauter Zähler)
        readings = (
            MeterReading.query
            .filter_by(billing_period_id=period.id)
            .join(WaterMeter)
            .join(Property)
            .filter(Property.active == True)
            .all()
        )

        # Nach Objekt gruppieren (ein Objekt kann mehrere Zähler in der Periode haben)
        from collections import defaultdict
        property_readings = defaultdict(list)
        for reading in readings:
            property_readings[reading.meter.property_id].append(reading)

        # Rechnungslauf anlegen (Tarif-Snapshot + Metadaten)
        billing_run = BillingRun(
            billing_period_id=period.id,
            created_by_id=current_user.id,
            tariff_name=tariff.name,
            tariff_valid_from=tariff.valid_from,
            tariff_valid_to=tariff.valid_to,
            tariff_base_fee=tariff.base_fee,
            tariff_base_fee_label=tariff.base_fee_label or "Grundgebühr",
            tariff_additional_fee=tariff.additional_fee,
            tariff_additional_fee_label=tariff.additional_fee_label or "Zusatzgebühr",
            tariff_price_per_m3=tariff.price_per_m3,
            tariff_notes=tariff.notes,
            sort_order=sort_order,
        )
        db.session.add(billing_run)
        db.session.flush()

        sorted_items = sorted(
            property_readings.items(),
            key=lambda kv: _billing_run_sort_key(kv[1], sort_order),
        )

        # Unterjaehrige Schlussrechnungen (Eigentuemerwechsel) je Objekt:
        # deren bereits verrechnete Mengen/Gebuehren-Tage werden unten vom
        # Jahresverbrauch des Nachbesitzers abgezogen.
        from app.owner_change.services import deductions_for_property

        created = 0
        skipped = 0
        # Objekte, bei denen die Schlussrechnung die Jahresmenge uebersteigt
        # (Rest auf 0 gekappt) — Sammel-Warnung nach dem Lauf.
        clamp_warnings = []
        # Pro Kunde nur EINE Rechnung im Lauf mit offenen Schaetz-Korrekturen
        # belasten/gutschreiben (ein Kunde kann mehrere Objekte haben).
        customers_corrected = set()
        for property_id, prop_readings in sorted_items:
            prop = prop_readings[0].meter.property
            ownership = prop.current_owner()
            if not ownership:
                skipped += 1
                continue

            # Bereits vorhanden? Nur regulaere (standard) Rechnungen blockieren
            # den Lauf — eine Schlussrechnung (final_settlement) tut das nicht.
            exists = Invoice.query.filter(
                Invoice.property_id == prop.id,
                Invoice.billing_period_id == period.id,
                Invoice.invoice_kind == Invoice.KIND_STANDARD,
            ).first()
            if exists:
                skipped += 1
                continue

            # Abzuege aus unterjaehrigen Schlussrechnungen dieses Objekts.
            ded = deductions_for_property(prop.id, period.id)

            customer = db.session.get(Customer, ownership.customer_id)

            # Priorität: Objekt > Kunde > Tarif
            # None bedeutet: keine Gebühr, keine Rechnungsposition
            effective_base_fee = (
                prop.base_fee_override if prop.base_fee_override is not None
                else customer.base_fee_override if customer.base_fee_override is not None
                else tariff.base_fee
            )
            effective_additional_fee = (
                prop.additional_fee_override if prop.additional_fee_override is not None
                else customer.additional_fee_override if customer.additional_fee_override is not None
                else tariff.additional_fee
            )
            base_fee_label = tariff.base_fee_label or "Grundgebühr"
            additional_fee_label = tariff.additional_fee_label or "Zusatzgebühr"

            inv = Invoice(
                invoice_number=_next_invoice_number(invoice_date.year),
                customer_id=ownership.customer_id,
                property_id=prop.id,
                billing_run_id=billing_run.id,
                billing_period_id=period.id,
                invoice_kind=Invoice.KIND_STANDARD,
                date=date.today(),
                due_date=date.today() + timedelta(days=due_days),
                status=Invoice.STATUS_DRAFT,
                created_by_id=current_user.id,
            )
            db.session.add(inv)
            db.session.flush()

            # Verbrauchspositionen (immer an erster Stelle)
            is_replacement = len(prop_readings) > 1
            print_meter_swap = AppSetting.get('invoice.print_meter_swap') == 'true'
            total_consumption = sum(
                (r.consumption or Decimal("0")) for r in prop_readings
            )
            # Schaetzung-Marker: beruht der Verbrauch (ganz/teils) auf einer
            # geschaetzten Ablesung? -> Position bekommt den "geschätzt"-Badge.
            is_any_estimated = any(
                getattr(r, "is_estimated", False) for r in prop_readings
            )

            # Bereits per Schlussrechnung verrechnete Mengen (Eigentuemer-
            # wechsel) abziehen; Rest nie negativ (auf 0 kappen + Warnung).
            ded_by_meter = ded["by_meter"] if ded else {}
            ded_total = ded["total"] if ded else Decimal("0")
            ded_numbers = ded["invoice_numbers"] if ded else []

            if is_replacement and print_meter_swap:
                # Separate Zeile je Zähler
                for reading in prop_readings:
                    consumption = reading.consumption or Decimal("0")
                    deducted = ded_by_meter.get(reading.meter_id, Decimal("0"))
                    if deducted > 0:
                        consumption = consumption - deducted
                        if consumption < 0:
                            clamp_warnings.append(prop.label())
                            consumption = Decimal("0")
                    meter = reading.meter
                    if meter.installed_to:
                        date_hint = f"ausgebaut {meter.installed_to.strftime('%d.%m.%Y')}"
                    elif (meter.installed_from
                          and period.start_date <= meter.installed_from <= period.end_date):
                        date_hint = f"eingebaut {meter.installed_from.strftime('%d.%m.%Y')}"
                    else:
                        date_hint = "ganzer Zeitraum"
                    ded_suffix = (
                        f", abzügl. {deducted.quantize(Decimal('1'))} m³ lt. Schlussrechnung"
                        if deducted > 0 else "")
                    desc = (
                        f"Wasserverbrauch {period.name} – Zähler {meter.meter_number}"
                        f" ({date_hint}, {consumption} m³{ded_suffix})"
                    )
                    amount = (consumption * tariff.price_per_m3).quantize(Decimal("0.01"))
                    db.session.add(InvoiceItem(
                        invoice_id=inv.id,
                        description=desc,
                        quantity=consumption,
                        unit="m³",
                        unit_price=tariff.price_per_m3,
                        amount=amount,
                        tax_rate=water_tax,
                        is_estimated=bool(getattr(reading, "is_estimated", False)),
                    ))
            else:
                # Eine Zeile mit Gesamtverbrauch (Standard)
                net_consumption = total_consumption
                if ded_total > 0:
                    net_consumption = total_consumption - ded_total
                    if net_consumption < 0:
                        clamp_warnings.append(prop.label())
                        net_consumption = Decimal("0")
                price_str = str(tariff.price_per_m3).replace(".", ",")
                desc = (
                    f"Wasserverbrauch {period.name}"
                    f" ({net_consumption.quantize(Decimal('1'))} m³"
                    f" × {price_str} €/m³)"
                )
                if ded_total > 0:
                    nums = ", ".join(ded_numbers)
                    ref = f" (Schlussrechnung {nums})" if nums else ""
                    desc += (f" – abzüglich {ded_total.quantize(Decimal('1'))} m³"
                             f" bereits verrechnet{ref}")
                amount = (net_consumption * tariff.price_per_m3).quantize(Decimal("0.01"))
                db.session.add(InvoiceItem(
                    invoice_id=inv.id,
                    description=desc,
                    quantity=net_consumption,
                    unit="m³",
                    unit_price=tariff.price_per_m3,
                    amount=amount,
                    tax_rate=water_tax,
                    is_estimated=is_any_estimated,
                ))

            # Gebuehren-Aufteilung bei pro-rata-Schlussrechnung: den bereits
            # dem Altbesitzer verrechneten Tage-Anteil dem Nachbesitzer kuerzen.
            fee_days = ded["fee_days"] if ded else 0
            period_days = (period.end_date - period.start_date).days + 1

            def _fee_line(label, fee):
                """(Beschreibung, Betrag) fuer eine Gebuehrenposition — anteilig
                gekuerzt, wenn eine pro-rata-Schlussrechnung Tage verrechnet hat."""
                if fee_days > 0 and period_days > 0:
                    remaining = max(period_days - fee_days, 0)
                    amt = (Decimal(str(fee)) * Decimal(remaining)
                           / Decimal(period_days)).quantize(Decimal("0.01"))
                    return f"{label} (anteilig {remaining}/{period_days} Tage)", amt
                return label, fee

            # Grundgebühr (nur wenn explizit hinterlegt, auch 0 erzeugt eine Position)
            if effective_base_fee is not None:
                base_desc, base_amount = _fee_line(base_fee_label, effective_base_fee)
                db.session.add(InvoiceItem(
                    invoice_id=inv.id,
                    description=base_desc,
                    quantity=1,
                    unit="Pauschal",
                    unit_price=base_amount,
                    amount=base_amount,
                    tax_rate=water_tax,
                ))

            # Zusatzgebühr (nur wenn explizit hinterlegt, auch 0 erzeugt eine Position)
            if effective_additional_fee is not None:
                add_desc, add_amount = _fee_line(additional_fee_label, effective_additional_fee)
                db.session.add(InvoiceItem(
                    invoice_id=inv.id,
                    description=add_desc,
                    quantity=1,
                    unit="Pauschal",
                    unit_price=add_amount,
                    amount=add_amount,
                    tax_rate=water_tax,
                ))

            # Damit USt (sofern vorhanden) im Gesamtbetrag berücksichtigt wird:
            db.session.flush()
            inv.recalculate_total()

            # Offene Schaetz-Korrekturen des Kunden in seine (erste) Rechnung
            # dieses Laufs einziehen — Nachforderung voll, Gutschrift nur bis
            # Betrag 0, Rest wandert auf die naechste Rechnung.
            if ownership.customer_id not in customers_corrected:
                apply_corrections_to_invoice(inv, ownership.customer_id)
                customers_corrected.add(ownership.customer_id)

            # Rechnung nie negativ (z.B. negativer Verbrauch nach zu hoher
            # Vorperioden-Schaetzung): auf 0 kappen, Rest als Gutschrift
            # auf die naechste Rechnung uebertragen.
            cap_invoice_at_zero(
                inv, customer_id=ownership.customer_id,
                meter_id=prop_readings[0].meter_id, period_id=period.id,
                tax_rate=water_tax, created_by_id=current_user.id,
            )

            created += 1

        billing_run.invoices_created = created
        billing_run.invoices_skipped = skipped

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Rechnungslauf – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("invoices.generate"))
        flash(f"{created} Rechnungen erstellt, {skipped} übersprungen.", "success")
        if clamp_warnings:
            objs = ", ".join(dict.fromkeys(clamp_warnings))
            flash(
                f"Bei folgenden Objekten übersteigt die Schlussrechnung den "
                f"Jahresverbrauch (Rest auf 0 gekappt) – bitte eine Gutschrift "
                f"an den Altbesitzer manuell prüfen: {objs}.", "warning")
        return redirect(url_for("invoices.billing_run_detail", run_id=billing_run.id))

    return render_template(
        "invoices/generate.html",
        tariffs=tariffs,
        periods=_billing_period_list(),
        active_period=BillingPeriod.current(),
        sort_order_choices=BillingRun.SORT_ORDER_CHOICES,
    )


@bp.route("/new", methods=["GET", "POST"])
@login_required
def new():
    """Direkte Rechnungs-Anlage (Entwurf, ohne vorherigen OpenItem).

    Nutzt denselben Positions-Editor wie der OpenItem→Rechnung-Flow und ist
    damit der primäre Weg zur manuellen Rechnungserstellung (siehe ADR-001).
    Kunde ist Pflichtfeld; das Quick-Create-Modal erlaubt die Neuanlage
    direkt aus diesem Formular heraus.
    """
    from app.accounting import services as acc_svc
    customers = Customer.query.order_by(Customer.name).all()
    tariffs = WaterTariff.query.order_by(WaterTariff.valid_from.desc()).all()
    editor_projects = Project.query.filter_by(closed=False).order_by(Project.name).all()

    if request.method == "POST":
        customer_id_raw = request.form.get("customer_id", "").strip()
        if not customer_id_raw:
            flash("Bitte einen Kunden auswählen.", "danger")
            return redirect(url_for("invoices.new"))
        try:
            customer_id = int(customer_id_raw)
        except ValueError:
            flash("Ungültige Kunden-ID.", "danger")
            return redirect(url_for("invoices.new"))
        customer = db.session.get(Customer, customer_id)
        if not customer:
            flash("Kunde nicht gefunden.", "danger")
            return redirect(url_for("invoices.new"))

        try:
            inv_date = date.fromisoformat(request.form["date"])
        except (KeyError, ValueError):
            flash("Ungültiges Rechnungsdatum.", "danger")
            return redirect(url_for("invoices.new"))

        fy_error = acc_svc.open_fiscal_year_error(inv_date)
        if fy_error:
            flash(f"{fy_error} Rechnung wurde nicht erstellt.", "danger")
            return redirect(url_for("invoices.new"))

        is_vat_liable_year = acc_svc.is_year_vat_liable(inv_date.year)
        due_date = (
            date.fromisoformat(request.form["due_date"])
            if request.form.get("due_date") else None
        )
        notes = request.form.get("notes", "").strip()

        inv = Invoice(
            invoice_number=_next_invoice_number(inv_date.year),
            customer_id=customer.id,
            date=inv_date,
            due_date=due_date,
            status=Invoice.STATUS_DRAFT,
            notes=notes,
            created_by_id=current_user.id,
        )
        db.session.add(inv)
        db.session.flush()

        _apply_row_items_to_invoice(inv, request.form, is_vat_liable_year)
        inv.recalculate_total()

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(
                f"Fehler beim Erstellen der Rechnung – alle Änderungen wurden zurückgesetzt: {e}",
                "danger",
            )
            return redirect(url_for("invoices.new"))

        flash(f"Rechnung {inv.invoice_number} erstellt.", "success")
        return redirect(url_for("invoices.detail", invoice_id=inv.id))

    return render_template(
        "invoices/new.html",
        customers=customers,
        tariffs=tariffs,
        today=date.today(),
        editor_projects=editor_projects,
    )


def _apply_row_items_to_invoice(inv, form, is_vat_liable_year):
    """Fügt die vom Positions-Editor gesendeten Zeilen als ``InvoiceItem``
    zur Rechnung hinzu. Shared zwischen ``invoices.new`` und
    ``accounting.open_item_invoice`` (dort noch inline), damit die
    Parse-Logik nur einmal existiert.
    """
    row_types = form.getlist("row_type[]")
    row_tariff_ids = form.getlist("row_tariff_id[]")
    row_consumptions = form.getlist("row_consumption_m3[]")
    row_descriptions = form.getlist("row_description[]")
    row_quantities = form.getlist("row_quantity[]")
    row_units = form.getlist("row_unit[]")
    row_unit_prices = form.getlist("row_unit_price[]")
    row_tax_rates = form.getlist("row_tax_rate[]")
    row_project_ids = form.getlist("row_project_id[]")

    def _dec(lst, idx, default="0"):
        v = lst[idx].replace(",", ".") if idx < len(lst) and lst[idx].strip() else default
        try:
            return Decimal(v)
        except Exception:
            return Decimal(default)

    def _int_or_none(lst, idx):
        if idx >= len(lst):
            return None
        raw = (lst[idx] or "").strip()
        if not raw:
            return None
        try:
            return int(raw)
        except ValueError:
            return None

    water_tax = Decimal("10") if is_vat_liable_year else None

    for i, rtype in enumerate(row_types):
        row_project_id = _int_or_none(row_project_ids, i)

        if rtype == "tariff":
            tariff_id_raw = row_tariff_ids[i] if i < len(row_tariff_ids) else ""
            if not tariff_id_raw:
                continue
            try:
                tariff_id = int(tariff_id_raw)
            except ValueError:
                continue
            tariff = db.session.get(WaterTariff, tariff_id)
            if not tariff:
                continue
            consumption = _dec(row_consumptions, i)
            price_str = str(tariff.price_per_m3).replace(".", ",")
            amount = (consumption * tariff.price_per_m3).quantize(Decimal("0.01"))
            db.session.add(InvoiceItem(
                invoice_id=inv.id,
                description=f"Wasserverbrauch ({consumption.quantize(Decimal('1'))} m³ × {price_str} €/m³)",
                quantity=consumption,
                unit="m³",
                unit_price=tariff.price_per_m3,
                amount=amount,
                tax_rate=water_tax,
                project_id=row_project_id,
            ))
            if tariff.base_fee is not None:
                db.session.add(InvoiceItem(
                    invoice_id=inv.id,
                    description=tariff.base_fee_label or "Grundgebühr",
                    quantity=Decimal("1"),
                    unit="Pauschal",
                    unit_price=tariff.base_fee,
                    amount=tariff.base_fee,
                    tax_rate=water_tax,
                    project_id=row_project_id,
                ))
            if tariff.additional_fee is not None:
                db.session.add(InvoiceItem(
                    invoice_id=inv.id,
                    description=tariff.additional_fee_label or "Zusatzgebühr",
                    quantity=Decimal("1"),
                    unit="Pauschal",
                    unit_price=tariff.additional_fee,
                    amount=tariff.additional_fee,
                    tax_rate=water_tax,
                    project_id=row_project_id,
                ))
        elif rtype == "water":
            consumption = _dec(row_consumptions, i)
            unit_price = _dec(row_unit_prices, i)
            desc = (
                row_descriptions[i].strip()
                if i < len(row_descriptions) and row_descriptions[i].strip()
                else f"Wasserverbrauch ({consumption.quantize(Decimal("1"))} m³)"
            )
            amount = (consumption * unit_price).quantize(Decimal("0.01"))
            db.session.add(InvoiceItem(
                invoice_id=inv.id,
                description=desc,
                quantity=consumption,
                unit="m³",
                unit_price=unit_price,
                amount=amount,
                tax_rate=water_tax,
                project_id=row_project_id,
            ))
        else:  # free
            desc = row_descriptions[i].strip() if i < len(row_descriptions) else ""
            if not desc:
                continue
            qty = _dec(row_quantities, i, "1")
            unit = row_units[i] if i < len(row_units) and row_units[i].strip() else "Stk"
            unit_price = _dec(row_unit_prices, i)
            tax_rate = _dec(row_tax_rates, i) if is_vat_liable_year else Decimal("0")
            amount = (qty * unit_price).quantize(Decimal("0.01"))
            db.session.add(InvoiceItem(
                invoice_id=inv.id,
                description=desc,
                quantity=qty,
                unit=unit,
                unit_price=unit_price,
                amount=amount,
                tax_rate=tax_rate if tax_rate > 0 else None,
                project_id=row_project_id,
            ))


@bp.route("/<int:invoice_id>")
@login_required
def detail(invoice_id):
    from app.accounting import services as acc_svc
    from app import tax_service
    invoice = db.get_or_404(Invoice, invoice_id)
    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    doc_format = AppSetting.get("invoice.document_format", "pdf")
    fy_vat_liable = acc_svc.is_year_vat_liable(invoice.date.year) if invoice.date else False
    tax_rates = tax_service.tax_rates() if fy_vat_liable else []
    # Tarife und Projekte werden nur im Entwurfs-Modus für den Positions-Editor benötigt.
    if invoice.status == Invoice.STATUS_DRAFT:
        tariffs = WaterTariff.query.order_by(WaterTariff.valid_from.desc()).all()
        editor_projects = Project.query.filter_by(closed=False).order_by(Project.name).all()
    else:
        tariffs = []
        # Auch in Read-Only-Ansicht brauchen wir Projekte, damit bestehende
        # Items korrekt als hidden-Inputs gerendert werden (Backward-Compat).
        editor_projects = Project.query.order_by(Project.name).all()
    return render_template(
        "invoices/detail.html",
        invoice=invoice,
        accounts=accounts,
        doc_format=doc_format,
        fy_vat_liable=fy_vat_liable,
        tax_rates=tax_rates,
        tariffs=tariffs,
        editor_projects=editor_projects,
        tax_summary=invoice.tax_breakdown,
        invoice_gross_total=invoice.total_amount,
    )


@bp.route("/<int:invoice_id>/edit", methods=["GET", "POST"])
@login_required
def edit(invoice_id):
    invoice = db.get_or_404(Invoice, invoice_id)
    if _invoice_is_locked(invoice):
        flash("Diese Rechnung kann nicht mehr bearbeitet werden (nur Stornierung möglich).", "warning")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))
    customers = Customer.query.filter_by(active=True).order_by(Customer.name).all()
    if request.method == "POST":
        invoice.date = date.fromisoformat(request.form["date"])
        invoice.due_date = date.fromisoformat(request.form["due_date"]) if request.form.get("due_date") else None
        invoice.notes = request.form.get("notes", "")
        db.session.commit()
        flash("Rechnung gespeichert.", "success")
        if invoice.date > date.today():
            flash("Achtung: Das Rechnungsdatum liegt in der Zukunft.", "warning")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))
    return render_template("invoices/edit.html", invoice=invoice, customers=customers)


@bp.route("/<int:invoice_id>/items/save", methods=["POST"])
@login_required
def items_save(invoice_id):
    """Speichert alle Rechnungspositionen eines Rechnungsentwurfs in einem
    einzigen Schritt. Ersetzt die frühere Kombination aus item_add/item_edit/
    item_delete und nutzt denselben Positions-Editor wie ``/invoices/new``.

    Nuclear-Replace: alle bisherigen Items werden gelöscht und aus den
    eingereichten Zeilen neu aufgebaut. Der zugeordnete Offene Posten wird
    anschließend auf den neuen Gesamtbetrag synchronisiert.
    """
    from app.accounting import services as acc_svc
    invoice = db.get_or_404(Invoice, invoice_id)
    if _invoice_is_locked(invoice):
        flash("Rechnung ist gesperrt und kann nicht mehr bearbeitet werden.", "warning")
        return redirect(url_for("invoices.detail", invoice_id=invoice_id))

    is_vat_liable_year = acc_svc.is_year_vat_liable(
        invoice.date.year if invoice.date else date.today().year
    )

    # Alle bestehenden Positionen entfernen — der Editor liefert den kompletten
    # neuen Zustand inklusive der (als "free" gerenderten) bestehenden Items.
    # ADR-003: Mahngebühr-Items bleiben erhalten (werden nur vom Dunning-Service verwaltet).
    for old_item in list(invoice.items):
        if getattr(old_item, "is_dunning_fee", 0):
            continue
        db.session.delete(old_item)
    db.session.flush()

    _apply_row_items_to_invoice(invoice, request.form, is_vat_liable_year)
    # Die neuen Items werden via invoice_id-FK (nicht ueber die Relationship-
    # Collection) angelegt. Da invoice.items oben bereits geladen wurde, ist die
    # In-Memory-Collection veraltet — flush + expire erzwingt ein frisches Reload,
    # damit recalculate_total die tatsaechlichen Positionen summiert.
    db.session.flush()
    db.session.expire(invoice, ["items"])
    invoice.recalculate_total()
    sync_message = _sync_open_item(invoice)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Speichern – alle Änderungen wurden zurückgesetzt: {e}", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice_id))

    flash("Rechnungspositionen gespeichert.", "success")
    if sync_message:
        flash("Offener Posten wurde auf den neuen Betrag aktualisiert.", "info")
    return redirect(url_for("invoices.detail", invoice_id=invoice_id))


@bp.route("/bulk-action", methods=["POST"])
@login_required
def bulk_action():
    invoice_ids = request.form.getlist("invoice_ids", type=int)
    action = request.form.get("action", "")

    if not invoice_ids:
        flash("Keine Rechnungen ausgewählt.", "warning")
        return redirect(url_for("invoices.index"))

    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()

    if action == "delete":
        deletable = [inv for inv in invoices if inv.status == Invoice.STATUS_DRAFT]
        skipped = len(invoices) - len(deletable)
        try:
            for inv in deletable:
                if inv.open_item:
                    db.session.delete(inv.open_item)
                db.session.delete(inv)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Löschen – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("invoices.index"))
        msg = f"{len(deletable)} Rechnung(en) gelöscht."
        if skipped:
            msg += f" {skipped} übersprungen (nur Entwürfe können gelöscht werden)."
        flash(msg, "success" if deletable else "warning")

    elif action in Invoice.ALL_STATUSES:
        if action == Invoice.STATUS_DRAFT:
            flash("Rechnungen können nicht auf 'Entwurf' zurückgesetzt werden.", "danger")
            return redirect(url_for("invoices.index"))
        bulk_account_id_raw = request.form.get("account_id") or None
        bulk_account_id = int(bulk_account_id_raw) if bulk_account_id_raw else None
        changed = 0
        try:
            for inv in invoices:
                inv.status = action
                if action == Invoice.STATUS_SENT:
                    account_id = _resolve_open_item_account_id(inv, bulk_account_id)
                    _create_or_update_open_item(inv, account_id=account_id)
                elif action == Invoice.STATUS_CANCELLED:
                    if inv.open_item:
                        inv.open_item.status = OpenItem.STATUS_PAID
                    # ADR-003: Storno → alle aktiven Mahnungen der Rechnung stornieren
                    from app.dunning.services import cancel_dunnings_for_invoice
                    cancel_dunnings_for_invoice(inv)
                changed += 1
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Statuswechsel – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("invoices.index"))
        flash(f"{changed} Rechnung(en) auf '{action}' gesetzt.", "success")

    elif action == "set-date":
        new_date_raw = request.form.get("new_date", "")
        try:
            new_date = date.fromisoformat(new_date_raw)
        except (ValueError, TypeError):
            flash("Ungültiges Datum.", "danger")
            return redirect(url_for("invoices.index"))
        # Nur Entwürfe dürfen ihr Rechnungsdatum ändern (analog Einzel-Edit).
        editable = [inv for inv in invoices if inv.status == Invoice.STATUS_DRAFT]
        skipped = len(invoices) - len(editable)
        try:
            for inv in editable:
                inv.date = new_date
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Setzen des Datums – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("invoices.index"))
        msg = f"Rechnungsdatum bei {len(editable)} Rechnung(en) auf {new_date.strftime('%d.%m.%Y')} gesetzt."
        if skipped:
            msg += f" {skipped} übersprungen (nur Entwürfe können geändert werden)."
        flash(msg, "success" if editable else "warning")
        if editable and new_date > date.today():
            flash("Achtung: Das gesetzte Rechnungsdatum liegt in der Zukunft.", "warning")

    elif action == "set-due-date":
        new_due_raw = request.form.get("new_due_date", "")
        try:
            new_due = date.fromisoformat(new_due_raw)
        except (ValueError, TypeError):
            flash("Ungültiges Datum.", "danger")
            return redirect(url_for("invoices.index"))
        # Nur Entwürfe dürfen ihr Fälligkeitsdatum ändern (analog Einzel-Edit).
        editable = [inv for inv in invoices if inv.status == Invoice.STATUS_DRAFT]
        skipped = len(invoices) - len(editable)
        try:
            for inv in editable:
                inv.due_date = new_due
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Setzen des Fälligkeitsdatums – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("invoices.index"))
        msg = f"Fälligkeitsdatum bei {len(editable)} Rechnung(en) auf {new_due.strftime('%d.%m.%Y')} gesetzt."
        if skipped:
            msg += f" {skipped} übersprungen (nur Entwürfe können geändert werden)."
        flash(msg, "success" if editable else "warning")

    else:
        flash("Ungültige Aktion.", "danger")

    return redirect(url_for("invoices.index"))


def _bulk_print_limit_exceeded(ids):
    """Sicherheitsnetz fuer Massendruck/-export: gibt eine Redirect-Response
    zurueck, wenn die Auswahl das konfigurierte Limit (BULK_PRINT_MAX) ueber-
    schreitet, sonst None. Die UI batcht bereits in Gruppen, der Cap faengt
    direkte/veraltete Clients ab und schuetzt vor RAM-/Timeout-Last."""
    limit = current_app.config.get("BULK_PRINT_MAX", 100)
    if len(ids) > limit:
        flash(f"Es können maximal {limit} Dokumente pro Durchgang erstellt werden. "
              f"Bitte die Auswahl in kleinere Gruppen aufteilen.", "warning")
        return redirect(url_for("invoices.index"))
    return None


@bp.route("/bulk-pdf-merged", methods=["POST"])
@login_required
def bulk_pdf_merged():
    """Alle markierten Rechnungen als zusammengeführtes PDF zum Download."""
    import io
    invoice_ids = request.form.getlist("invoice_ids", type=int)
    if not invoice_ids:
        flash("Keine Rechnungen ausgewählt.", "warning")
        return redirect(url_for("invoices.index"))
    if (resp := _bulk_print_limit_exceeded(invoice_ids)):
        return resp
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        if current_app.debug:
            flash("WeasyPrint nicht verfügbar. Einzelne PDF-Vorschau unter /invoices/<id>/pdf-preview.", "info")
        else:
            flash("WeasyPrint ist nicht installiert. PDF-Export nur im Docker-Container verfügbar.", "danger")
        return redirect(url_for("invoices.index"))
    from pypdf import PdfWriter
    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).order_by(Invoice.invoice_number).all()
    if not invoices:
        flash("Keine Rechnungen gefunden.", "warning")
        return redirect(url_for("invoices.index"))
    # Jede Rechnung einzeln rendern und als fertiges PDF an den Merger haengen.
    # WICHTIG — bewusst KEIN WeasyPrint-copy() ueber mehrere Dokumente:
    # dessen Bild-Deduplizierung (_use_references) kippt auf dem Server
    # reproduzierbar mit PIL.UnidentifiedImageError, sobald das geteilte Logo
    # ueber viele Seiten referenziert wird. Schlimmer noch: der copy()-Versuch
    # konsumiert die Bild-Buffer der beteiligten Dokumente, sodass danach auch
    # ein Einzel-write_pdf desselben Dokuments scheitert (kein sauberer Fallback
    # moeglich). Einzel-Render + pypdf-Concat laeuft dagegen stabil durch.
    writer = PdfWriter()
    for invoice in invoices:
        html_content = _render_pdf_html(invoice)
        pdf_bytes = HTML(string=html_content).render().write_pdf()
        if _invoice_is_locked(invoice):
            pdf_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "pdf")
            if not invoice.pdf_path or not os.path.exists(invoice.pdf_path):
                with open(pdf_path, "wb") as fh:
                    fh.write(pdf_bytes)
                invoice.pdf_path = pdf_path
        writer.append(io.BytesIO(pdf_bytes))
    db.session.commit()
    # Byte-identische Objekte zusammenfassen (hilft bei deckenden Logos; bei
    # transparenten PNGs mit SMask greift es kaum — bekannte Einschraenkung).
    writer.compress_identical_objects()
    merged_path = os.path.join(current_app.config["PDF_DIR"], "_bulk_merged.pdf")
    os.makedirs(current_app.config["PDF_DIR"], exist_ok=True)
    with open(merged_path, "wb") as fh:
        writer.write(fh)
    writer.close()
    return send_file(merged_path, as_attachment=True, download_name="Rechnungen_gesamt.pdf")


@bp.route("/bulk-pdf-zip", methods=["POST"])
@login_required
def bulk_pdf_zip():
    """Alle markierten Rechnungen als einzelne PDFs in einer ZIP-Datei."""
    import zipfile
    import io
    invoice_ids = request.form.getlist("invoice_ids", type=int)
    if not invoice_ids:
        flash("Keine Rechnungen ausgewählt.", "warning")
        return redirect(url_for("invoices.index"))
    if (resp := _bulk_print_limit_exceeded(invoice_ids)):
        return resp
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        if current_app.debug:
            flash("WeasyPrint nicht verfügbar. Einzelne PDF-Vorschau unter /invoices/<id>/pdf-preview.", "info")
        else:
            flash("WeasyPrint ist nicht installiert. PDF-Export nur im Docker-Container verfügbar.", "danger")
        return redirect(url_for("invoices.index"))
    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).order_by(Invoice.invoice_number).all()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for invoice in invoices:
            if _invoice_is_locked(invoice) and invoice.pdf_path and os.path.exists(invoice.pdf_path):
                zf.write(invoice.pdf_path, f"{invoice.invoice_number}.pdf")
            else:
                pdf_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "pdf")
                html_content = _render_pdf_html(invoice)
                HTML(string=html_content).write_pdf(pdf_path)
                if _invoice_is_locked(invoice):
                    invoice.pdf_path = pdf_path
                zf.write(pdf_path, f"{invoice.invoice_number}.pdf")
    db.session.commit()
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Rechnungen.zip",
                     mimetype="application/zip")


@bp.route("/bulk-docx-zip", methods=["POST"])
@login_required
def bulk_docx_zip():
    """Alle markierten Rechnungen als einzelne .docx-Dateien in einer ZIP-Datei."""
    import zipfile
    import io
    from app.invoices.document_service import generate_docx
    invoice_ids = request.form.getlist("invoice_ids", type=int)
    if not invoice_ids:
        flash("Keine Rechnungen ausgewählt.", "warning")
        return redirect(url_for("invoices.index"))
    if (resp := _bulk_print_limit_exceeded(invoice_ids)):
        return resp
    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).order_by(Invoice.invoice_number).all()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for invoice in invoices:
            if _invoice_is_locked(invoice) and invoice.doc_path and os.path.exists(invoice.doc_path):
                zf.write(invoice.doc_path, f"{invoice.invoice_number}.docx")
            else:
                doc_data = generate_docx(invoice, wg_settings(), design=_current_design())
                if _invoice_is_locked(invoice):
                    doc_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "docx")
                    with open(doc_path, "wb") as f:
                        f.write(doc_data)
                    invoice.doc_path = doc_path
                zf.writestr(f"{invoice.invoice_number}.docx", doc_data)
    db.session.commit()
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Rechnungen_docx.zip",
                     mimetype="application/zip")


@bp.route("/bulk-docx-merged", methods=["POST"])
@login_required
def bulk_docx_merged():
    """Alle markierten Rechnungen als einzelnes zusammengeführtes .docx zum Download."""
    import io as _io
    from app.invoices.document_service import generate_docx, merge_docx_files
    invoice_ids = request.form.getlist("invoice_ids", type=int)
    if not invoice_ids:
        flash("Keine Rechnungen ausgewählt.", "warning")
        return redirect(url_for("invoices.index"))
    if (resp := _bulk_print_limit_exceeded(invoice_ids)):
        return resp
    invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).order_by(Invoice.invoice_number).all()
    sources = []
    for invoice in invoices:
        if _invoice_is_locked(invoice) and invoice.doc_path and os.path.exists(invoice.doc_path):
            sources.append(invoice.doc_path)
        else:
            doc_data = generate_docx(invoice, wg_settings(), design=_current_design())
            if _invoice_is_locked(invoice):
                doc_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "docx")
                with open(doc_path, "wb") as f:
                    f.write(doc_data)
                invoice.doc_path = doc_path
            sources.append(doc_data)
    db.session.commit()
    merged = merge_docx_files(sources)
    return send_file(_io.BytesIO(merged), as_attachment=True,
                     download_name="Rechnungen_gesamt.docx",
                     mimetype=_DOCX_MIME)


@bp.route("/<int:invoice_id>/status", methods=["POST"])
@login_required
def set_status(invoice_id):
    invoice = db.get_or_404(Invoice, invoice_id)
    new_status = request.form.get("status")
    if new_status not in Invoice.ALL_STATUSES:
        abort(400)

    old_status = invoice.status

    # Kein Wechsel → keine Aktion (verhindert versehentliches Neu-Verbuchen).
    if new_status == old_status:
        flash("Status unverändert.", "info")
        return _status_response(invoice)

    # State-Machine durchsetzen (siehe Invoice.ALLOWED_TRANSITIONS).
    if new_status not in Invoice.ALLOWED_TRANSITIONS.get(old_status, []):
        if old_status == Invoice.STATUS_DRAFT and new_status == Invoice.STATUS_CANCELLED:
            msg = ("Ein Entwurf wird nicht storniert, sondern gelöscht "
                   "(Schaltfläche „Löschen“). Stornieren ist nur für bereits "
                   "versendete Rechnungen vorgesehen.")
        elif old_status == Invoice.STATUS_DRAFT:
            msg = ("Eine Entwurfs-Rechnung muss zuerst auf „Versendet“ gesetzt "
                   "werden, bevor sie als bezahlt oder als Guthaben markiert "
                   "werden kann.")
        elif old_status == Invoice.STATUS_CANCELLED:
            msg = "Eine stornierte Rechnung kann nicht wieder aktiviert werden."
        elif new_status == Invoice.STATUS_DRAFT:
            msg = "Eine Rechnung kann nicht auf „Entwurf“ zurückgesetzt werden."
        else:
            msg = (f"Statuswechsel von „{old_status}“ auf „{new_status}“ "
                   f"ist nicht zulässig.")
        flash(msg, "danger")
        return _status_response(invoice)

    try:
        if new_status == Invoice.STATUS_SENT:
            form_account_id_raw = request.form.get("account_id") or None
            form_account_id = int(form_account_id_raw) if form_account_id_raw else None
            account_id = _resolve_open_item_account_id(invoice, form_account_id)
            invoice.status = new_status
            _create_or_update_open_item(invoice, account_id=account_id)

        elif new_status == Invoice.STATUS_CANCELLED:
            # Verknüpfte Buchungen rückabwickeln, sonst bleibt eine zuvor
            # erzeugte Zahlungsbuchung als Phantom-Einnahme in den Auswertungen.
            from app.accounting import services as acc_svc
            err = acc_svc.storno_invoice_bookings(
                invoice,
                reason=f"Storno Rechnung {invoice.invoice_number}",
                created_by_id=current_user.id,
            )
            if err:
                raise ValueError(err)
            invoice.status = new_status
            if invoice.open_item:
                invoice.open_item.status = OpenItem.STATUS_PAID
            # ADR-003: Storno → alle aktiven Mahnungen der Rechnung stornieren
            from app.dunning.services import cancel_dunnings_for_invoice
            cancel_dunnings_for_invoice(invoice)

        elif new_status == Invoice.STATUS_PAID:
            invoice.status = new_status
            _book_invoice_payment_if_needed(invoice)

        else:  # STATUS_CREDIT
            invoice.status = new_status

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Statuswechsel – alle Änderungen wurden zurückgesetzt: {e}", "danger")
        return _status_response(invoice)

    flash(f"Status auf '{new_status}' gesetzt.", "success")
    return _status_response(invoice)


@bp.route("/<int:invoice_id>/delete", methods=["POST"])
@login_required
def delete(invoice_id):
    """Löscht eine Entwurfs-Rechnung endgültig.

    Nur im Entwurfsstatus erlaubt — eine versendete Rechnung wurde ausgestellt
    und darf nur noch storniert (nicht gelöscht) werden. Da der Entwurf nie
    ausgestellt war, ist die endgültige Löschung der saubere Weg statt einer
    Pseudo-Stornierung. Optional wird der Rechnungsnummern-Zähler des Jahres
    zurückgesetzt (analog zum Rechnungslauf-Löschen).
    """
    invoice = db.get_or_404(Invoice, invoice_id)
    if invoice.status != Invoice.STATUS_DRAFT:
        flash("Nur Entwürfe können gelöscht werden. Eine bereits ausgestellte "
              "Rechnung muss storniert werden.", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    number = invoice.invoice_number
    reset_counter = request.form.get("reset_counter") == "1"
    try:
        year = int(number.split("-")[0])
    except (ValueError, IndexError):
        year = None

    try:
        # Schätz-Korrekturen rückabwickeln (verrechnete zurückgeben, aus der
        # Kappung erzeugte entfernen) BEVOR die Rechnung verschwindet.
        reverse_corrections_for_invoice(invoice)
        # Verweis eines Eigentuemerwechsels auf diese (Schluss-)Rechnung lösen —
        # SQLite feuert das FK-``SET NULL`` nicht; ohne das bliebe ein toter
        # Verweis, den der Deduktions-Join (Status-Filter) zwar ignoriert, aber
        # der die Detailanzeige verfälscht.
        from app.models import OwnerChange
        OwnerChange.query.filter_by(settlement_invoice_id=invoice.id).update(
            {"settlement_invoice_id": None, "fee_days_billed": None},
            synchronize_session=False)
        # Verknüpften Offenen Posten entfernen (bei Entwürfen normalerweise keiner).
        if invoice.open_item is not None:
            db.session.delete(invoice.open_item)
        db.session.delete(invoice)  # InvoiceItems via cascade="all, delete-orphan"
        db.session.flush()

        if reset_counter and year is not None:
            prefix = f"{year}-"
            last = (
                Invoice.query
                .filter(Invoice.invoice_number.like(f"{prefix}%"))
                .order_by(Invoice.invoice_number.desc())
                .first()
            )
            if last:
                try:
                    new_seq = int(last.invoice_number.split("-")[-1]) + 1
                except ValueError:
                    new_seq = 1
            else:
                new_seq = 1
            counter = db.session.get(InvoiceCounter, year)
            if counter is None:
                db.session.add(InvoiceCounter(year=year, next_seq=new_seq))
            else:
                counter.next_seq = new_seq

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Löschen – nichts wurde verändert: {e}", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    flash(f"Entwurf {number} wurde gelöscht.", "success")
    return redirect(url_for("invoices.index"))


@bp.route("/<int:invoice_id>/pay", methods=["POST"])
@login_required
def pay(invoice_id):
    """Zahlung (Teil- oder Vollzahlung) auf eine Rechnung buchen."""
    invoice = db.get_or_404(Invoice, invoice_id)
    amount_raw = request.form.get("amount", "0").replace(",", ".")
    try:
        amount = Decimal(amount_raw)
    except Exception:
        flash("Ungültiger Betrag.", "danger")
        return redirect(url_for("accounting.open_items"))
    if amount <= 0:
        flash("Betrag muss positiv sein.", "danger")
        return redirect(url_for("accounting.open_items"))

    # Offenes Buchungsjahr für das Zahlungsdatum prüfen
    from app.accounting import services as acc_svc
    payment_date = date.today()
    fy_error = acc_svc.open_fiscal_year_error(payment_date)
    if fy_error:
        flash(fy_error, "danger")
        return redirect(url_for("accounting.open_items"))

    # Standard-Bankkonto verwenden
    default_ra = RealAccount.query.filter_by(is_default=True, active=True).first() \
        or RealAccount.query.filter_by(active=True).first()
    real_account_id = default_ra.id if default_ra else None

    try:
        # Zahlung über Service-Layer erzeugen. Bei mehreren Dimensionen
        # (Konto/Projekt/Steuersatz) entsteht automatisch eine Sammelbuchung,
        # sonst eine einzelne Buchung (ADR-002).
        acc_svc.booking_group_from_invoice_payment(
            invoice=invoice,
            amount=amount,
            payment_date=payment_date,
            real_account_id=real_account_id,
            created_by_id=current_user.id,
            open_item=invoice.open_item,
            reference=invoice.invoice_number,
        )

        from sqlalchemy import func
        paid_total = db.session.query(func.sum(Booking.amount)).filter(
            Booking.invoice_id == invoice.id
        ).scalar() or Decimal("0")
        balance = Decimal(str(invoice.total_amount)) - Decimal(str(paid_total))

        if balance > Decimal("0"):
            invoice.status = Invoice.STATUS_SENT
        elif balance == Decimal("0"):
            invoice.status = Invoice.STATUS_PAID
        else:
            invoice.status = Invoice.STATUS_CREDIT

        if invoice.open_item:
            oi = invoice.open_item
            if balance > Decimal("0"):
                oi.status = OpenItem.STATUS_PARTIAL
            elif balance == Decimal("0"):
                oi.status = OpenItem.STATUS_PAID
            else:
                oi.status = OpenItem.STATUS_CREDIT

        db.session.commit()
    except ValueError as ve:
        db.session.rollback()
        flash(f"Fehler bei der Zahlung: {ve}", "danger")
        return redirect(url_for("accounting.open_items"))
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler bei der Zahlung – alle Änderungen wurden zurückgesetzt: {e}", "danger")
        return redirect(url_for("accounting.open_items"))

    if balance > Decimal("0"):
        flash(f"Teilzahlung von {amount:.2f} \u20ac gebucht. Offener Restbetrag: {balance:.2f} \u20ac", "success")
    elif balance == Decimal("0"):
        flash(f"Rechnung {invoice.invoice_number} vollst\u00e4ndig bezahlt.", "success")
    else:
        flash(f"\u00dcberzahlung von {abs(balance):.2f} \u20ac. Rechnung als Gutschrift markiert.", "info")
    return redirect(url_for("accounting.open_items"))


@bp.route("/<int:invoice_id>/pdf")
@login_required
def pdf(invoice_id):
    import io as _io
    invoice = db.get_or_404(Invoice, invoice_id)
    fmt = _get_document_format(request.args.get("fmt"))

    if fmt == "docx":
        # Gecachte .docx ausliefern wenn vorhanden
        if _invoice_is_locked(invoice) and invoice.doc_path and os.path.exists(invoice.doc_path):
            return send_file(invoice.doc_path, as_attachment=True,
                             download_name=f"{invoice.invoice_number}.docx",
                             mimetype=_DOCX_MIME)
        from app.invoices.document_service import generate_docx
        doc_data = generate_docx(invoice, wg_settings(), design=_current_design())
        if _invoice_is_locked(invoice):
            doc_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "docx")
            with open(doc_path, "wb") as f:
                f.write(doc_data)
            invoice.doc_path = doc_path
            db.session.commit()
        return send_file(_io.BytesIO(doc_data), as_attachment=True,
                         download_name=f"{invoice.invoice_number}.docx",
                         mimetype=_DOCX_MIME)

    # ── PDF (Standard) ────────────────────────────────────────────────────
    # Gesperrte Rechnungen: gecachte PDF ausliefern wenn vorhanden
    if _invoice_is_locked(invoice) and invoice.pdf_path and os.path.exists(invoice.pdf_path):
        return send_file(invoice.pdf_path, as_attachment=False,
                         download_name=f"{invoice.invoice_number}.pdf")
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        if current_app.debug:
            flash("WeasyPrint nicht verfügbar – HTML-Vorschau geöffnet (Strg+P → Als PDF speichern).", "info")
            return redirect(url_for("invoices.pdf_preview", invoice_id=invoice.id))
        flash("WeasyPrint ist nicht installiert. PDF-Export nur im Docker-Container verfügbar.", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))
    html_content = _render_pdf_html(invoice)
    pdf_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "pdf")
    HTML(string=html_content).write_pdf(pdf_path)
    # Nur für Nicht-Entwürfe persistieren
    if _invoice_is_locked(invoice):
        invoice.pdf_path = pdf_path
        db.session.commit()
    return send_file(pdf_path, as_attachment=False,
                     download_name=f"{invoice.invoice_number}.pdf")


@bp.route("/<int:invoice_id>/pdf-preview")
@login_required
def pdf_preview(invoice_id):
    """HTML-Vorschau des PDF-Templates im Browser (nur Entwicklung).

    Ruft denselben Render-Pfad wie WeasyPrint auf, gibt das HTML aber direkt
    aus. Im Browser: Strg+P → Als PDF speichern simuliert den echten Output.
    """
    invoice = db.get_or_404(Invoice, invoice_id)
    return _render_pdf_html(invoice)


def _record_invoice_sent(invoice, msg, recipient):
    """Setzt Versand-Status + Tracking-Felder am Invoice und legt ein Sent-Event an.

    Wird sowohl von der klassischen send_email-Route als auch von send_email_ajax
    nach erfolgreichem ``send_mail(msg)`` aufgerufen. ``read_message_id`` liefert
    die Postmark-MessageID, sofern der SaaS-Hook sie vorbelegt hat — bei
    reinem SMTP bleibt sie None.
    """
    invoice.status = Invoice.STATUS_SENT
    record_email_sent(invoice, recipient, read_message_id(msg))


@bp.route("/<int:invoice_id>/send-email", methods=["POST"])
@login_required
def send_email(invoice_id):
    import io as _io
    from flask_mail import Message

    invoice = db.get_or_404(Invoice, invoice_id)
    if not invoice.customer.email:
        flash("Keine E-Mail-Adresse beim Kunden hinterlegt.", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    test_mode = request.form.get("test_mode") == "1"
    # Einwilligung pruefen: ohne aktivierten Schalter „Schriftverkehr per
    # E-Mail" darf nur der Test an die eigene Admin-Adresse gehen.
    if not test_mode and not invoice.customer.rechnung_per_email:
        flash("Der Kunde hat den Schriftverkehr per E-Mail nicht aktiviert.", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))
    if test_mode:
        recipient = current_user.email
        if not recipient:
            flash("Kein eigener E-Mail-Account für Testmodus hinterlegt.", "danger")
            return redirect(url_for("invoices.detail", invoice_id=invoice.id))
    else:
        recipient = invoice.customer.email

    # Sperrliste: an eine als unzustellbar gesperrte Adresse gar nicht erst
    # senden (der Testversand an die Admin-Adresse bleibt erlaubt).
    if not test_mode:
        from app.email_suppression import suppression_notice
        notice = suppression_notice(recipient)
        if notice:
            flash(notice, "danger")
            return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    fmt = _get_document_format()

    subject = _render_email_subject(invoice)
    body = _render_email_body(invoice)
    if test_mode:
        subject = f"[TEST – an: {invoice.customer.email}] {subject}"
        body = f"[TESTMODUS – eigentlicher Empfänger: {invoice.customer.email}]\n\n{body}"

    msg = Message(subject=subject, recipients=[recipient], body=body)

    doc_data = None
    pdf_path = None
    pdf_ok = False

    if fmt in ("pdf", "both"):
        try:
            import weasyprint
            html_content = _render_pdf_html(invoice, for_email=True)
            pdf_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "pdf")
            weasyprint.HTML(string=html_content).write_pdf(pdf_path)
            with open(pdf_path, "rb") as fp:
                msg.attach(f"{invoice.invoice_number}.pdf", "application/pdf", fp.read())
            pdf_ok = True
        except (ImportError, OSError):
            pdf_path = None
            if fmt == "pdf":
                flash("WeasyPrint ist nicht installiert. E-Mail-Versand mit PDF-Anhang "
                      "nur im Docker-Container verfügbar.", "danger")
                return redirect(url_for("invoices.detail", invoice_id=invoice_id))
            # bei 'both': docx-Fallback unten

    if fmt == "docx" or (fmt == "both" and not pdf_ok):
        from app.invoices.document_service import generate_docx
        doc_data = generate_docx(invoice, wg_settings(), design=_current_design())
        msg.attach(f"{invoice.invoice_number}.docx", _DOCX_MIME, doc_data)

    if not msg.attachments:
        flash("Kein Dokument konnte generiert werden.", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    run_before_send(invoice, msg)
    try:
        send_mail(msg)
    except Exception as exc:
        current_app.logger.exception("E-Mail-Versand fehlgeschlagen")
        flash(f"E-Mail-Versand fehlgeschlagen: {exc}", "danger")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    if test_mode:
        flash(f"Test-Mail an {recipient} versendet (eigentlicher Empfänger: {invoice.customer.email}).", "info")
        return redirect(url_for("invoices.detail", invoice_id=invoice.id))

    _record_invoice_sent(invoice, msg, recipient=recipient)

    if doc_data and _invoice_is_locked(invoice):
        doc_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "docx")
        with open(doc_path, "wb") as f:
            f.write(doc_data)
        invoice.doc_path = doc_path
    if pdf_path:
        invoice.pdf_path = pdf_path

    form_account_id_raw = request.form.get("account_id") or None
    form_account_id = int(form_account_id_raw) if form_account_id_raw else None
    account_id = _resolve_open_item_account_id(invoice, form_account_id)
    _create_or_update_open_item(invoice, account_id=account_id)
    db.session.commit()
    flash(f"Rechnung an {recipient} versendet.", "success")
    return redirect(url_for("invoices.detail", invoice_id=invoice.id))


@bp.route("/<int:invoice_id>/send-email-ajax", methods=["POST"])
@login_required
def send_email_ajax(invoice_id):
    """JSON-Variante für den Massenmail-Versand per JavaScript."""
    from flask import jsonify
    from flask_mail import Message

    test_mode = request.form.get("test_mode") == "1"

    invoice = db.get_or_404(Invoice, invoice_id)
    if not invoice.customer.email:
        return jsonify({"ok": False, "error": "Keine E-Mail-Adresse hinterlegt"}), 400
    if not invoice.customer.rechnung_per_email:
        return jsonify({"ok": False, "error": "E-Mail-Versand nicht aktiviert"}), 400

    if test_mode:
        recipient = current_user.email
        if not recipient:
            return jsonify({"ok": False, "error": "Kein Admin-E-Mail für Testmodus hinterlegt"}), 400
    else:
        recipient = invoice.customer.email

    # Sperrliste: gesperrte Adressen im Massenversand pro Zeile als Fehler
    # melden (das Bulk-JS zeigt die Meldung an), nicht senden.
    if not test_mode:
        from app.email_suppression import suppression_notice
        notice = suppression_notice(recipient)
        if notice:
            return jsonify({"ok": False, "error": notice}), 400

    fmt = _get_document_format()

    try:
        subject = _render_email_subject(invoice)
        if test_mode:
            subject = f"[TEST – an: {invoice.customer.email}] {subject}"
        body = _render_email_body(invoice)
        if test_mode:
            body = f"[TESTMODUS – eigentlicher Empfänger: {invoice.customer.email}]\n\n{body}"
        msg = Message(subject=subject, recipients=[recipient], body=body)

        doc_data = None
        pdf_path = None
        pdf_ok = False

        if fmt in ("pdf", "both"):
            try:
                import weasyprint
                html_content = _render_pdf_html(invoice, for_email=True)
                pdf_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "pdf")
                weasyprint.HTML(string=html_content).write_pdf(pdf_path)
                with open(pdf_path, "rb") as fp:
                    msg.attach(f"{invoice.invoice_number}.pdf", "application/pdf", fp.read())
                pdf_ok = True
            except (ImportError, OSError):
                pdf_path = None
                if fmt == "pdf":
                    preview_url = url_for("invoices.pdf_preview", invoice_id=invoice_id)
                    return jsonify({"ok": False, "error": "WeasyPrint nicht verfügbar",
                                    "preview_url": preview_url if current_app.debug else None}), 503
                # bei 'both': docx-Fallback unten

        if fmt == "docx" or (fmt == "both" and not pdf_ok):
            from app.invoices.document_service import generate_docx
            doc_data = generate_docx(invoice, wg_settings(), design=_current_design())
            msg.attach(f"{invoice.invoice_number}.docx", _DOCX_MIME, doc_data)

        if not msg.attachments:
            return jsonify({"ok": False, "error": "Kein Dokument konnte generiert werden"}), 500

        run_before_send(invoice, msg)
        send_mail(msg)

        if not test_mode:
            _record_invoice_sent(invoice, msg, recipient=recipient)
            if doc_data and _invoice_is_locked(invoice):
                doc_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "docx")
                with open(doc_path, "wb") as f:
                    f.write(doc_data)
                invoice.doc_path = doc_path
            if pdf_path:
                invoice.pdf_path = pdf_path
            ajax_account_id_raw = request.form.get("account_id") or None
            ajax_account_id = int(ajax_account_id_raw) if ajax_account_id_raw else None
            account_id = _resolve_open_item_account_id(invoice, ajax_account_id)
            _create_or_update_open_item(invoice, account_id=account_id)
            db.session.commit()

        return jsonify({"ok": True, "invoice_number": invoice.invoice_number,
                        "email": recipient, "test_mode": test_mode})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(exc)}), 500


@bp.route("/<int:invoice_id>/email-events")
@login_required
def email_events(invoice_id):
    """Read-only Audit-Trail aller E-Mail-Versand-/Webhook-Events zur Rechnung."""
    invoice = db.get_or_404(Invoice, invoice_id)
    # Tiebreaker auf id desc: Postmark-Timestamps (DeliveredAt/BouncedAt) sind
    # nur sekundengenau, unsere Sent-Events haben datetime.utcnow()-Mikrosekunden.
    # Bei Zustellung binnen weniger ms ist Postmark.DeliveredAt zwar konzeptionell
    # juenger, vergleicht aber als kleiner — ohne Tiebreaker wuerde Delivery
    # unter Sent rutschen.
    events = (EmailEvent.query
              .filter_by(subject_type="invoice", subject_id=invoice.id)
              .order_by(EmailEvent.occurred_at.desc(),
                        EmailEvent.id.desc())
              .all())
    return render_template("invoices/email_events.html", invoice=invoice, events=events)


@bp.route("/email-einstellungen", methods=["GET", "POST"])
@login_required
def email_settings():
    """E-Mail-Vorlage für den Rechnungsversand konfigurieren (Verwaltungs-Recht)."""
    import pathlib
    if request.method == "POST":
        AppSetting.set("email_subject_template", request.form.get("email_subject_template", "").strip())
        AppSetting.set("email_body_template", request.form.get("email_body_template", "").strip())
        db.session.commit()
        flash("E-Mail-Vorlage gespeichert.", "success")
        return redirect(url_for("invoices.email_settings"))
    current_subject = AppSetting.get("email_subject_template") or DEFAULT_INVOICE_EMAIL_SUBJECT
    current_body = AppSetting.get("email_body_template")
    if not current_body:
        tpl_path = pathlib.Path(current_app.root_path) / "templates" / "invoices" / "email_body.txt"
        current_body = tpl_path.read_text(encoding="utf-8") if tpl_path.exists() else ""
    return render_template(
        "invoices/email_settings.html",
        current_subject=current_subject,
        current_body=current_body,
    )


# ---------------------------------------------------------------------------
# Rechnungsläufe – Übersicht, Detail, Löschen
# ---------------------------------------------------------------------------

@bp.route("/billing-runs")
@login_required
def billing_runs():
    runs = BillingRun.query.order_by(BillingRun.created_at.desc()).all()
    return render_template("invoices/billing_runs.html", runs=runs)


def _billing_run_amounts(invoice, base_label, add_label):
    """Kennzahlen einer Lauf-Rechnung für die Detail-Tabelle (eine Items-Iteration).

    Klassifiziert die Positionen anhand der Lauf-Snapshot-Labels: ``unit == "m³"``
    ist der Wasserverbrauch, ``description`` == Grund-/Zusatzgebühr-Label die
    Pauschalen. ``ust`` = Brutto (gespeichert in ``total_amount``) − Netto;
    bei nicht USt-pflichtigen Läufen 0. Mahngebühr-Items werden ignoriert.
    """
    m3 = Decimal("0")
    water_net = Decimal("0")
    base_fee = Decimal("0")
    add_fee = Decimal("0")
    net = Decimal("0")
    for item in invoice.items:
        if getattr(item, "is_dunning_fee", 0):
            continue
        amount = Decimal(str(item.amount or 0))
        net += amount
        if item.unit == "m³":
            water_net += amount
            m3 += item.quantity or Decimal("0")
        elif item.description == base_label:
            base_fee += amount
        elif item.description == add_label:
            add_fee += amount
    gross = Decimal(str(invoice.total_amount or 0))
    return {
        "m3": m3 or None,
        "water_net": water_net,
        "base_fee": base_fee,
        "additional_fee": add_fee,
        "net": net,
        "ust": gross - net,
        "gross": gross,
    }


def _billing_run_overview(run):
    """Aggregiert die Kennzahlen eines Rechnungslaufs für die Detailseite und
    den Komplett-Ausdruck (Deckblatt).

    Liefert ein Dict mit der Rechnungsliste (`rows`/`invoices`), Versand- und
    Zahlungs-Zählern sowie den Summen. Stornierte Rechnungen sind gegenstandslos
    und fließen nicht in die Summen ein — es gilt: Gesamtsumme = Bezahlt + Offen.
    """
    from sqlalchemy.orm import selectinload, joinedload

    invoices = (
        run.invoices
        .options(
            selectinload(Invoice.items),
            joinedload(Invoice.customer),
            joinedload(Invoice.property),
        )
        .order_by(Invoice.invoice_number, Invoice.id)
        .all()
    )

    base_label = run.tariff_base_fee_label or "Grundgebühr"
    add_label = run.tariff_additional_fee_label or "Zusatzgebühr"

    rows = []
    count_draft = count_sent = count_mail = count_post = 0
    count_paid = count_open = 0
    other_status_counts = {}
    sum_total = Decimal("0")   # Σ aller gültigen (nicht stornierten) Rechnungen
    sum_paid = Decimal("0")    # davon bereits bezahlt
    sum_open = Decimal("0")    # davon noch offen (inkl. noch nicht versandte)
    run_has_vat = False
    mailable = []   # versandbereite Entwürfe per Mail (für den Versenden-Dialog)
    post_ids = []   # versandbereite Entwürfe per Post

    for inv in invoices:
        amt = _billing_run_amounts(inv, base_label, add_label)
        if amt["ust"] and amt["ust"] > 0:
            run_has_vat = True
        wants_email = inv.customer.wants_email
        rows.append({"inv": inv, "amt": amt, "wants_email": wants_email})

        gross = amt["gross"]
        status = inv.status

        # --- Versand-Fortschritt --------------------------------------------
        # „Schon versendet" zählt dauerhaft jede Rechnung, die den Entwurfs-
        # status verlassen hat (Versendet/Bezahlt/Guthaben) — sinkt also NICHT,
        # wenn eine versendete Rechnung später bezahlt wird. Stornierte zählen
        # weder als Entwurf noch als versendet.
        if status == Invoice.STATUS_DRAFT:
            count_draft += 1
            if wants_email:
                count_mail += 1
                mailable.append({
                    "id": inv.id,
                    "number": inv.invoice_number,
                    "name": inv.customer.name,
                    "email": inv.customer.email,
                })
            else:
                count_post += 1
                post_ids.append(inv.id)
        elif status != Invoice.STATUS_CANCELLED:
            count_sent += 1

        # --- Finanzen -------------------------------------------------------
        # Stornierte Rechnungen sind gegenstandslos und fließen nicht in die
        # Summen ein. Es gilt: Gesamtsumme = Bezahlt + Offen.
        if status != Invoice.STATUS_CANCELLED:
            sum_total += gross
            if status == Invoice.STATUS_PAID:
                count_paid += 1
                sum_paid += gross
            else:
                count_open += 1
                sum_open += gross

        # --- Weitere Status (Storniert/Guthaben) gesondert ausweisen --------
        if status in (Invoice.STATUS_CANCELLED, Invoice.STATUS_CREDIT):
            other_status_counts[status] = other_status_counts.get(status, 0) + 1

    count_total = len(invoices)
    count_live = count_draft + count_sent            # Nenner für Versand-Balken
    pct_sent = round(count_sent / count_live * 100) if count_live else 0
    pct_paid = round(float(sum_paid) / float(sum_total) * 100) if sum_total else 0

    return {
        "invoices": invoices,
        "rows": rows,
        "base_label": base_label,
        "add_label": add_label,
        "count_draft": count_draft, "count_sent": count_sent,
        "count_mail": count_mail, "count_post": count_post,
        "count_paid": count_paid, "count_open": count_open,
        "count_total": count_total, "count_live": count_live,
        "other_status_counts": other_status_counts,
        "sum_total": sum_total, "sum_paid": sum_paid, "sum_open": sum_open,
        "pct_sent": pct_sent, "pct_paid": pct_paid,
        "run_has_vat": run_has_vat, "mailable": mailable, "post_ids": post_ids,
    }


@bp.route("/billing-runs/<int:run_id>")
@login_required
def billing_run_detail(run_id):
    run = db.get_or_404(BillingRun, run_id)
    doc_format = AppSetting.get("invoice.document_format", "pdf")
    return render_template(
        "invoices/billing_run_detail.html",
        run=run, doc_format=doc_format, **_billing_run_overview(run),
    )


@bp.route("/billing-runs/<int:run_id>/delete", methods=["POST"])
@login_required
def billing_run_delete(run_id):
    run = db.get_or_404(BillingRun, run_id)
    invoices = run.invoices.all()

    deletable = [inv for inv in invoices if inv.status == Invoice.STATUS_DRAFT]
    locked = [inv for inv in invoices if inv.status != Invoice.STATUS_DRAFT]

    if locked:
        # Teilweise gesperrt: Feedback-Seite zeigen (kein Löschen)
        flash(
            f"{len(locked)} Rechnung(en) können nicht gelöscht werden (nicht mehr im Entwurfsstatus). "
            f"Bitte diese Rechnungen manuell stornieren oder entfernen, bevor der Rechnungslauf gelöscht werden kann.",
            "warning",
        )
        return redirect(url_for("invoices.billing_run_detail", run_id=run_id))

    # Jahre der gelöschten Rechnungen für optionalen Zähler-Reset merken
    reset_counter = request.form.get("reset_counter") == "1"
    years_to_reset = set()
    if reset_counter and deletable:
        for inv in deletable:
            try:
                years_to_reset.add(int(inv.invoice_number.split("-")[0]))
            except (ValueError, IndexError):
                pass

    # Alle löschbar — vorher Schätz-Korrekturen rückabwickeln (verrechnete
    # zurückgeben, aus Kappung erzeugte entfernen). Schlägt fehl, wenn eine
    # erzeugte Gutschrift schon in einer späteren Rechnung verrechnet wurde.
    try:
        for inv in deletable:
            reverse_corrections_for_invoice(inv)
        for inv in deletable:
            db.session.delete(inv)
        db.session.delete(run)
        db.session.commit()
    except ValueError as e:
        db.session.rollback()
        flash(str(e), "warning")
        return redirect(url_for("invoices.billing_run_detail", run_id=run_id))

    if years_to_reset:
        reset_parts = []
        for year in sorted(years_to_reset):
            prefix = f"{year}-"
            last = (
                Invoice.query
                .filter(Invoice.invoice_number.like(f"{prefix}%"))
                .order_by(Invoice.invoice_number.desc())
                .first()
            )
            if last:
                try:
                    new_seq = int(last.invoice_number.split("-")[-1]) + 1
                except ValueError:
                    new_seq = 1
            else:
                new_seq = 1
            counter = db.session.get(InvoiceCounter, year)
            if counter is None:
                counter = InvoiceCounter(year=year, next_seq=new_seq)
                db.session.add(counter)
            else:
                counter.next_seq = new_seq
            reset_parts.append(f"Zähler {year} → {new_seq:05d}")
        db.session.commit()
        flash(
            f"Rechnungslauf und {len(deletable)} Rechnung(en) gelöscht. "
            f"Rechnungsnummern-Zähler zurückgesetzt: {', '.join(reset_parts)}.",
            "success",
        )
    else:
        flash(f"Rechnungslauf und {len(deletable)} Rechnung(en) wurden gelöscht.", "success")
    return redirect(url_for("invoices.billing_runs"))


@bp.route("/billing-runs/<int:run_id>/post-bulk-merged", methods=["POST"])
@login_required
def billing_run_post_bulk(run_id):
    """Post-Versand eines Rechnungslaufs: markierte Rechnungen als zusammengeführtes
    PDF zum Download — und markiert dabei alle enthaltenen Entwürfe als „Versendet".

    Spiegelt die Seiteneffekte des Mailversands: das erzeugte PDF wird persistiert
    (Post-Versand-Beleg) und für Entwürfe der Offene Posten angelegt
    (``_create_or_update_open_item``). Bereits versendete/bezahlte Rechnungen im
    Bulk bleiben statusseitig unverändert, werden aber mit ausgeliefert.
    """
    import io

    run = db.get_or_404(BillingRun, run_id)
    invoice_ids = request.form.getlist("invoice_ids", type=int)
    if not invoice_ids:
        flash("Keine Rechnungen ausgewählt.", "warning")
        return redirect(url_for("invoices.billing_run_detail", run_id=run_id))
    if (resp := _bulk_print_limit_exceeded(invoice_ids)):
        return resp
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        if current_app.debug:
            flash("WeasyPrint nicht verfügbar. PDF-Export nur im Docker-Container verfügbar.", "info")
        else:
            flash("WeasyPrint ist nicht installiert. PDF-Export nur im Docker-Container verfügbar.", "danger")
        return redirect(url_for("invoices.billing_run_detail", run_id=run_id))
    from pypdf import PdfWriter

    # Auf den Lauf einschränken: keine fremden Rechnungen über manipulierte IDs.
    invoices = (
        run.invoices
        .filter(Invoice.id.in_(invoice_ids))
        .order_by(Invoice.invoice_number)
        .all()
    )
    if not invoices:
        flash("Keine Rechnungen gefunden.", "warning")
        return redirect(url_for("invoices.billing_run_detail", run_id=run_id))

    writer = PdfWriter()
    for invoice in invoices:
        html_content = _render_pdf_html(invoice)
        pdf_bytes = HTML(string=html_content).render().write_pdf()
        # Erzeugtes PDF immer persistieren (der heruntergeladene Post-Beleg).
        pdf_path = _versioned_path(_get_doc_dir(invoice), invoice.invoice_number, "pdf")
        with open(pdf_path, "wb") as fh:
            fh.write(pdf_bytes)
        invoice.pdf_path = pdf_path
        # Nur Entwürfe auf „Versendet" setzen; Offenen Posten wie beim Mailversand anlegen.
        if invoice.status == Invoice.STATUS_DRAFT:
            invoice.status = Invoice.STATUS_SENT
            account_id = _resolve_open_item_account_id(invoice, None)
            _create_or_update_open_item(invoice, account_id=account_id)
        writer.append(io.BytesIO(pdf_bytes))
    db.session.commit()
    writer.compress_identical_objects()
    merged_path = os.path.join(current_app.config["PDF_DIR"], "_bulk_merged.pdf")
    os.makedirs(current_app.config["PDF_DIR"], exist_ok=True)
    with open(merged_path, "wb") as fh:
        writer.write(fh)
    writer.close()
    return send_file(merged_path, as_attachment=True, download_name="Rechnungen_Post.pdf")


@bp.route("/billing-runs/<int:run_id>/export/excel")
@login_required
def billing_run_export_excel(run_id):
    """Übersicht eines Rechnungslaufs als Excel: oben das Deckblatt
    (Lauf-Metadaten, Tarif-Snapshot, Summen), darunter die Rechnungstabelle.

    Exportiert ausschließlich die **Übersichtsdaten** — nicht die einzelnen
    Rechnungen. Die Tabelle entspricht der Ansicht auf der Detailseite (alle
    Rechnungen inkl. stornierter, die Summen rechnen stornierte heraus).
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    run = db.get_or_404(BillingRun, run_id)
    ov = _billing_run_overview(run)
    rows = ov["rows"]
    run_has_vat = ov["run_has_vat"]
    base_label = ov["base_label"]
    add_label = ov["add_label"]

    EUR = '#,##0.00 "€"'
    EUR4 = '#,##0.0000 "€"'
    NUM = '#,##0.00'
    TITLE_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, size=11)
    LBL_FONT = Font(bold=True)
    HDR_FILL = PatternFill("solid", fgColor="2F5496")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    TOTAL_FONT = Font(bold=True)
    BORDER_TOP = Border(top=Side(style="thin"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rechnungslauf"

    def kv(row, label, value, fmt=None):
        ws.cell(row=row, column=1, value=label).font = LBL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        if fmt:
            cell.number_format = fmt

    r = 1
    title = "Rechnungslauf" + (f" — {run.billing_period.name}" if run.billing_period else "")
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=f"Erstellt am {run.created_at.strftime('%d.%m.%Y %H:%M')} Uhr")
    r += 2

    # --- Deckblatt: Details ---
    ws.cell(row=r, column=1, value="Details").font = SECTION_FONT
    r += 1
    kv(r, "Benutzer", run.created_by.username if run.created_by else "—"); r += 1
    kv(r, "Abrechnungsperiode", run.billing_period.name if run.billing_period else "—"); r += 1
    if run.sort_order:
        kv(r, "Sortierung", dict(BillingRun.SORT_ORDER_CHOICES).get(run.sort_order, run.sort_order)); r += 1
    kv(r, "Rechnungen erstellt", run.invoices_created); r += 1
    kv(r, "Übersprungen", run.invoices_skipped); r += 1
    r += 1

    # --- Deckblatt: Tarif-Snapshot ---
    ws.cell(row=r, column=1, value="Verwendeter Tarif (Kopie)").font = SECTION_FONT
    r += 1
    kv(r, "Name", run.tariff_name); r += 1
    kv(r, "Gültig", f"{run.tariff_valid_from or '—'} – {run.tariff_valid_to or 'aktuell'}"); r += 1
    if run.tariff_base_fee is not None:
        kv(r, base_label, float(run.tariff_base_fee), EUR); r += 1
    if run.tariff_additional_fee is not None:
        kv(r, add_label, float(run.tariff_additional_fee), EUR); r += 1
    kv(r, "Preis/m³", float(run.tariff_price_per_m3), EUR4); r += 1
    r += 1

    # --- Deckblatt: Summen ---
    ws.cell(row=r, column=1, value="Zusammenfassung").font = SECTION_FONT
    r += 1
    kv(r, "Gesamtsumme", float(ov["sum_total"]), EUR)
    ws.cell(row=r, column=3, value=f"{ov['count_paid'] + ov['count_open']} Rechnung(en)"); r += 1
    kv(r, "Bezahlt", float(ov["sum_paid"]), EUR)
    ws.cell(row=r, column=3, value=f"{ov['count_paid']} Rechnung(en)"); r += 1
    kv(r, "Offen", float(ov["sum_open"]), EUR)
    ws.cell(row=r, column=3, value=f"{ov['count_open']} Rechnung(en)"); r += 1
    kv(r, "Versendet", f"{ov['count_sent']} von {ov['count_live']}"); r += 1
    if ov["other_status_counts"]:
        kv(r, "Weitere Status",
           ", ".join(f"{k}: {v}" for k, v in ov["other_status_counts"].items())); r += 1
    r += 1

    # --- Rechnungstabelle ---
    headers = ["Rechnungsnr.", "Kunde", "Liegenschaft", "Adresse", "Versandart",
               "m³", "Wasser", base_label, add_label, "Netto"]
    if run_has_vat:
        headers.append("USt")
    headers += ["Brutto", "Mail-Status", "Status"]
    brutto_col = len(headers) - 2

    for c, val in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    r += 1

    for row in rows:
        inv = row["inv"]
        amt = row["amt"]
        addr = " ".join(p for p in [inv.property.strasse, inv.property.hausnummer] if p) if inv.property else ""
        liegenschaft = (inv.property.object_number or str(inv.property_id)) if inv.property else "—"
        c = 1
        ws.cell(row=r, column=c, value=inv.invoice_number); c += 1
        ws.cell(row=r, column=c, value=inv.customer.name); c += 1
        ws.cell(row=r, column=c, value=liegenschaft); c += 1
        ws.cell(row=r, column=c, value=addr); c += 1
        ws.cell(row=r, column=c, value="Mail" if row["wants_email"] else "Post"); c += 1
        mc = ws.cell(row=r, column=c, value=(float(amt["m3"]) if amt["m3"] else None)); mc.number_format = NUM; c += 1
        for key in ("water_net", "base_fee", "additional_fee", "net"):
            ws.cell(row=r, column=c, value=float(amt[key])).number_format = EUR; c += 1
        if run_has_vat:
            ws.cell(row=r, column=c, value=float(amt["ust"])).number_format = EUR; c += 1
        ws.cell(row=r, column=c, value=float(amt["gross"])).number_format = EUR; c += 1
        ws.cell(row=r, column=c, value=inv.last_email_status_de or ""); c += 1
        ws.cell(row=r, column=c, value=inv.status); c += 1
        r += 1

    if rows:
        ws.cell(row=r, column=1, value="Gesamt (ohne stornierte)").font = TOTAL_FONT
        tc = ws.cell(row=r, column=brutto_col, value=float(ov["sum_total"]))
        tc.number_format = EUR
        tc.font = TOTAL_FONT
        tc.border = BORDER_TOP

    for col in ws.columns:
        length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(45, max(11, length + 2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    period = run.billing_period.name if run.billing_period else "Lauf"
    safe_period = re.sub(r"[^0-9A-Za-zÄÖÜäöüß._-]+", "_", period)
    fname = f"Rechnungslauf_{safe_period}_{run.created_at.strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Tarife
# ---------------------------------------------------------------------------

@bp.route("/tariffs")
@login_required
def tariffs():
    all_tariffs = WaterTariff.query.order_by(WaterTariff.valid_from.desc()).all()
    return render_template("invoices/tariffs.html", tariffs=all_tariffs)


@bp.route("/tariffs/new", methods=["GET", "POST"])
@login_required
def tariff_new():
    if request.method == "POST":
        def _fee_or_none(field):
            v = request.form.get(field, "").strip().replace(",", ".")
            return Decimal(v) if v else None

        t = WaterTariff(
            name=request.form["name"].strip(),
            valid_from=int(request.form["valid_from"]),
            valid_to=int(request.form["valid_to"]) if request.form.get("valid_to") else None,
            base_fee=_fee_or_none("base_fee"),
            base_fee_label=request.form.get("base_fee_label", "").strip() or "Grundgebühr",
            additional_fee=_fee_or_none("additional_fee"),
            additional_fee_label=request.form.get("additional_fee_label", "").strip() or "Zusatzgebühr",
            price_per_m3=Decimal(request.form["price_per_m3"].replace(",", ".")),
            notes=request.form.get("notes", ""),
        )
        db.session.add(t)
        db.session.commit()
        flash("Tarif angelegt.", "success")
        return redirect(url_for("invoices.tariffs"))
    return render_template("invoices/tariff_form.html", tariff=None)


@bp.route("/tariffs/<int:tariff_id>/edit", methods=["GET", "POST"])
@login_required
def tariff_edit(tariff_id):
    t = db.get_or_404(WaterTariff, tariff_id)
    if request.method == "POST":
        def _fee_or_none(field):
            v = request.form.get(field, "").strip().replace(",", ".")
            return Decimal(v) if v else None

        t.name = request.form["name"].strip()
        t.valid_from = int(request.form["valid_from"])
        t.valid_to = int(request.form["valid_to"]) if request.form.get("valid_to") else None
        t.base_fee = _fee_or_none("base_fee")
        t.base_fee_label = request.form.get("base_fee_label", "").strip() or "Grundgebühr"
        t.additional_fee = _fee_or_none("additional_fee")
        t.additional_fee_label = request.form.get("additional_fee_label", "").strip() or "Zusatzgebühr"
        t.price_per_m3 = Decimal(request.form["price_per_m3"].replace(",", "."))
        t.notes = request.form.get("notes", "")
        db.session.commit()
        flash("Tarif aktualisiert.", "success")
        return redirect(url_for("invoices.tariffs"))
    return render_template("invoices/tariff_form.html", tariff=t)


# ---------------------------------------------------------------------------
# Rechnungsnummer-Zähler
# ---------------------------------------------------------------------------

@bp.route("/counters")
@login_required
def counters():
    from sqlalchemy import func
    # Alle Jahre aus bestehenden Rechnungen + vorhandenen Countern zusammenführen
    invoice_years = db.session.query(
        func.substr(Invoice.invoice_number, 1, 4).label("year"),
        func.count(Invoice.id).label("count"),
        func.max(Invoice.invoice_number).label("max_nr"),
    ).group_by(func.substr(Invoice.invoice_number, 1, 4)).all()

    all_counters = {c.year: c for c in InvoiceCounter.query.all()}

    rows = []
    for row in sorted(invoice_years, key=lambda r: r.year, reverse=True):
        try:
            y = int(row.year)
        except (TypeError, ValueError):
            continue
        counter = all_counters.get(y)
        rows.append({
            "year": y,
            "count": row.count,
            "max_nr": row.max_nr,
            "next_seq": counter.next_seq if counter else "–",
        })
    # Jahre mit Counter aber ohne Rechnungen ergänzen
    for y, counter in all_counters.items():
        if not any(r["year"] == y for r in rows):
            rows.append({
                "year": y,
                "count": 0,
                "max_nr": "–",
                "next_seq": counter.next_seq,
            })
    rows.sort(key=lambda r: r["year"], reverse=True)
    return render_template("invoices/counters.html", rows=rows)


@bp.route("/counters/<int:year>/reset", methods=["POST"])
@login_required
def counter_reset(year):
    from sqlalchemy import func
    mode = request.form.get("mode", "auto")
    counter = db.session.get(InvoiceCounter, year)

    if mode == "manual":
        try:
            new_seq = int(request.form.get("next_seq", 1))
            if new_seq < 1:
                raise ValueError
        except ValueError:
            flash("Ungültiger Wert für den Zähler.", "danger")
            return redirect(url_for("invoices.counters"))
    else:
        # auto: max vorhandene Nummer + 1
        prefix = f"{year}-"
        last = (
            Invoice.query
            .filter(Invoice.invoice_number.like(f"{prefix}%"))
            .order_by(Invoice.invoice_number.desc())
            .first()
        )
        if last:
            try:
                new_seq = int(last.invoice_number.split("-")[-1]) + 1
            except ValueError:
                new_seq = 1
        else:
            new_seq = 1

    if counter is None:
        counter = InvoiceCounter(year=year, next_seq=new_seq)
        db.session.add(counter)
    else:
        counter.next_seq = new_seq
    db.session.commit()
    flash(f"Zähler für {year} auf {new_seq:05d} zurückgesetzt.", "success")
    return redirect(url_for("invoices.counters"))
