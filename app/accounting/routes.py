import io
import csv
import json
import base64
import difflib
from collections import OrderedDict
from datetime import date
from decimal import Decimal, InvalidOperation

from flask import (
    render_template, redirect, url_for, flash, request,
    Response, stream_with_context, make_response,
)
from flask_login import login_required, current_user
from sqlalchemy import extract, func

from app.accounting import bp
from app.accounting import services as acc_svc
from app import tax_service
from app.extensions import db
from app.models import Account, Booking, BookingGroup, Invoice, OpenItem, WaterTariff, Customer, Project, RealAccount, RealAccountYearBalance, FiscalYear, FiscalYearReopenLog, Transfer, BillingPeriod, Property, PropertyOwnership, WaterMeter
from app.utils import next_invoice_number
from app.pagination import paginate_list, paginate_query


@bp.route("/")
@login_required
def index():
    year = request.args.get("year", date.today().year, type=int)
    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    return render_template("accounting/index.html", accounts=accounts, year=year)


# ---------------------------------------------------------------------------
# Kontenplan
# ---------------------------------------------------------------------------

@bp.route("/accounts")
@login_required
def accounts():
    all_accounts = Account.query.order_by(Account.name).all()
    return render_template("accounting/accounts.html", accounts=all_accounts)


def _validate_code(code_raw, model_cls, exclude_id=None):
    """Validiert ein 3-stelliges Kürzel (A-Z, 0-9). Gibt (code, fehlermeldung) zurück."""
    import re
    code = code_raw.strip().upper() or None
    if code:
        if not re.match(r'^[A-Z0-9]{3}$', code):
            return None, "Kürzel muss genau 3 Zeichen bestehen (Großbuchstaben A–Z oder Ziffern 0–9)."
        q = model_cls.query.filter(model_cls.code == code)
        if exclude_id is not None:
            q = q.filter(model_cls.id != exclude_id)
        existing = q.first()
        if existing:
            return None, f"Kürzel '{code}' wird bereits von '{existing.name}' verwendet."
    return code, None


@bp.route("/accounts/new", methods=["GET", "POST"])
@login_required
def account_new():
    if request.method == "POST":
        code, err = _validate_code(request.form.get("code", ""), Account)
        if err:
            flash(err, "danger")
            return render_template("accounting/account_form.html", account=None)
        a = Account(
            code=code,
            name=request.form["name"].strip(),
            description=request.form.get("description", ""),
        )
        db.session.add(a)
        db.session.commit()
        flash("Konto angelegt.", "success")
        return redirect(url_for("accounting.accounts"))
    return render_template("accounting/account_form.html", account=None)


@bp.route("/accounts/<int:account_id>/edit", methods=["GET", "POST"])
@login_required
def account_edit(account_id):
    a = db.get_or_404(Account, account_id)
    if request.method == "POST":
        code, err = _validate_code(request.form.get("code", ""), Account, exclude_id=a.id)
        if err:
            flash(err, "danger")
            return render_template("accounting/account_form.html", account=a)
        a.code = code
        a.name = request.form["name"].strip()
        a.description = request.form.get("description", "")
        a.active = "active" in request.form
        db.session.commit()
        flash("Konto aktualisiert.", "success")
        return redirect(url_for("accounting.accounts"))
    return render_template("accounting/account_form.html", account=a)


# ---------------------------------------------------------------------------
# Buchungen
# ---------------------------------------------------------------------------

# Buchhaltungs-Berechnungen sind im zentralen Service ``app.accounting.services``
# gebündelt. Die folgenden Aliase erhalten die bisherigen lokalen Aufrufstellen
# kompatibel und sorgen dafür, dass es nur eine einzige Quelle der Wahrheit gibt.
_locked_fiscal_year = acc_svc.locked_fiscal_year
_jan1_balance = acc_svc.jan1_balance
_year_end_balance = acc_svc.year_end_balance
_current_balance = acc_svc.current_balance
_year_movements = acc_svc.year_movements
_ust_period = acc_svc.ust_period
_ust_berechnen = acc_svc.ust_compute


# Selektoren der Filter-Leiste — fuer hx-include in den Filter-Controls UND im
# Massen-Bearbeiten-Modal (damit der Tabellen-Refresh die aktiven Filter behaelt).
_BOOKINGS_FILTER_INCLUDE = (
    "[name='year'],[name='month'],[name='quarter'],[name='real_account_id'],"
    "[name='kind'],[name='tax'],[name='account_id'],[name='project_id']"
)

# Wie _BOOKINGS_FILTER_INCLUDE, aber auf die Filter-Leiste eingeschraenkt
# (#bookings-filter-bar). Wird vom Tabellen-Refresher nach dem Speichern einer
# Buchung benutzt: ohne diese Eingrenzung wuerde das offene Buchungs-Modal (das
# ebenfalls Felder wie account_id/project_id traegt) doppelte Parameter
# einschleusen, sobald „Speichern und weitere Buchung" das Modal offen laesst.
_BOOKINGS_REFRESH_INCLUDE = ", ".join(
    "#bookings-filter-bar [name='%s']" % _n
    for _n in ("year", "month", "quarter", "real_account_id",
               "kind", "tax", "account_id", "project_id")
)


def _bookings_table_ctx(params, bulk_done_count=None):
    """Baut den Render-Kontext der Buchungstabelle aus den Filter-Parametern.

    ``params`` ist eine MultiDict — ``request.args`` beim Filter-GET,
    ``request.form`` beim Massen-Bearbeiten-Refresh; beide unterstuetzen
    ``.get(key, default, type=...)``. Liefert ``(table_ctx, filters)``:
    ``table_ctx`` geht direkt in ``_bookings_table.html``, ``filters`` haelt die
    geparsten Filter-Echo-Werte fuer das Voll-Seiten-Template.
    """
    year = params.get("year", date.today().year, type=int)
    account_id = params.get("account_id", "", type=str)
    project_id = params.get("project_id", "", type=str)
    real_account_id = params.get("real_account_id", "", type=str)
    # kind: "" / "income" / "expense" — schränkt auf Einnahmen oder Ausgaben ein
    kind = params.get("kind", "", type=str)
    if kind not in ("income", "expense"):
        kind = ""
    # month (1–12) / quarter (1–4): zusätzliche Datumsfilter innerhalb des Jahres
    month = params.get("month", 0, type=int)
    if month not in range(1, 13):
        month = 0
    quarter = params.get("quarter", 0, type=int)
    if quarter not in range(1, 5):
        quarter = 0
    # tax: "" (alle) / "any" (alle mit Steuer ≠ 0%) / Satz-Wert ("10", "20", …)
    tax = params.get("tax", "", type=str)
    _tax_values = {str(int(r)) for r in tax_service.tax_rate_values()}
    if tax not in ({"any"} | _tax_values):
        tax = ""

    query = (
        Booking.query
        .filter(extract("year", Booking.date) == year)
        .order_by(
            Booking.date.desc(),
            func.coalesce(Booking.storno_of_id, Booking.id).desc(),
        )
    )
    if account_id:
        query = query.filter(Booking.account_id == int(account_id))
    if project_id:
        query = query.filter(Booking.project_id == int(project_id))
    if real_account_id:
        query = query.filter(Booking.real_account_id == int(real_account_id))
    if month:
        query = query.filter(extract("month", Booking.date) == month)
    if quarter:
        query = query.filter(
            extract("month", Booking.date).between(quarter * 3 - 2, quarter * 3)
        )
    if kind == "income":
        query = query.filter(Booking.amount > 0)
    elif kind == "expense":
        query = query.filter(Booking.amount < 0)
    # 0 % wird als tax_rate NULL gespeichert (siehe booking_new/_edit).
    if tax == "any":
        query = query.filter(Booking.tax_rate.isnot(None), Booking.tax_rate > 0)
    elif tax == "0":
        query = query.filter(db.or_(Booking.tax_rate.is_(None), Booking.tax_rate == 0))
    elif tax:
        query = query.filter(Booking.tax_rate == Decimal(tax))

    bkgs = query.all()

    # Umbuchungen laden (gleicher Jahres- und Bankkontofilter)
    transfer_query = (
        Transfer.query
        .filter(extract("year", Transfer.date) == year)
        .order_by(Transfer.date.desc())
    )
    if real_account_id:
        ra_id = int(real_account_id)
        transfer_query = transfer_query.filter(
            db.or_(
                Transfer.from_real_account_id == ra_id,
                Transfer.to_real_account_id == ra_id,
            )
        )
    if month:
        transfer_query = transfer_query.filter(extract("month", Transfer.date) == month)
    if quarter:
        transfer_query = transfer_query.filter(
            extract("month", Transfer.date).between(quarter * 3 - 2, quarter * 3)
        )
    # account_id/project_id/kind/tax gelten nur für Buchungen, nicht für
    # Umbuchungen. Bei aktivem Einnahmen-/Ausgaben- oder Steuerfilter werden
    # Umbuchungen ebenfalls ausgeblendet, da sie weder echte Einnahmen noch
    # Ausgaben sind und keinen Steuersatz tragen.
    transfers = [] if (account_id or project_id or kind or tax) else transfer_query.all()

    # Grouping-Modus: Sammelbuchungen werden nur im ungefilterten Jahresview
    # gruppiert. Sobald Konto-/Projekt-/Bankkonto-Filter aktiv sind, fällt die
    # Ansicht auf eine flache Zeilenliste zurück (ADR-002), damit Filter
    # innerhalb einer Sammelbuchung konsistent wirken.
    group_mode = not (account_id or project_id or real_account_id or kind or tax)

    rows = []
    seen_groups = OrderedDict()  # group_id → row-dict (für In-Place-Update)
    for b in bkgs:
        if group_mode and b.group_id:
            entry = seen_groups.get(b.group_id)
            if entry is None:
                group = b.group
                entry = {
                    "type": "booking_group",
                    "group": group,
                    "children": [],
                    "date": group.date if group else b.date,
                }
                seen_groups[b.group_id] = entry
                rows.append(entry)
            entry["children"].append(b)
        else:
            rows.append({"type": "booking", "obj": b, "date": b.date})

    ra_id_int = int(real_account_id) if real_account_id else None
    for t in transfers:
        if ra_id_int:
            # Nur die relevante Seite zeigen
            if t.from_real_account_id == ra_id_int:
                rows.append({"type": "transfer", "obj": t, "date": t.date,
                             "side": "out", "amount": -t.amount})
            else:
                rows.append({"type": "transfer", "obj": t, "date": t.date,
                             "side": "in", "amount": t.amount})
        else:
            # Beide Seiten anzeigen
            rows.append({"type": "transfer", "obj": t, "date": t.date,
                         "side": "out", "amount": -t.amount})
            rows.append({"type": "transfer", "obj": t, "date": t.date,
                         "side": "in", "amount": t.amount})
    rows.sort(key=lambda r: r["date"], reverse=True)

    closed_fys = FiscalYear.query.filter_by(closed=True).all()
    locked_booking_ids = set()
    for b in bkgs:
        for fy in closed_fys:
            if fy.start_date <= b.date <= fy.end_date:
                locked_booking_ids.add(b.id)
                break

    # Stornopaare (Original + Gegenbuchung) müssen gemeinsam ausgeschlossen werden,
    # sonst zählt der frühere "status != STORNIERT"-Filter nur die Gegenbuchung mit
    # und verfälscht die Summe.
    effective_bkgs = [b for b in bkgs if acc_svc.is_effective_booking(b)]
    total_bookings = sum((b.amount for b in effective_bkgs), Decimal("0"))
    total_transfers = sum((r["amount"] for r in rows if r["type"] == "transfer"), Decimal("0"))
    total_amount = total_bookings + total_transfers

    total_vorsteuer = sum(
        (acc_svc.booking_tax(b) for b in effective_bkgs if b.amount < 0),
        Decimal("0"),
    )
    total_ust = sum(
        (acc_svc.booking_tax(b) for b in effective_bkgs if b.amount > 0),
        Decimal("0"),
    )

    pagination = paginate_list(rows, page_key="bookings")

    table_ctx = dict(
        rows=pagination.items, year=year,
        now_year=date.today().year,
        total_amount=total_amount,
        total_vorsteuer=total_vorsteuer,
        total_ust=total_ust,
        locked_booking_ids=locked_booking_ids,
        pagination=pagination,
        bulk_done_count=bulk_done_count,
    )
    filters = dict(
        account_id=account_id, project_id=project_id,
        real_account_id=real_account_id,
        kind=kind, month=month, quarter=quarter, tax=tax,
    )
    return table_ctx, filters


@bp.route("/bookings")
@login_required
def bookings():
    table_ctx, filters = _bookings_table_ctx(request.args)

    if request.headers.get("HX-Request"):
        return render_template("accounting/_bookings_table.html", **table_ctx)

    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    projects = Project.query.order_by(Project.name).all()
    real_accounts = RealAccount.query.filter_by(active=True).order_by(RealAccount.name).all()
    # Fuer das Massen-Bearbeiten-Modal: aktive Kontakte + nicht abgeschlossene
    # Projekte als Ziel-Auswahl (analog booking_form).
    customers = Customer.query.filter_by(active=True).order_by(Customer.name).all()
    assignable_projects = Project.query.filter_by(closed=False).order_by(Project.name).all()

    return render_template(
        "accounting/bookings.html",
        accounts=accounts, projects=projects,
        real_accounts=real_accounts,
        customers=customers, assignable_projects=assignable_projects,
        tax_rates=tax_service.tax_rates(),
        filter_include=_BOOKINGS_FILTER_INCLUDE,
        refresh_include=_BOOKINGS_REFRESH_INCLUDE,
        **filters,
        **table_ctx,
    )


@bp.route("/bookings/bulk-edit", methods=["POST"])
@login_required
def bookings_bulk_edit():
    """Setzt Konto/Projekt/Kontakt auf mehreren ausgewaehlten Buchungen.

    Nur ausgefuellte Felder werden uebernommen (leeres Feld = unveraendert).
    Nicht editierbare Buchungen werden uebersprungen — dieselben Guards wie in
    ``booking_edit``: Sammelbuchungs-Zeilen (``group_id``), stornierte bzw.
    Storno-Buchungen und Buchungen in abgeschlossenen Buchungsjahren. Die
    Antwort ist das neu gerenderte Tabellen-Fragment; die aktiven Filter kommen
    via hx-include mit und bleiben dadurch erhalten.
    """
    raw_ids = request.form.getlist("booking_ids")
    ids = []
    for r in raw_ids:
        try:
            ids.append(int(r))
        except (TypeError, ValueError):
            continue

    def _resolve(field, model):
        """(neue_id, soll_gesetzt_werden) — leeres/ungueltiges Feld ⇒ unveraendert."""
        raw = (request.form.get(field) or "").strip()
        if not raw:
            return None, False
        try:
            obj = db.session.get(model, int(raw))
        except (TypeError, ValueError):
            return None, False
        if obj is None:
            return None, False
        return obj.id, True

    new_account_id, set_account = _resolve("bulk_account_id", Account)
    new_project_id, set_project = _resolve("bulk_project_id", Project)
    new_customer_id, set_customer = _resolve("bulk_customer_id", Customer)

    updated = 0
    if ids and (set_account or set_project or set_customer):
        for b in Booking.query.filter(Booking.id.in_(ids)).all():
            if b.group_id is not None:
                continue
            if b.status == Booking.STATUS_STORNIERT or b.storno_of_id is not None:
                continue
            if _locked_fiscal_year(b.date):
                continue
            if set_account:
                b.account_id = new_account_id
            if set_project:
                b.project_id = new_project_id
            if set_customer:
                b.customer_id = new_customer_id
            updated += 1
        db.session.commit()

    table_ctx, _ = _bookings_table_ctx(request.form, bulk_done_count=updated)
    return render_template("accounting/_bookings_table.html", **table_ctx)


# ---------------------------------------------------------------------------
# Buchung anlegen / bearbeiten — Modal-first, htmx- und dialekt-robust.
#
# Beide Routen liefern bei einem HX-Request (Modal) nur das Formular-Fragment
# (`_booking_form.html`) bzw. nach dem Speichern ein frisches Fragment + den
# `booking-saved`-HX-Trigger; bei normalen Requests (Direktlink/Fallback) die
# Vollseite (`booking_form.html`). Die Validierung sitzt zentral in
# `_parse_booking_form`, sodass kein roher `int()`/`Decimal()`-Cast mehr einen
# 500er werfen kann und Eingaben bei Fehlern erhalten bleiben.
# ---------------------------------------------------------------------------

def _opt_int(raw):
    """Parst einen optionalen Integer aus einem Formularwert (leer → None)."""
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return int(raw)
    except (ValueError, TypeError):
        return None


def _valid_fk(raw, model_cls):
    """Parst eine optionale FK-Id und prueft Existenz; sonst None.

    Schuetzt vor IntegrityError beim Commit, falls ein manipulierter oder
    veralteter Wert ankommt — er wird dann schlicht verworfen.
    """
    pk = _opt_int(raw)
    if pk is None:
        return None
    return pk if db.session.get(model_cls, pk) is not None else None


def _booking_form_context(booking=None):
    """Gemeinsamer Render-Kontext fuer das Buchungsformular (Modal + Vollseite).

    Bei einer Bearbeitung werden referenzierte, aber inaktive/abgeschlossene
    Entitaeten den Auswahllisten hinzugefuegt, damit sie sichtbar bleiben und
    beim Speichern nicht verloren gehen.
    """
    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    projects = Project.query.filter_by(closed=False).order_by(Project.name).all()
    real_accounts = RealAccount.query.filter_by(active=True).order_by(RealAccount.name).all()
    customers = Customer.query.filter_by(active=True).order_by(Customer.name).all()
    if booking is not None:
        if booking.account is not None and booking.account not in accounts:
            accounts.append(booking.account)
        if booking.project is not None and booking.project not in projects:
            projects.append(booking.project)
        if booking.real_account is not None and booking.real_account not in real_accounts:
            real_accounts.append(booking.real_account)
        if booking.customer is not None and booking.customer not in customers:
            customers.append(booking.customer)
    return dict(
        accounts=accounts,
        projects=projects,
        real_accounts=real_accounts,
        customers=customers,
        tax_rates=tax_service.tax_rates(),
        default_real_account=RealAccount.query.filter_by(is_default=True, active=True).first(),
        today=date.today(),
    )


def _parse_booking_form(form, today, *, booking=None):
    """Validiert + parst die Buchungs-Formulardaten.

    Liefert ``(data, errors)``. ``errors`` ist ein dict {feld → deutsche Meldung}
    und leer, wenn alles gueltig ist. Bei verbuchten Buchungen werden
    Datum/Betrag/Steuersatz/Belegnummer nicht geparst (gesperrt) — der Aufrufer
    behaelt die bestehenden Werte.
    """
    errors = {}
    is_verbucht = booking is not None and booking.status == Booking.STATUS_VERBUCHT
    data = {}

    # Konto (Pflicht, muss existieren)
    account_id_raw = (form.get("account_id") or "").strip()
    account = db.session.get(Account, int(account_id_raw)) if account_id_raw.isdigit() else None
    if account is None:
        errors["account_id"] = "Bitte ein gültiges Konto wählen."
    data["account"] = account

    # Beschreibung (Pflicht)
    description = (form.get("description") or "").strip()
    if not description:
        errors["description"] = "Bitte eine Beschreibung eingeben."
    data["description"] = description

    # Optionale Zuordnungen (existenzgeprueft, gegen FK-Verletzungen)
    data["project_id"] = _valid_fk(form.get("project_id"), Project)
    data["customer_id"] = _valid_fk(form.get("customer_id"), Customer)
    data["real_account_id"] = _valid_fk(form.get("real_account_id"), RealAccount)

    if not is_verbucht:
        # Datum (Pflicht, gueltig, nicht in der Zukunft, offenes Buchungsjahr)
        date_raw = (form.get("date") or "").strip()
        booking_date = None
        if not date_raw:
            errors["date"] = "Bitte ein Datum wählen."
        else:
            try:
                booking_date = date.fromisoformat(date_raw)
            except ValueError:
                errors["date"] = "Ungültiges Datum."
        if booking_date is not None:
            if booking_date > today:
                errors["date"] = "Das Buchungsdatum darf nicht in der Zukunft liegen."
            else:
                fy_error = acc_svc.open_fiscal_year_error(booking_date)
                if fy_error:
                    errors["date"] = fy_error
        data["date"] = booking_date

        # Betrag (Pflicht, gueltige Zahl, ungleich 0)
        amount_raw = (form.get("amount") or "").strip().replace(",", ".")
        amount = None
        if amount_raw == "":
            errors["amount"] = "Bitte einen Betrag eingeben."
        else:
            try:
                amount = Decimal(amount_raw)
            except (InvalidOperation, ValueError):
                errors["amount"] = "Ungültiger Betrag."
        if amount is not None and amount == 0:
            errors["amount"] = "Der Betrag darf nicht 0 sein."
        data["amount"] = amount

        # Steuersatz (optional; 0/ungueltig/unbekannt → kein Steuersatz)
        tax_rate_raw = (form.get("tax_rate") or "").strip().replace(",", ".")
        try:
            tax_rate = Decimal(tax_rate_raw) if tax_rate_raw else Decimal("0")
        except (InvalidOperation, ValueError):
            tax_rate = Decimal("0")
        known_rates = set(tax_service.tax_rate_values())
        data["tax_rate"] = tax_rate if (tax_rate > 0 and tax_rate in known_rates) else None

        # Belegnummer (optional)
        data["reference"] = (form.get("reference") or "").strip()

    return data, errors


def _render_booking_form(*, booking, ctx, hx, keep_date="", form_data=None,
                         errors=None, sticky=None):
    """Rendert das Buchungsformular — als Fragment (hx/Modal) oder Vollseite."""
    is_verbucht = booking is not None and booking.status == Booking.STATUS_VERBUCHT
    form_action_url = (
        url_for("accounting.booking_edit", booking_id=booking.id) if booking
        else url_for("accounting.booking_new")
    )
    template = "accounting/_booking_form.html" if hx else "accounting/booking_form.html"
    return render_template(
        template,
        booking=booking,
        is_modal=hx,
        is_verbucht=is_verbucht,
        form_action_url=form_action_url,
        form_data=form_data,
        errors=errors or {},
        keep_date=keep_date,
        sticky=sticky or {},
        **ctx,
    )


def _booking_form_error(errors, form_data, *, booking, hx):
    """Antwort bei Validierungsfehlern: Formular mit erhaltenen Werten + Fehlern."""
    ctx = _booking_form_context(booking=booking)
    if not hx:
        for msg in errors.values():
            flash(msg, "danger")
    return _render_booking_form(
        booking=booking, ctx=ctx, hx=hx, form_data=form_data, errors=errors,
    )


def _booking_blocked(message, hx):
    """Antwort, wenn die Bearbeitung nicht erlaubt ist (Gruppe/Storno/FY)."""
    if hx:
        return render_template("accounting/_booking_blocked.html", message=message)
    flash(message, "warning")
    return redirect(url_for("accounting.bookings"))


# Platzhalter, der nach dem Speichern ins Modal geswappt wird. Bewusst KEIN
# Formular: der Client schliesst das Modal vollstaendig (und oeffnet es bei
# „weiteres" frisch wieder). So entstehen keine doppelten TomSelects durch
# In-Place-Swaps in einem offenen Modal.
_BOOKING_SAVED_PLACEHOLDER = (
    '<div class="text-center py-5 text-muted">'
    '<div class="spinner-border" role="status"><span class="visually-hidden">…</span></div>'
    '</div>'
)


def _booking_saved_response(*, action, last_date=None, last_real_account_id=None):
    """HX-Antwort nach erfolgreichem Speichern.

    Liefert nur einen Platzhalter plus einen ``booking-saved``-HX-Trigger. Der
    Client aktualisiert daraufhin die Tabelle und schliesst das Modal — bei
    „weiteres" oeffnet er es frisch wieder (Datum/Bankkonto werden im Trigger
    fuer die naechste Erfassung mitgegeben).
    """
    weiteres = action == "weiteres"
    detail = {"action": "weiteres" if weiteres else "save"}
    if weiteres:
        if last_date:
            detail["date"] = last_date.isoformat()
        if last_real_account_id:
            detail["real_account_id"] = last_real_account_id
    resp = make_response(_BOOKING_SAVED_PLACEHOLDER)
    resp.headers["HX-Trigger"] = json.dumps({"booking-saved": detail})
    return resp


@bp.route("/bookings/new", methods=["GET", "POST"])
@login_required
def booking_new():
    today = date.today()
    hx = bool(request.headers.get("HX-Request"))

    if request.method == "POST":
        data, errors = _parse_booking_form(request.form, today)
        if errors:
            return _booking_form_error(errors, request.form, booking=None, hx=hx)
        b = Booking(
            date=data["date"],
            account_id=data["account"].id,
            amount=data["amount"],
            description=data["description"],
            reference=data["reference"],
            project_id=data["project_id"],
            real_account_id=data["real_account_id"],
            customer_id=data["customer_id"],
            tax_rate=data["tax_rate"],
            created_by_id=current_user.id,
        )
        db.session.add(b)
        db.session.commit()
        action = request.form.get("action")
        if hx:
            return _booking_saved_response(
                action=action, last_date=data["date"],
                last_real_account_id=data["real_account_id"],
            )
        flash("Buchung gespeichert.", "success")
        if action == "weiteres":
            return redirect(url_for("accounting.booking_new", date=data["date"].isoformat()))
        return redirect(url_for("accounting.bookings"))

    keep_date = request.args.get("date", "")
    # Sticky-Bankkonto beim erneuten Oeffnen via „weiteres" (Query-Param vom Client).
    sticky = {}
    ra_raw = (request.args.get("real_account_id") or "").strip()
    if ra_raw.isdigit() and db.session.get(RealAccount, int(ra_raw)) is not None:
        sticky["real_account_id"] = int(ra_raw)
    ctx = _booking_form_context(booking=None)
    return _render_booking_form(booking=None, ctx=ctx, hx=hx, keep_date=keep_date, sticky=sticky)


@bp.route("/bookings/<int:booking_id>/edit", methods=["GET", "POST"])
@login_required
def booking_edit(booking_id):
    b = db.get_or_404(Booking, booking_id)
    today = date.today()
    hx = bool(request.headers.get("HX-Request"))

    if b.group_id is not None:
        if hx:
            return _booking_blocked(
                "Kinder einer Sammelbuchung können nicht einzeln bearbeitet werden. "
                "Bitte die Sammelbuchung stornieren und neu anlegen.", True,
            )
        flash(
            "Kinder einer Sammelbuchung können nicht einzeln bearbeitet werden. "
            "Bitte Sammelbuchung stornieren und neu anlegen.",
            "warning",
        )
        return redirect(url_for("accounting.booking_group_edit", group_id=b.group_id))
    if b.status == Booking.STATUS_STORNIERT:
        return _booking_blocked("Stornierte Buchungen können nicht bearbeitet werden.", hx)
    fy_locked = _locked_fiscal_year(b.date)
    if fy_locked:
        return _booking_blocked(
            f"Das Buchungsjahr {fy_locked.year} ist abgeschlossen. "
            "Diese Buchung kann nicht bearbeitet werden.", hx,
        )

    is_verbucht = b.status == Booking.STATUS_VERBUCHT

    if request.method == "POST":
        data, errors = _parse_booking_form(request.form, today, booking=b)
        if errors:
            return _booking_form_error(errors, request.form, booking=b, hx=hx)
        b.account_id = data["account"].id
        b.description = data["description"]
        b.project_id = data["project_id"]
        b.real_account_id = data["real_account_id"]
        b.customer_id = data["customer_id"]
        if not is_verbucht:
            b.amount = data["amount"]
            b.date = data["date"]
            b.reference = data["reference"]
            b.tax_rate = data["tax_rate"]
        db.session.commit()
        if hx:
            return _booking_saved_response(action="save")
        flash("Buchung aktualisiert.", "success")
        return redirect(url_for("accounting.bookings"))

    ctx = _booking_form_context(booking=b)
    return _render_booking_form(booking=b, ctx=ctx, hx=hx)


@bp.route("/bookings/<int:booking_id>/delete", methods=["POST"])
@login_required
def booking_delete(booking_id):
    b = db.get_or_404(Booking, booking_id)
    # Sammelbuchungs-Kinder dürfen nicht einzeln gelöscht werden (ADR-002,
    # Regel 4: Storno/Löschung immer der ganzen Gruppe).
    if b.group_id is not None:
        flash(
            "Diese Buchung gehört zu einer Sammelbuchung und kann nicht "
            "einzeln gelöscht werden. Bitte Sammelbuchung stornieren.",
            "warning",
        )
        return redirect(url_for("accounting.booking_group_edit", group_id=b.group_id))
    if b.status != Booking.STATUS_OFFEN:
        flash("Nur offene Buchungen können gelöscht werden.", "warning")
        return redirect(url_for("accounting.bookings"))
    fy_locked = _locked_fiscal_year(b.date)
    if fy_locked:
        flash(f"Das Buchungsjahr {fy_locked.year} ist abgeschlossen. Diese Buchung kann nicht gelöscht werden.", "danger")
        return redirect(url_for("accounting.bookings"))
    db.session.delete(b)
    db.session.commit()
    flash("Buchung gelöscht.", "info")
    return redirect(url_for("accounting.bookings"))


@bp.route("/bookings/<int:booking_id>/stornieren", methods=["GET", "POST"])
@login_required
def booking_stornieren(booking_id):
    b = db.get_or_404(Booking, booking_id)

    # Sammelbuchungs-Kinder dürfen nicht einzeln storniert werden (ADR-002,
    # Regel 4: Storno immer der ganzen Gruppe). Weiterleitung auf das
    # Gruppen-Storno-Formular, damit der Anwender die gesamte Sammelbuchung
    # zurücksetzt.
    if b.group_id is not None:
        flash(
            "Diese Buchung gehört zu einer Sammelbuchung und kann nur als "
            "Ganzes storniert werden.",
            "warning",
        )
        return redirect(url_for("accounting.booking_group_stornieren", group_id=b.group_id))

    if b.status == Booking.STATUS_STORNIERT:
        flash("Diese Buchung ist bereits storniert.", "warning")
        return redirect(url_for("accounting.bookings"))
    if b.storno_of_id is not None:
        flash("Eine Storno-Buchung kann nicht erneut storniert werden.", "warning")
        return redirect(url_for("accounting.bookings"))
    fy_locked = _locked_fiscal_year(b.date)
    if fy_locked:
        flash(f"Das Buchungsjahr {fy_locked.year} ist abgeschlossen. Diese Buchung kann nicht storniert werden.", "danger")
        return redirect(url_for("accounting.bookings"))
    if b.date.year != date.today().year:
        flash("Buchungen aus Vorjahren können nicht storniert werden.", "warning")
        return redirect(url_for("accounting.bookings"))

    if request.method == "POST":
        reason = request.form.get("storno_reason", "").strip()
        if not reason:
            flash("Bitte einen Storno-Grund angeben.", "danger")
            return render_template("accounting/storno_form.html", booking=b)

        try:
            # Storno-Buchung anlegen (gleiches Datum wie Ursprungsbuchung)
            storno = Booking(
                date=b.date,
                account_id=b.account_id,
                amount=b.amount * -1,
                description=f"Storno: {b.description}",
                invoice_id=b.invoice_id,
                open_item_id=b.open_item_id,
                project_id=b.project_id,
                tax_rate=b.tax_rate,
                storno_of_id=b.id,
                storno_reason=reason,
                storno_date=date.today(),
                status=Booking.STATUS_VERBUCHT,
                created_by_id=current_user.id,
            )
            db.session.add(storno)

            # Ursprungsbuchung als storniert markieren
            b.status = Booking.STATUS_STORNIERT

            # Verknüpfte Rechnung stornieren
            cancelled_invoice_number = None
            if b.invoice_id:
                inv = db.session.get(Invoice, b.invoice_id)
                if inv and inv.status not in (Invoice.STATUS_CANCELLED,):
                    inv.status = Invoice.STATUS_CANCELLED
                    cancelled_invoice_number = inv.invoice_number

            # Offenen Posten zurücksetzen wenn gewünscht
            if b.open_item_id and request.form.get("close_open_item"):
                oi = db.session.get(OpenItem, b.open_item_id)
                if oi:
                    oi.status = OpenItem.STATUS_OPEN

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Stornieren – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("accounting.bookings"))

        if cancelled_invoice_number:
            flash(
                f"Rechnung {cancelled_invoice_number} wurde storniert. "
                f"Bitte eine neue Rechnung ausstellen.",
                "warning",
            )
        flash("Buchung erfolgreich storniert.", "success")
        return redirect(url_for("accounting.bookings"))

    return render_template("accounting/storno_form.html", booking=b)


# ---------------------------------------------------------------------------
# Sammelbuchungen (BookingGroup) — ADR-002
# ---------------------------------------------------------------------------

@bp.route("/booking-groups/<int:group_id>")
@login_required
def booking_group_detail(group_id):
    """Alt-URL – leitet auf die Bearbeiten-Ansicht um.

    Die reine Ansichts-Route wurde entfernt; ``booking_group_edit`` rendert
    das Formular readonly, sobald die Gruppe nicht mehr bearbeitbar ist.
    """
    return redirect(url_for("accounting.booking_group_edit", group_id=group_id))


@bp.route("/booking-groups/new", methods=["GET", "POST"])
@login_required
def booking_group_new():
    """Manuelle Anlage einer Sammelbuchung.

    Der Header enthält Datum, Beschreibung, Belegnummer, optional Kunde/Rechnung/
    Bankkonto. Die Kinder werden als dynamische Zeilen (Konto, Projekt, Steuersatz,
    Betrag, Beschreibung) geliefert. Es müssen mindestens 2 Zeilen übergeben
    werden — einzeilige Sammelbuchungen werden vom Editor abgewiesen (ADR-002,
    Regel 1).
    """
    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    active_projects = Project.query.filter_by(closed=False).order_by(Project.name).all()
    real_accounts = RealAccount.query.filter_by(active=True).order_by(RealAccount.name).all()
    customers = Customer.query.filter_by(active=True).order_by(Customer.name).all()
    tax_rates = tax_service.tax_rates()
    default_real_account = RealAccount.query.filter_by(is_default=True, active=True).first()
    today = date.today()

    def _render_new(**extra):
        return render_template(
            "accounting/booking_group_form.html",
            group=None,
            accounts=accounts,
            projects=active_projects,
            real_accounts=real_accounts,
            customers=customers,
            tax_rates=tax_rates,
            default_real_account=default_real_account,
            today=today,
            **extra,
        )

    if request.method == "POST":
        try:
            group_date = date.fromisoformat(request.form["date"])
        except (KeyError, ValueError):
            flash("Ungültiges Datum.", "danger")
            return _render_new(form_data=request.form)
        if group_date > today:
            flash("Das Datum darf nicht in der Zukunft liegen.", "danger")
            return _render_new(form_data=request.form)
        fy_error = acc_svc.open_fiscal_year_error(group_date)
        if fy_error:
            flash(f"{fy_error} Sammelbuchung wurde nicht gespeichert.", "danger")
            return _render_new(form_data=request.form)

        description = request.form.get("description", "").strip()
        if not description:
            flash("Bitte eine Beschreibung angeben.", "danger")
            return _render_new(form_data=request.form)

        reference = request.form.get("reference", "").strip() or None
        customer_id_raw = request.form.get("customer_id") or None
        invoice_id_raw = request.form.get("invoice_id") or None
        real_account_id_raw = request.form.get("real_account_id") or None

        customer_id = int(customer_id_raw) if customer_id_raw else None
        invoice_id = int(invoice_id_raw) if invoice_id_raw else None
        real_account_id = int(real_account_id_raw) if real_account_id_raw else None

        # Kind-Zeilen einsammeln
        child_accounts = request.form.getlist("child_account_id[]")
        child_projects = request.form.getlist("child_project_id[]")
        child_tax_rates = request.form.getlist("child_tax_rate[]")
        child_amounts = request.form.getlist("child_amount[]")
        child_descriptions = request.form.getlist("child_description[]")

        parsed_children = []
        for i, acc_raw in enumerate(child_accounts):
            if not acc_raw:
                continue
            try:
                acc_id = int(acc_raw)
            except ValueError:
                continue
            amount_raw = child_amounts[i].replace(",", ".") if i < len(child_amounts) else "0"
            try:
                amount = Decimal(amount_raw)
            except Exception:
                continue
            if amount == 0:
                continue
            proj_raw = child_projects[i] if i < len(child_projects) else ""
            try:
                proj_id = int(proj_raw) if proj_raw else None
            except ValueError:
                proj_id = None
            tax_raw = child_tax_rates[i] if i < len(child_tax_rates) else ""
            try:
                tax_rate = Decimal(tax_raw) if tax_raw else Decimal("0")
            except Exception:
                tax_rate = Decimal("0")
            desc = (
                child_descriptions[i].strip()
                if i < len(child_descriptions) and child_descriptions[i].strip()
                else description
            )
            parsed_children.append({
                "account_id": acc_id,
                "project_id": proj_id,
                "tax_rate": tax_rate,
                "amount": amount,
                "description": desc,
            })

        if len(parsed_children) < 2:
            flash(
                "Eine Sammelbuchung benötigt mindestens 2 Zeilen. "
                "Für Einzelbuchungen bitte 'Neue Buchung' verwenden.",
                "danger",
            )
            return _render_new(form_data=request.form)

        try:
            group = BookingGroup(
                date=group_date,
                description=description,
                reference=reference,
                invoice_id=invoice_id,
                customer_id=customer_id,
                total_amount=Decimal("0"),
                status=BookingGroup.STATUS_AKTIV,
                created_by_id=current_user.id,
            )
            db.session.add(group)
            db.session.flush()

            for c in parsed_children:
                child = Booking(
                    date=group_date,
                    account_id=c["account_id"],
                    project_id=c["project_id"],
                    amount=c["amount"],
                    description=c["description"],
                    reference=reference,
                    invoice_id=invoice_id,
                    customer_id=customer_id,
                    real_account_id=real_account_id,
                    tax_rate=c["tax_rate"] if c["tax_rate"] and c["tax_rate"] > 0 else None,
                    group_id=group.id,
                    created_by_id=current_user.id,
                )
                db.session.add(child)

            db.session.flush()
            acc_svc.recompute_group_total(group.id)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Anlegen der Sammelbuchung – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return _render_new(form_data=request.form)

        flash("Sammelbuchung angelegt.", "success")
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))

    return _render_new()


def _group_is_editable(group):
    """Liefert (ok, msg). Eine Sammelbuchung darf bearbeitet/gelöscht werden,
    solange sie aktiv ist, kein Kind bereits verbucht wurde und das
    Buchungsjahr nicht abgeschlossen ist (analog zur Einzelbuchung)."""
    if group.status == BookingGroup.STATUS_STORNIERT:
        return False, "Stornierte Sammelbuchungen können nicht bearbeitet werden."
    fy_locked = _locked_fiscal_year(group.date)
    if fy_locked:
        return False, (
            f"Das Buchungsjahr {fy_locked.year} ist abgeschlossen. "
            f"Diese Sammelbuchung kann nicht bearbeitet werden."
        )
    for c in group.children:
        if c.status != Booking.STATUS_OFFEN:
            return False, (
                "Diese Sammelbuchung ist bereits verbucht und kann "
                "nicht mehr geändert werden."
            )
    return True, None


@bp.route("/booking-groups/<int:group_id>/edit", methods=["GET", "POST"])
@login_required
def booking_group_edit(group_id):
    """Bearbeiten / Anzeigen einer Sammelbuchung (ADR-002).

    Bearbeitung ist erlaubt, solange die Gruppe aktiv und noch nicht verbucht
    ist. Ist das nicht der Fall, wird dasselbe Formular im Readonly-Modus
    angezeigt – die reine Ansichts-Route ``booking_group_detail`` wurde
    abgeschafft und leitet hierher weiter.

    Beim Speichern werden *alle* aktiven Kinder ersetzt (delete + create),
    damit Zeilen frei hinzugefügt/entfernt werden können. Die Kopplung an
    Rechnung/OpenItem/Kunde bleibt erhalten.
    """
    group = db.get_or_404(BookingGroup, group_id)
    ok, msg = _group_is_editable(group)
    fy_locked = _locked_fiscal_year(group.date)

    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    active_projects = Project.query.filter_by(closed=False).order_by(Project.name).all()
    real_accounts = RealAccount.query.filter_by(active=True).order_by(RealAccount.name).all()
    customers = Customer.query.filter_by(active=True).order_by(Customer.name).all()
    tax_rates = tax_service.tax_rates()
    default_real_account = RealAccount.query.filter_by(is_default=True, active=True).first()
    today = date.today()

    def _render_edit(**extra):
        return render_template(
            "accounting/booking_group_form.html",
            group=group,
            accounts=accounts,
            projects=active_projects,
            real_accounts=real_accounts,
            customers=customers,
            tax_rates=tax_rates,
            default_real_account=default_real_account,
            today=today,
            is_edit=True,
            readonly=not ok,
            readonly_reason=msg,
            fy_locked=fy_locked,
            **extra,
        )

    # Bei nicht bearbeitbaren Gruppen (storniert / verbucht / FY gesperrt)
    # wird das Formular readonly gerendert – POST-Anfragen werden abgewiesen.
    if not ok:
        if request.method == "POST":
            flash(msg, "warning")
            return redirect(url_for("accounting.booking_group_edit", group_id=group.id))
        return _render_edit()

    if request.method == "POST":
        try:
            group_date = date.fromisoformat(request.form["date"])
        except (KeyError, ValueError):
            flash("Ungültiges Datum.", "danger")
            return _render_edit(form_data=request.form)
        if group_date > today:
            flash("Das Datum darf nicht in der Zukunft liegen.", "danger")
            return _render_edit(form_data=request.form)
        fy_error = acc_svc.open_fiscal_year_error(group_date)
        if fy_error:
            flash(f"{fy_error} Sammelbuchung wurde nicht gespeichert.", "danger")
            return _render_edit(form_data=request.form)

        description = request.form.get("description", "").strip()
        if not description:
            flash("Bitte eine Beschreibung angeben.", "danger")
            return _render_edit(form_data=request.form)

        reference = request.form.get("reference", "").strip() or None
        customer_id_raw = request.form.get("customer_id") or None
        invoice_id_raw = request.form.get("invoice_id") or None
        real_account_id_raw = request.form.get("real_account_id") or None

        customer_id = int(customer_id_raw) if customer_id_raw else None
        invoice_id = int(invoice_id_raw) if invoice_id_raw else None
        real_account_id = int(real_account_id_raw) if real_account_id_raw else None

        # Kind-Zeilen einsammeln
        child_accounts = request.form.getlist("child_account_id[]")
        child_projects = request.form.getlist("child_project_id[]")
        child_tax_rates = request.form.getlist("child_tax_rate[]")
        child_amounts = request.form.getlist("child_amount[]")
        child_descriptions = request.form.getlist("child_description[]")

        parsed_children = []
        for i, acc_raw in enumerate(child_accounts):
            if not acc_raw:
                continue
            try:
                acc_id = int(acc_raw)
            except ValueError:
                continue
            amount_raw = child_amounts[i].replace(",", ".") if i < len(child_amounts) else "0"
            try:
                amount = Decimal(amount_raw)
            except Exception:
                continue
            if amount == 0:
                continue
            proj_raw = child_projects[i] if i < len(child_projects) else ""
            try:
                proj_id = int(proj_raw) if proj_raw else None
            except ValueError:
                proj_id = None
            tax_raw = child_tax_rates[i] if i < len(child_tax_rates) else ""
            try:
                tax_rate = Decimal(tax_raw) if tax_raw else Decimal("0")
            except Exception:
                tax_rate = Decimal("0")
            desc = (
                child_descriptions[i].strip()
                if i < len(child_descriptions) and child_descriptions[i].strip()
                else description
            )
            parsed_children.append({
                "account_id": acc_id,
                "project_id": proj_id,
                "tax_rate": tax_rate,
                "amount": amount,
                "description": desc,
            })

        if len(parsed_children) < 2:
            flash(
                "Eine Sammelbuchung benötigt mindestens 2 Zeilen. "
                "Für Einzelbuchungen bitte 'Neue Buchung' verwenden.",
                "danger",
            )
            return _render_edit(form_data=request.form)

        # OpenItem-Referenz aus bestehenden Kindern übernehmen (kann nicht
        # über das Form geändert werden — bleibt am Offenen Posten gebunden).
        existing_open_item_ids = {c.open_item_id for c in group.children if c.open_item_id}
        open_item_id = next(iter(existing_open_item_ids), None) if len(existing_open_item_ids) == 1 else None

        try:
            # Header aktualisieren
            group.date = group_date
            group.description = description
            group.reference = reference
            group.invoice_id = invoice_id
            group.customer_id = customer_id

            # Alte Kinder komplett ersetzen
            for old in list(group.children):
                db.session.delete(old)
            db.session.flush()

            for c in parsed_children:
                child = Booking(
                    date=group_date,
                    account_id=c["account_id"],
                    project_id=c["project_id"],
                    amount=c["amount"],
                    description=c["description"],
                    reference=reference,
                    invoice_id=invoice_id,
                    open_item_id=open_item_id,
                    customer_id=customer_id,
                    real_account_id=real_account_id,
                    tax_rate=c["tax_rate"] if c["tax_rate"] and c["tax_rate"] > 0 else None,
                    group_id=group.id,
                    created_by_id=current_user.id,
                )
                db.session.add(child)

            db.session.flush()
            acc_svc.recompute_group_total(group.id)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(
                f"Fehler beim Aktualisieren der Sammelbuchung – alle Änderungen "
                f"wurden zurückgesetzt: {e}",
                "danger",
            )
            return _render_edit(form_data=request.form)

        flash("Sammelbuchung aktualisiert.", "success")
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))

    return _render_edit()


@bp.route("/booking-groups/<int:group_id>/delete", methods=["POST"])
@login_required
def booking_group_delete(group_id):
    """Löscht eine Sammelbuchung inklusive aller Kinder (ADR-002).

    Erlaubt solange die Gruppe aktiv und keine Buchungszeile bereits verbucht
    wurde — analog zur Einzelbuchung (``booking_delete``).
    """
    group = db.get_or_404(BookingGroup, group_id)
    ok, msg = _group_is_editable(group)
    if not ok:
        flash(msg, "warning")
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))
    try:
        for child in list(group.children):
            db.session.delete(child)
        db.session.delete(group)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Löschen der Sammelbuchung: {e}", "danger")
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))
    flash("Sammelbuchung gelöscht.", "info")
    return redirect(url_for("accounting.bookings"))


@bp.route("/booking-groups/<int:group_id>/stornieren", methods=["GET", "POST"])
@login_required
def booking_group_stornieren(group_id):
    group = db.get_or_404(BookingGroup, group_id)

    if group.status == BookingGroup.STATUS_STORNIERT:
        flash("Diese Sammelbuchung ist bereits storniert.", "warning")
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))

    fy_locked = _locked_fiscal_year(group.date)
    if fy_locked:
        flash(
            f"Das Buchungsjahr {fy_locked.year} ist abgeschlossen. "
            f"Diese Sammelbuchung kann nicht storniert werden.",
            "danger",
        )
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))
    if group.date.year != date.today().year:
        flash("Sammelbuchungen aus Vorjahren können nicht storniert werden.", "warning")
        return redirect(url_for("accounting.booking_group_edit", group_id=group.id))

    if request.method == "POST":
        reason = request.form.get("storno_reason", "").strip()
        if not reason:
            flash("Bitte einen Storno-Grund angeben.", "danger")
            return render_template("accounting/booking_group_storno_form.html", group=group)

        try:
            acc_svc.storno_booking_group(group, reason, current_user.id)

            # Verknüpfte Rechnung analog zur Einzel-Storno-Kaskade behandeln.
            cancelled_invoice_number = None
            if group.invoice_id:
                inv = db.session.get(Invoice, group.invoice_id)
                if inv and inv.status != Invoice.STATUS_CANCELLED:
                    inv.status = Invoice.STATUS_CANCELLED
                    cancelled_invoice_number = inv.invoice_number

            # Verknüpften Offenen Posten (über die Kinder) ggf. wieder öffnen.
            if request.form.get("close_open_item"):
                open_item_ids = {c.open_item_id for c in group.children if c.open_item_id}
                for oid in open_item_ids:
                    oi = db.session.get(OpenItem, oid)
                    if oi:
                        oi.status = OpenItem.STATUS_OPEN

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Stornieren – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("accounting.booking_group_edit", group_id=group.id))

        if cancelled_invoice_number:
            flash(
                f"Rechnung {cancelled_invoice_number} wurde storniert. "
                f"Bitte eine neue Rechnung ausstellen.",
                "warning",
            )
        flash("Sammelbuchung erfolgreich storniert.", "success")
        return redirect(url_for("accounting.bookings"))

    return render_template("accounting/booking_group_storno_form.html", group=group)


# ---------------------------------------------------------------------------
# Offene Posten
# ---------------------------------------------------------------------------

@bp.route("/open-items")
@login_required
def open_items():
    # Status-Auswahl: "open" (offen+teilbezahlt) | "closed" (bezahlt+gutschrift) | "all".
    status_filter = request.args.get("status", "").strip().lower()
    if status_filter not in ("open", "closed", "all"):
        # Rückwärtskompatibel: alte Links/Bookmarks mit ?show_closed=1 zeigten alle Posten.
        status_filter = "all" if request.args.get("show_closed") == "1" else "open"
    amount_min_raw = request.args.get("amount_min", "").strip()
    amount_max_raw = request.args.get("amount_max", "").strip()
    customer_q = request.args.get("customer", "").strip()
    ref_q = request.args.get("ref", "").strip()  # Rechnungsnr. oder Beschreibung
    period_q = request.args.get("period", "").strip()  # BillingPeriod-ID

    item_q = OpenItem.query.join(Customer, OpenItem.customer_id == Customer.id)

    status_groups = {
        "open": [OpenItem.STATUS_OPEN, OpenItem.STATUS_PARTIAL],
        "closed": [OpenItem.STATUS_PAID, OpenItem.STATUS_CREDIT],
    }
    if status_filter in status_groups:
        item_q = item_q.filter(OpenItem.status.in_(status_groups[status_filter]))
    # "all" → kein Status-Filter

    if customer_q:
        item_q = item_q.filter(Customer.name.ilike(f"%{customer_q}%"))
    if ref_q:
        item_q = item_q.filter(OpenItem.description.ilike(f"%{ref_q}%"))
    if period_q:
        # Abrechnungsperiode wird ueber die verknuepfte Rechnung ermittelt —
        # OpenItems ohne Rechnung haben keine Periode und fallen hier raus.
        try:
            item_q = item_q.join(Invoice, OpenItem.invoice_id == Invoice.id).filter(
                Invoice.billing_period_id == int(period_q)
            )
        except ValueError:
            pass
    if amount_min_raw:
        try:
            item_q = item_q.filter(OpenItem.amount >= Decimal(amount_min_raw.replace(",", ".")))
        except Exception:
            pass
    if amount_max_raw:
        try:
            item_q = item_q.filter(OpenItem.amount <= Decimal(amount_max_raw.replace(",", ".")))
        except Exception:
            pass

    # Total ueber ALLE gefilterten Posten — unabhaengig von der Pagination.
    all_items = item_q.order_by(OpenItem.due_date).all()
    total_open = sum(item.open_balance for item in all_items)

    pagination = paginate_list(all_items, page_key="open_items")
    accounts = Account.query.filter_by(active=True).order_by(Account.name).all()
    billing_periods = BillingPeriod.query.order_by(
        BillingPeriod.start_date.desc(), BillingPeriod.id.desc()
    ).all()

    return render_template(
        "accounting/open_items.html",
        items=pagination.items,
        total_open=total_open,
        today=date.today(),
        status_filter=status_filter,
        f_customer=customer_q,
        f_ref=ref_q,
        f_period=period_q,
        f_amount_min=amount_min_raw,
        f_amount_max=amount_max_raw,
        pagination=pagination,
        accounts=accounts,
        billing_periods=billing_periods,
    )


@bp.route("/open-items/set-account", methods=["POST"])
@login_required
def open_items_set_account():
    """Setzt das Buchungskonto auf allen noch nicht abgeschlossenen Posten."""
    account_id_raw = request.form.get("account_id") or None
    if not account_id_raw:
        flash("Bitte ein Buchungskonto wählen.", "warning")
        return redirect(url_for("accounting.open_items"))
    account_id = int(account_id_raw)
    updated = OpenItem.query.filter(
        OpenItem.status.in_([OpenItem.STATUS_OPEN, OpenItem.STATUS_PARTIAL])
    ).update({OpenItem.account_id: account_id}, synchronize_session=False)
    db.session.commit()
    flash(f"Buchungskonto für {updated} offene(n) Posten gesetzt.", "success")
    return redirect(url_for("accounting.open_items"))


@bp.route("/open-items/new", methods=["GET", "POST"])
@login_required
def open_item_new():
    customers = Customer.query.filter_by(active=True).order_by(Customer.name).all()
    if request.method == "POST":
        from decimal import Decimal
        item_date = date.fromisoformat(request.form["date"])
        fy_error = acc_svc.open_fiscal_year_error(item_date)
        if fy_error:
            flash(f"{fy_error} Offener Posten wurde nicht gespeichert.", "danger")
            return render_template(
                "accounting/open_item_form.html",
                item=None, customers=customers, today=date.today(),
                form_data=request.form,
            )
        amount_raw = request.form.get("amount", "0").replace(",", ".")
        item = OpenItem(
            customer_id=int(request.form["customer_id"]),
            description=request.form["description"].strip(),
            notes=request.form.get("notes", "").strip(),
            amount=Decimal(amount_raw),
            date=item_date,
            due_date=date.fromisoformat(request.form["due_date"]) if request.form.get("due_date") else None,
            status=OpenItem.STATUS_OPEN,
            created_by_id=current_user.id,
        )
        db.session.add(item)
        db.session.commit()
        flash("Offener Posten angelegt.", "success")
        return redirect(url_for("accounting.open_items"))
    return render_template("accounting/open_item_form.html", item=None, customers=customers, today=date.today())


@bp.route("/open-items/<int:item_id>/pay", methods=["POST"])
@login_required
def open_item_pay(item_id):
    """Zahlung (Teil- oder Vollzahlung) auf einen offenen Posten buchen.

    Ist der Posten mit einer Rechnung verknüpft, läuft die Buchungserzeugung
    über ``booking_group_from_invoice_payment`` (ADR-002) — bei mehreren
    Dimensionen (Konto/Projekt/Steuersatz) wird dabei eine Sammelbuchung
    angelegt, sonst fällt sie auf eine einzelne Buchung zurück.
    """
    item = db.get_or_404(OpenItem, item_id)
    from decimal import Decimal
    amount_raw = request.form.get("amount", "0").replace(",", ".")
    form_account_id_raw = request.form.get("account_id") or None
    form_account_id = int(form_account_id_raw) if form_account_id_raw else None
    try:
        amount = Decimal(amount_raw)
    except Exception:
        flash("Ungültiger Betrag.", "danger")
        return redirect(url_for("accounting.open_items"))
    if amount <= 0:
        flash("Betrag muss positiv sein.", "danger")
        return redirect(url_for("accounting.open_items"))

    # Standard-Bankkonto verwenden
    default_ra = RealAccount.query.filter_by(is_default=True, active=True).first() \
        or RealAccount.query.filter_by(active=True).first()
    real_account_id = default_ra.id if default_ra else None

    try:
        payment_date = date.today()

        if item.invoice_id:
            # Rechnungs-Zahlung → Sammelbuchung oder Einzelbuchung via Service.
            invoice = db.session.get(Invoice, item.invoice_id)
            if invoice is None:
                flash("Verknüpfte Rechnung nicht gefunden.", "danger")
                return redirect(url_for("accounting.open_items"))
            if form_account_id and item.account_id != form_account_id:
                item.account_id = form_account_id
            group, children = acc_svc.booking_group_from_invoice_payment(
                invoice=invoice,
                amount=amount,
                payment_date=payment_date,
                real_account_id=real_account_id,
                created_by_id=current_user.id,
                open_item=item,
                reference=invoice.invoice_number,
                fallback_account_id=form_account_id,
            )
        else:
            # Manueller OpenItem ohne Rechnung → einfache Einzelbuchung.
            account_id_to_use = form_account_id or item.account_id
            if account_id_to_use:
                acc = Account.query.get(account_id_to_use)
            else:
                acc = Account.query.filter_by(active=True).first()
            if form_account_id and item.account_id != form_account_id:
                item.account_id = form_account_id
            if not acc:
                flash("Kein aktives Konto gefunden.", "danger")
                return redirect(url_for("accounting.open_items"))

            booking = Booking(
                date=payment_date,
                account_id=acc.id,
                amount=amount,
                description=f"Zahlung – {item.description} – {item.customer.name}",
                reference=f"OP-{item.id}",
                open_item_id=item.id,
                real_account_id=real_account_id,
                customer_id=item.customer_id,
                created_by_id=current_user.id,
            )
            db.session.add(booking)
            db.session.flush()

        paid_total = db.session.query(func.sum(Booking.amount)).filter(
            Booking.open_item_id == item.id
        ).scalar() or Decimal("0")
        balance = Decimal(str(item.amount)) - Decimal(str(paid_total))

        if balance > Decimal("0"):
            item.status = OpenItem.STATUS_PARTIAL
        elif balance == Decimal("0"):
            item.status = OpenItem.STATUS_PAID
        else:
            item.status = OpenItem.STATUS_CREDIT

        # Verknüpfte Rechnung synchronisieren
        if item.invoice_id:
            inv = db.session.get(Invoice, item.invoice_id)
            if inv:
                if balance > Decimal("0"):
                    inv.status = Invoice.STATUS_SENT
                elif balance == Decimal("0"):
                    inv.status = Invoice.STATUS_PAID
                else:
                    inv.status = Invoice.STATUS_CREDIT

        db.session.commit()
    except ValueError as ve:
        db.session.rollback()
        flash(f"Fehler bei der Zahlung: {ve}", "danger")
        return redirect(url_for("accounting.open_items"))
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler bei der Zahlung – alle Änderungen wurden zurückgesetzt: {e}", "danger")
        return redirect(url_for("accounting.open_items"))

    if balance > Decimal("0"):
        flash(f"Teilzahlung von {amount:.2f} \u20ac gebucht. Offener Restbetrag: {balance:.2f} \u20ac", "success")
    elif balance == Decimal("0"):
        flash("Offener Posten vollst\u00e4ndig bezahlt.", "success")
    else:
        flash(f"\u00dcberzahlung von {abs(balance):.2f} \u20ac. Offener Posten als Gutschrift markiert.", "info")
    return redirect(url_for("accounting.open_items"))


@bp.route("/open-items/<int:item_id>/invoice", methods=["GET", "POST"])
@login_required
def open_item_invoice(item_id):
    """Rechnung aus einem manuellen offenen Posten generieren."""
    item = db.get_or_404(OpenItem, item_id)
    tariffs = WaterTariff.query.order_by(WaterTariff.valid_from.desc()).all()
    editor_projects = Project.query.filter_by(closed=False).order_by(Project.name).all()

    if request.method == "POST":
        from app.models import Invoice
        from app.invoices.routes import _apply_row_items_to_invoice

        inv_date = date.fromisoformat(request.form["date"])
        fy_error = acc_svc.open_fiscal_year_error(inv_date)
        if fy_error:
            flash(f"{fy_error} Rechnung wurde nicht erstellt.", "danger")
            return redirect(url_for("accounting.open_item_invoice", item_id=item_id))
        is_vat_liable_year = acc_svc.is_year_vat_liable(inv_date.year)
        due_date = date.fromisoformat(request.form["due_date"]) if request.form.get("due_date") else None
        notes = request.form.get("notes", "").strip()

        inv = Invoice(
            invoice_number=next_invoice_number(inv_date.year),
            customer_id=item.customer_id,
            date=inv_date,
            due_date=due_date,
            status=Invoice.STATUS_DRAFT,
            notes=notes,
            created_by_id=current_user.id,
        )
        db.session.add(inv)
        db.session.flush()

        _apply_row_items_to_invoice(inv, request.form, is_vat_liable_year)

        inv.recalculate_total()
        item.invoice_id = inv.id
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Erstellen der Rechnung – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("accounting.open_items"))
        flash(f"Rechnung {inv.invoice_number} erstellt.", "success")
        return redirect(url_for("invoices.detail", invoice_id=inv.id))

    return render_template(
        "accounting/open_item_invoice.html",
        item=item,
        tariffs=tariffs,
        today=date.today(),
        editor_projects=editor_projects,
    )


# ---------------------------------------------------------------------------
# Jahresabschluss / Berichte
# ---------------------------------------------------------------------------

@bp.route("/report")
@login_required
def report():
    year = request.args.get("year", date.today().year, type=int)
    real_account_id = request.args.get("real_account_id", 0, type=int)  # 0 = alle

    all_real_accounts = RealAccount.query.order_by(RealAccount.name).all()

    income_rows, expense_rows, total_income, total_expense, balance = (
        acc_svc.year_income_expense(year, real_account_id=real_account_id or None)
    )

    project_summary_list = acc_svc.year_project_summary(
        year, real_account_id=real_account_id or None
    )

    # Kontenentwicklung Bankkonten
    konten_list = []
    for ra in all_real_accounts:
        jan1 = acc_svc.jan1_balance(ra, year)
        dec31 = acc_svc.year_end_balance(ra, year)
        bewegung = acc_svc.year_booking_total(ra.id, year)
        konten_list.append({
            "id": ra.id,
            "name": ra.name,
            "iban": ra.iban or "",
            "jan1": jan1,
            "bewegung": Decimal(str(bewegung)),
            "dec31": dec31,
        })

    return render_template(
        "accounting/report.html",
        year=year,
        real_account_id=real_account_id,
        all_real_accounts=all_real_accounts,
        income_rows=income_rows,
        expense_rows=expense_rows,
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        project_summary_list=project_summary_list,
        konten_list=konten_list,
        fy_vat_liable=acc_svc.is_year_vat_liable(year),
    )


@bp.route("/report/export/excel")
@login_required
def report_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    year = request.args.get("year", date.today().year, type=int)
    fy_vat_liable = acc_svc.is_year_vat_liable(year)

    # ---- Hilfsfunktionen ----
    HDR_FILL = PatternFill("solid", fgColor="2F5496")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
    SUBHDR_FONT = Font(bold=True)
    TOTAL_FONT = Font(bold=True)
    EUR_FMT = '#,##0.00 "€"'
    DATE_FMT = "DD.MM.YYYY"
    thin = Side(style="thin")
    BORDER = Border(bottom=thin)

    def _hdr(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

    def _subhdr(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = SUBHDR_FONT
            cell.fill = SUBHDR_FILL

    def _autowidth(ws, min_w=10, max_w=50):
        for col in ws.columns:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, length + 2))

    def _eur(ws, row, col, val):
        cell = ws.cell(row=row, column=col, value=float(val))
        cell.number_format = EUR_FMT
        return cell

    def _total_row(ws, row, label, val, ncols):
        ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
        c = _eur(ws, row, ncols, val)
        c.font = TOTAL_FONT
        c.border = BORDER

    # ---- Daten ermitteln ----
    all_real_accounts = RealAccount.query.order_by(RealAccount.name).all()

    bookings_year = acc_svc.year_bookings(year)

    income_rows, expense_rows, total_income, total_expense, balance = (
        acc_svc.year_income_expense(year)
    )

    # Kontenentwicklung
    konten_list = []
    for ra in all_real_accounts:
        jan1 = acc_svc.jan1_balance(ra, year)
        dec31 = acc_svc.year_end_balance(ra, year)
        bewegung = acc_svc.year_booking_total(ra.id, year)
        konten_list.append({
            "name": ra.name, "iban": ra.iban or "",
            "jan1": jan1, "bewegung": Decimal(str(bewegung)), "dec31": dec31,
        })

    # Projektübersicht
    project_list = acc_svc.year_project_summary(year)

    # ---- Workbook aufbauen ----
    wb = openpyxl.Workbook()

    # ================================================================
    # BLATT 1: Übersicht
    # ================================================================
    ws = wb.active
    ws.title = "Übersicht"
    r = 1
    ws.merge_cells(f"A{r}:D{r}")
    cell = ws.cell(row=r, column=1, value=f"Jahresbericht {year}")
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    r += 2

    _subhdr(ws, r, ["Kennzahl", "", "Betrag"])
    r += 1
    _eur(ws, r, 3, total_income); ws.cell(row=r, column=1, value="Gesamteinnahmen"); r += 1
    _eur(ws, r, 3, total_expense); ws.cell(row=r, column=1, value="Gesamtausgaben"); r += 1
    c = _eur(ws, r, 3, balance)
    c.font = Font(bold=True, color="375623" if balance >= 0 else "9C0006")
    ws.cell(row=r, column=1, value="Saldo").font = TOTAL_FONT
    r += 2

    _subhdr(ws, r, ["Bankkonto", "IBAN", "Stand 1.1.", "Stand 31.12."])
    r += 1
    total_jan1 = Decimal("0"); total_dec31 = Decimal("0")
    for k in konten_list:
        ws.cell(row=r, column=1, value=k["name"])
        ws.cell(row=r, column=2, value=k["iban"])
        _eur(ws, r, 3, k["jan1"]); _eur(ws, r, 4, k["dec31"])
        total_jan1 += k["jan1"]; total_dec31 += k["dec31"]
        r += 1
    ws.cell(row=r, column=1, value="Gesamt").font = TOTAL_FONT
    c3 = _eur(ws, r, 3, total_jan1); c3.font = TOTAL_FONT; c3.border = BORDER
    c4 = _eur(ws, r, 4, total_dec31); c4.font = TOTAL_FONT; c4.border = BORDER
    r += 2

    # USt-Zahllast Jahresübersicht (nur wenn Jahr umsatzsteuerpflichtig ist)
    if fy_vat_liable:
        ust_rows_y, vst_rows_y = _ust_berechnen(year, 0)
        zahllast_y = sum(v["steuer"] for _, v in ust_rows_y) - sum(v["steuer"] for _, v in vst_rows_y)
        _subhdr(ws, r, ["USt/VSt", "", "Betrag"])
        r += 1
        _eur(ws, r, 3, sum(v["steuer"] for _, v in ust_rows_y)); ws.cell(row=r, column=1, value="Umsatzsteuer gesamt"); r += 1
        _eur(ws, r, 3, sum(v["steuer"] for _, v in vst_rows_y)); ws.cell(row=r, column=1, value="Vorsteuer gesamt"); r += 1
        c = _eur(ws, r, 3, zahllast_y); c.font = TOTAL_FONT
        ws.cell(row=r, column=1, value="Zahllast").font = TOTAL_FONT
    else:
        ust_rows_y, vst_rows_y = [], []
    _autowidth(ws)

    # ================================================================
    # BLATT 2: Einnahmen & Ausgaben
    # ================================================================
    ws2 = wb.create_sheet("Einnahmen & Ausgaben")
    _hdr(ws2, 1, ["Konto", "Typ", "Betrag"])
    r2 = 2
    for _acc_id, name, amt in income_rows:
        ws2.cell(row=r2, column=1, value=name)
        ws2.cell(row=r2, column=2, value="Einnahme")
        _eur(ws2, r2, 3, amt); r2 += 1
    _total_row(ws2, r2, "Einnahmen gesamt", total_income, 3); r2 += 2
    for _acc_id, name, amt in expense_rows:
        ws2.cell(row=r2, column=1, value=name)
        ws2.cell(row=r2, column=2, value="Ausgabe")
        _eur(ws2, r2, 3, amt); r2 += 1
    _total_row(ws2, r2, "Ausgaben gesamt", total_expense, 3); r2 += 2
    ws2.cell(row=r2, column=1, value="Saldo").font = TOTAL_FONT
    c = _eur(ws2, r2, 3, balance); c.font = TOTAL_FONT; c.border = BORDER
    _autowidth(ws2)

    # ================================================================
    # BLATT 3: Buchungen
    # ================================================================
    ws3 = wb.create_sheet("Buchungen")
    if fy_vat_liable:
        _hdr(ws3, 1, ["Datum", "Bankkonto", "Konto", "Typ", "Beschreibung", "Belegnummer", "Projekt", "Kunde", "USt %", "USt Betrag", "Betrag", "Status", "Sammel-ID"])
    else:
        _hdr(ws3, 1, ["Datum", "Bankkonto", "Konto", "Typ", "Beschreibung", "Belegnummer", "Projekt", "Kunde", "Betrag", "Status", "Sammel-ID"])
    r3 = 2
    for b in bookings_year:
        ws3.cell(row=r3, column=1, value=b.date).number_format = DATE_FMT
        ws3.cell(row=r3, column=2, value=b.real_account.name if b.real_account else "")
        ws3.cell(row=r3, column=3, value=b.account.name)
        ws3.cell(row=r3, column=4, value="Einnahme" if b.amount >= 0 else "Ausgabe")
        ws3.cell(row=r3, column=5, value=b.description)
        ws3.cell(row=r3, column=6, value=b.reference or "")
        ws3.cell(row=r3, column=7, value=b.project.name if b.project else "")
        ws3.cell(row=r3, column=8, value=b.customer.name if b.customer else "")
        # Sammel-ID: zeigt die BookingGroup-ID an — macht in Excel-Pivot
        # die Zusammengehörigkeit von Split-Buchungen nachvollziehbar (ADR-002).
        group_label = f"#{b.group_id}" if b.group_id else ""
        if fy_vat_liable:
            tax_amt = ""
            if b.tax_rate and b.tax_rate > 0:
                tax_amt = float((abs(b.amount) * Decimal(str(b.tax_rate)) / (100 + Decimal(str(b.tax_rate)))).quantize(Decimal("0.01")))
            ws3.cell(row=r3, column=9, value=int(b.tax_rate) if b.tax_rate else "")
            if tax_amt != "":
                _eur(ws3, r3, 10, tax_amt)
            _eur(ws3, r3, 11, float(b.amount))
            ws3.cell(row=r3, column=12, value=b.status or "")
            ws3.cell(row=r3, column=13, value=group_label)
        else:
            _eur(ws3, r3, 9, float(b.amount))
            ws3.cell(row=r3, column=10, value=b.status or "")
            ws3.cell(row=r3, column=11, value=group_label)
        r3 += 1
    _autowidth(ws3)

    # ================================================================
    # BLATT 4: Bankkonten-Entwicklung
    # ================================================================
    ws4 = wb.create_sheet("Bankkonten")
    _hdr(ws4, 1, ["Bankkonto", "IBAN", f"Stand 1.1.{year}", f"Bewegung {year}", f"Stand 31.12.{year}"])
    r4 = 2
    sum_jan1 = Decimal("0"); sum_dec31 = Decimal("0")
    for k in konten_list:
        ws4.cell(row=r4, column=1, value=k["name"])
        ws4.cell(row=r4, column=2, value=k["iban"])
        _eur(ws4, r4, 3, k["jan1"])
        _eur(ws4, r4, 4, k["bewegung"])
        _eur(ws4, r4, 5, k["dec31"])
        sum_jan1 += k["jan1"]; sum_dec31 += k["dec31"]
        r4 += 1
    ws4.cell(row=r4, column=1, value="Gesamt").font = TOTAL_FONT
    for col, val in [(3, sum_jan1), (4, sum_dec31 - sum_jan1), (5, sum_dec31)]:
        c = _eur(ws4, r4, col, val); c.font = TOTAL_FONT; c.border = BORDER
    _autowidth(ws4)

    # ================================================================
    # BLATT 5: Projekte
    # ================================================================
    ws5 = wb.create_sheet("Projekte")
    _hdr(ws5, 1, ["Projekt", "Konto", "Einnahmen", "Ausgaben", "Saldo"])
    r5 = 2
    for ps in project_list:
        start_r = r5
        for _acc_id, acc_name, inc_amt, exp_amt in ps["accounts"]:
            ws5.cell(row=r5, column=1, value=ps["name"])
            ws5.cell(row=r5, column=2, value=acc_name)
            if inc_amt > 0:
                _eur(ws5, r5, 3, inc_amt)
            if exp_amt > 0:
                _eur(ws5, r5, 4, exp_amt)
            r5 += 1
        # Projektsumme
        ws5.cell(row=r5, column=1, value=ps["name"]).font = TOTAL_FONT
        ws5.cell(row=r5, column=2, value="Gesamt").font = TOTAL_FONT
        c3 = _eur(ws5, r5, 3, ps["income"]); c3.font = TOTAL_FONT; c3.border = BORDER
        c4 = _eur(ws5, r5, 4, ps["expense"]); c4.font = TOTAL_FONT; c4.border = BORDER
        c5 = _eur(ws5, r5, 5, ps["income"] - ps["expense"]); c5.font = TOTAL_FONT; c5.border = BORDER
        r5 += 2
    _autowidth(ws5)

    # ================================================================
    # BLÄTTER 6-10: USt-Voranmeldungen Q1–Q4 + Gesamtjahr
    # ================================================================
    def _ust_sheet(ws_ust, label, date_from, date_to, ust_r, vst_r):
        ws_ust.cell(row=1, column=1, value="Zeitraum").font = SUBHDR_FONT
        ws_ust.cell(row=1, column=2, value=label)
        ws_ust.cell(row=2, column=1, value="Von").font = SUBHDR_FONT
        ws_ust.cell(row=2, column=2, value=date_from).number_format = DATE_FMT
        ws_ust.cell(row=3, column=1, value="Bis").font = SUBHDR_FONT
        ws_ust.cell(row=3, column=2, value=date_to).number_format = DATE_FMT

        _hdr(ws_ust, 5, ["Abschnitt", "Steuersatz %", "Bruttobetrag", "Nettobetrag", "Steuerbetrag"])
        ru = 6
        for rate, v in ust_r:
            ws_ust.cell(row=ru, column=1, value="Umsatzsteuer")
            ws_ust.cell(row=ru, column=2, value=rate)
            _eur(ws_ust, ru, 3, v["brutto"]); _eur(ws_ust, ru, 4, v["netto"]); _eur(ws_ust, ru, 5, v["steuer"])
            ru += 1
        for rate, v in vst_r:
            ws_ust.cell(row=ru, column=1, value="Vorsteuer")
            ws_ust.cell(row=ru, column=2, value=rate)
            _eur(ws_ust, ru, 3, v["brutto"]); _eur(ws_ust, ru, 4, v["netto"]); _eur(ws_ust, ru, 5, v["steuer"])
            ru += 1
        ru += 1
        total_u = sum(v["steuer"] for _, v in ust_r)
        total_v = sum(v["steuer"] for _, v in vst_r)
        zahllast = total_u - total_v
        for lbl, val in [("Umsatzsteuer gesamt", total_u), ("Vorsteuer gesamt", total_v), ("Zahllast", zahllast)]:
            ws_ust.cell(row=ru, column=1, value=lbl).font = TOTAL_FONT
            c = _eur(ws_ust, ru, 5, val); c.font = TOTAL_FONT; c.border = BORDER
            ru += 1
        _autowidth(ws_ust)

    if fy_vat_liable:
        for q in range(1, 5):
            df, dt = _ust_period(year, q)
            ur, vr = _ust_berechnen(year, q)
            ws_q = wb.create_sheet(f"USt Q{q}")
            _ust_sheet(ws_q, f"Q{q}/{year}", df, dt, ur, vr)

        df_y, dt_y = _ust_period(year, 0)
        ws_y = wb.create_sheet("USt Gesamtjahr")
        _ust_sheet(ws_y, str(year), df_y, dt_y, ust_rows_y, vst_rows_y)

    # ---- Ausgabe ----
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=jahresbericht_{year}.xlsx"},
    )


# ---------------------------------------------------------------------------
# Buchungen-Import
# ---------------------------------------------------------------------------

def _parse_at_number(raw):
    """Österreichisches Zahlenformat: Leerzeichen/Punkt als Tausender, Komma als Dezimal."""
    raw = str(raw).strip().replace('\xa0', '').replace(' ', '')
    if not raw or raw == 'nan':
        return None
    if ',' in raw and '.' in raw:
        raw = raw.replace('.', '')
    raw = raw.replace(',', '.')
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _ensure_umbuchungs_account(account_cache, kst_name):
    """Stellt sicher, dass ein Account für den Umbuchungs-KST im Cache vorhanden ist."""
    if kst_name not in account_cache:
        acc = Account.query.filter_by(name=kst_name).first()
        if not acc:
            acc = Account(name=kst_name)
            db.session.add(acc)
            db.session.flush()
        account_cache[kst_name] = acc


def _find_customer(name, customers, customer_name_map):
    """Exakte Suche nach Kunde anhand des Namens (Vor- und Nachname müssen übereinstimmen)."""
    if not name or not name.strip():
        return None
    name_lower = name.strip().lower()

    # Exakte Übereinstimmung
    if name_lower in customer_name_map:
        return customer_name_map[name_lower]

    # Umgekehrte Reihenfolge (Nachname Vorname → Vorname Nachname)
    parts = name.strip().split()
    if len(parts) == 2:
        reversed_name = f"{parts[1]} {parts[0]}".lower()
        if reversed_name in customer_name_map:
            return customer_name_map[reversed_name]

    return None


@bp.route("/bookings/import", methods=["GET", "POST"])
@login_required
def import_bookings():
    import pandas as pd

    if request.method == "POST":
        # ------------------------------------------------------------------
        # Stufe 1: Datei hochladen → Spaltenvorschau
        # ------------------------------------------------------------------
        if "file" in request.files and request.files["file"].filename:
            f = request.files["file"]
            try:
                raw_bytes = f.read()
                # Versuche UTF-8-sig, dann latin-1
                for enc in ("utf-8-sig", "latin-1"):
                    try:
                        df = pd.read_csv(
                            io.BytesIO(raw_bytes), sep=";", dtype=str,
                            encoding=enc, keep_default_na=False,
                        )
                        break
                    except Exception:
                        continue
                else:
                    flash("Fehler beim Lesen der Datei.", "danger")
                    return redirect(url_for("accounting.import_bookings"))
            except Exception as e:
                flash(f"Fehler beim Lesen der Datei: {e}", "danger")
                return redirect(url_for("accounting.import_bookings"))

            file_content_b64 = base64.b64encode(raw_bytes).decode("ascii")
            columns = list(df.columns)
            preview = df.head(5).to_dict(orient="records")

            # Spalten automatisch vorauswählen
            def _auto(names):
                for n in names:
                    if n in columns:
                        return n
                return ""

            auto = {
                "datum": _auto(["Datum"]),
                "kst": _auto(["KST"]),
                "ausgaben": _auto(["Ausgaben"]),
                "einnahmen": _auto(["Einnahmen"]),
                "konto": _auto(["Konto"]),
                "ktr": _auto(["KTR"]),
                "name": _auto(["Name"]),
                "beschreibung": _auto(["Beschreibung"]),
                "steuer": _auto(["Steuer"]),
            }
            return render_template(
                "accounting/import_bookings_mapping.html",
                columns=columns,
                preview=preview,
                file_content=file_content_b64,
                auto=auto,
            )

        # ------------------------------------------------------------------
        # Stufe 2: Mapping bestätigen → Buchungen anlegen
        # ------------------------------------------------------------------
        if request.form.get("confirm") == "1":
            file_content_b64 = request.form.get("file_content", "")
            if not file_content_b64:
                flash("Import-Daten fehlen, bitte Datei erneut hochladen.", "danger")
                return redirect(url_for("accounting.import_bookings"))

            try:
                raw_bytes = base64.b64decode(file_content_b64)
                for enc in ("utf-8-sig", "latin-1"):
                    try:
                        df = pd.read_csv(
                            io.BytesIO(raw_bytes), sep=";", dtype=str,
                            encoding=enc, keep_default_na=False,
                        )
                        break
                    except Exception:
                        continue
                else:
                    flash("Fehler beim Lesen der Import-Daten.", "danger")
                    return redirect(url_for("accounting.import_bookings"))
            except Exception as e:
                flash(f"Fehler: {e}", "danger")
                return redirect(url_for("accounting.import_bookings"))

            # Spaltenmapping aus Formular
            col_datum = request.form.get("col_datum", "")
            col_kst = request.form.get("col_kst", "")
            col_ausgaben = request.form.get("col_ausgaben", "")
            col_einnahmen = request.form.get("col_einnahmen", "")
            col_konto = request.form.get("col_konto", "")
            col_ktr = request.form.get("col_ktr", "")
            col_name = request.form.get("col_name", "")
            col_beschreibung = request.form.get("col_beschreibung", "")
            col_steuer = request.form.get("col_steuer", "")
            umbuchungs_kst = request.form.get("umbuchungs_kst", "").strip()

            if not col_datum or not col_kst:
                flash("Pflichtfelder Datum und KST (Konto) müssen zugeordnet sein.", "danger")
                return redirect(url_for("accounting.import_bookings"))

            # Kunden-Cache aufbauen
            alle_kunden = Customer.query.filter_by(active=True).all()
            customer_name_map = {c.name.lower(): c.id for c in alle_kunden}

            # Konto/Projekt/Bankkonto-Caches
            account_cache = {}      # name → Account
            project_cache = {}      # name → Project
            real_account_cache = {} # name → RealAccount

            results = {"ok": 0, "skip": 0, "matched": 0, "transfers": 0, "transfer_warnings": []}

            # Umbuchungs-Kandidaten: Liste von dicts mit geparsten Feldern
            transfer_candidates = []  # {"amount": Decimal, "date": date, "real_account_id": int|None, "description": str}

            for _, row in df.iterrows():
                def _col(c):
                    v = str(row.get(c, "")).strip() if c else ""
                    return v if v and v.lower() != "nan" else ""

                # Betrag bestimmen
                amount = None
                is_ausgabe = False
                ausgaben_raw = _col(col_ausgaben)
                einnahmen_raw = _col(col_einnahmen)

                if ausgaben_raw:
                    amount = _parse_at_number(ausgaben_raw)
                    is_ausgabe = True
                elif einnahmen_raw:
                    amount = _parse_at_number(einnahmen_raw)

                if amount is None:
                    results["skip"] += 1
                    continue

                # Ausgaben-Spalte: Betrag muss negativ sein
                if is_ausgabe and amount > 0:
                    amount = -amount

                # Datum parsen
                datum_raw = _col(col_datum)
                if not datum_raw:
                    results["skip"] += 1
                    continue
                try:
                    if "." in datum_raw:
                        from datetime import datetime as _dt
                        booking_date = _dt.strptime(datum_raw, "%d.%m.%Y").date()
                    else:
                        booking_date = date.fromisoformat(datum_raw)
                except Exception:
                    results["skip"] += 1
                    continue

                # Konto (KST) ermitteln
                kst_name = _col(col_kst)
                if not kst_name:
                    results["skip"] += 1
                    continue

                # Reales Bankkonto ermitteln / anlegen
                real_account_id = None
                konto_name = _col(col_konto)
                if konto_name:
                    if konto_name not in real_account_cache:
                        ra = RealAccount.query.filter_by(name=konto_name).first()
                        if not ra:
                            ra = RealAccount(name=konto_name)
                            db.session.add(ra)
                            db.session.flush()
                        real_account_cache[konto_name] = ra
                    real_account_id = real_account_cache[konto_name].id

                # Umbuchungs-Kandidat: in separate Liste stellen
                if umbuchungs_kst and kst_name == umbuchungs_kst:
                    beschreibung = _col(col_beschreibung)
                    import_name = _col(col_name)
                    parts = [p for p in [import_name, beschreibung] if p]
                    description = " – ".join(parts) if parts else "Umbuchung"
                    transfer_candidates.append({
                        "amount": amount,
                        "date": booking_date,
                        "real_account_id": real_account_id,
                        "description": description[:500],
                    })
                    continue

                if kst_name not in account_cache:
                    acc = Account.query.filter_by(name=kst_name).first()
                    if not acc:
                        acc = Account(name=kst_name)
                        db.session.add(acc)
                        db.session.flush()
                    account_cache[kst_name] = acc
                acc = account_cache[kst_name]

                # Projekt ermitteln / anlegen
                project_id = None
                ktr_name = _col(col_ktr)
                if ktr_name:
                    if ktr_name not in project_cache:
                        proj = Project.query.filter_by(name=ktr_name).first()
                        if not proj:
                            proj = Project(name=ktr_name)
                            db.session.add(proj)
                            db.session.flush()
                        project_cache[ktr_name] = proj
                    project_id = project_cache[ktr_name].id

                # Kunde suchen
                import_name = _col(col_name)
                customer_id = _find_customer(import_name, alle_kunden, customer_name_map)
                if customer_id:
                    results["matched"] += 1

                # Beschreibung
                beschreibung = _col(col_beschreibung)
                if customer_id:
                    description = beschreibung or import_name or "—"
                else:
                    parts = [p for p in [import_name, beschreibung] if p]
                    description = " – ".join(parts) if parts else "—"

                # Steuersatz
                tax_rate = None
                steuer_raw = _col(col_steuer)
                if steuer_raw:
                    try:
                        tr = Decimal(steuer_raw.replace(",", "."))
                        if tr > 0:
                            tax_rate = tr
                    except Exception:
                        pass

                b = Booking(
                    date=booking_date,
                    account_id=acc.id,
                    amount=amount,
                    description=description[:500],
                    real_account_id=real_account_id,
                    project_id=project_id,
                    customer_id=customer_id,
                    tax_rate=tax_rate,
                    created_by_id=current_user.id,
                    status=Booking.STATUS_OFFEN,
                )
                db.session.add(b)
                results["ok"] += 1

            # ------------------------------------------------------------------
            # Umbuchungs-Kandidaten paarweise matchen
            # Regel: Ausgabe (amount < 0) von Konto A + Einnahme (amount > 0) auf Konto B
            # mit gleichem abs(amount) → Transfer
            # ------------------------------------------------------------------
            if transfer_candidates:
                from collections import defaultdict
                # Gruppieren nach abs(amount)
                by_amount = defaultdict(lambda: {"ausgaben": [], "einnahmen": []})
                for tc in transfer_candidates:
                    key = abs(tc["amount"])
                    if tc["amount"] < 0:
                        by_amount[key]["ausgaben"].append(tc)
                    else:
                        by_amount[key]["einnahmen"].append(tc)

                for abs_amt, group in by_amount.items():
                    ausgaben = group["ausgaben"]
                    einnahmen = group["einnahmen"]

                    while ausgaben and einnahmen:
                        aus = ausgaben.pop(0)
                        ein = einnahmen.pop(0)

                        from_ra_id = aus["real_account_id"]
                        to_ra_id = ein["real_account_id"]

                        if from_ra_id is None or to_ra_id is None:
                            results["transfer_warnings"].append(
                                f"Umbuchung {abs_amt:.2f}: Bankkonto fehlt — als normale Buchung importiert."
                            )
                            # Fallback: als normale Buchungen anlegen
                            for tc_fb in [aus, ein]:
                                _ensure_umbuchungs_account(account_cache, umbuchungs_kst)
                                acc_fb = account_cache[umbuchungs_kst]
                                b_fb = Booking(
                                    date=tc_fb["date"],
                                    account_id=acc_fb.id,
                                    amount=tc_fb["amount"],
                                    description=tc_fb["description"],
                                    real_account_id=tc_fb["real_account_id"],
                                    created_by_id=current_user.id,
                                    status=Booking.STATUS_OFFEN,
                                )
                                db.session.add(b_fb)
                                results["ok"] += 1
                            continue

                        if from_ra_id == to_ra_id:
                            results["transfer_warnings"].append(
                                f"Umbuchung {abs_amt:.2f}: Ausgangs- und Zielkonto identisch — übersprungen."
                            )
                            results["skip"] += 2
                            continue

                        t = Transfer(
                            date=aus["date"],
                            amount=abs_amt,
                            description=aus["description"] or ein["description"],
                            from_real_account_id=from_ra_id,
                            to_real_account_id=to_ra_id,
                            created_by_id=current_user.id,
                        )
                        db.session.add(t)
                        results["transfers"] += 1

                    # Nicht gematchte Kandidaten → Warnung + normale Buchung
                    for tc_unmatched in ausgaben + einnahmen:
                        results["transfer_warnings"].append(
                            f"Umbuchung {abs_amt:.2f} ({tc_unmatched['date']}): "
                            f"Keine Gegenbuchung gefunden — als normale Buchung importiert."
                        )
                        _ensure_umbuchungs_account(account_cache, umbuchungs_kst)
                        acc_um = account_cache[umbuchungs_kst]
                        b_um = Booking(
                            date=tc_unmatched["date"],
                            account_id=acc_um.id,
                            amount=tc_unmatched["amount"],
                            description=tc_unmatched["description"],
                            real_account_id=tc_unmatched["real_account_id"],
                            created_by_id=current_user.id,
                            status=Booking.STATUS_OFFEN,
                        )
                        db.session.add(b_um)
                        results["ok"] += 1

            db.session.commit()
            msg = (
                f"Import abgeschlossen: {results['ok']} Buchungen importiert, "
                f"{results['skip']} übersprungen"
            )
            if results["transfers"]:
                msg += f", {results['transfers']} Umbuchungen erstellt"
            if results["matched"]:
                msg += f", {results['matched']} Kunden automatisch zugeordnet"
            msg += "."
            flash(msg, "success" if (results["ok"] or results["transfers"]) else "warning")
            for w in results["transfer_warnings"]:
                flash(w, "warning")
            return redirect(url_for("accounting.bookings"))

    return render_template("accounting/import_bookings.html")


# ---------------------------------------------------------------------------
# CSV-Export
# ---------------------------------------------------------------------------

@bp.route("/bookings/export")
@login_required
def export_csv():
    year = request.args.get("year", date.today().year, type=int)
    # Filter werden nur angewandt, wenn die jeweiligen Parameter mitgegeben
    # werden — der "ganzes Jahr"-Export ruft die Route ohne sie auf.
    real_account_id = request.args.get("real_account_id", "", type=str)
    account_id = request.args.get("account_id", "", type=str)
    project_id = request.args.get("project_id", "", type=str)
    kind = request.args.get("kind", "", type=str)
    month = request.args.get("month", 0, type=int)
    quarter = request.args.get("quarter", 0, type=int)
    tax = request.args.get("tax", "", type=str)

    query = Booking.query.filter(extract("year", Booking.date) == year)
    if real_account_id:
        query = query.filter(Booking.real_account_id == int(real_account_id))
    if account_id:
        query = query.filter(Booking.account_id == int(account_id))
    if project_id:
        query = query.filter(Booking.project_id == int(project_id))
    if month in range(1, 13):
        query = query.filter(extract("month", Booking.date) == month)
    if quarter in range(1, 5):
        query = query.filter(
            extract("month", Booking.date).between(quarter * 3 - 2, quarter * 3)
        )
    if kind == "income":
        query = query.filter(Booking.amount > 0)
    elif kind == "expense":
        query = query.filter(Booking.amount < 0)
    if tax == "any":
        query = query.filter(Booking.tax_rate.isnot(None), Booking.tax_rate > 0)
    elif tax == "0":
        query = query.filter(db.or_(Booking.tax_rate.is_(None), Booking.tax_rate == 0))
    elif tax in {str(int(r)) for r in tax_service.tax_rate_values()}:
        query = query.filter(Booking.tax_rate == Decimal(tax))
    bookings = query.order_by(Booking.date).all()

    def generate():
        output = io.StringIO()
        output.write("\ufeff")  # UTF-8 BOM für korrekte Darstellung in Excel
        writer = csv.writer(output, delimiter=";")
        writer.writerow([
            "Datum", "Bankkonto", "Konto", "Typ", "Beschreibung",
            "Belegnummer", "Projekt", "Kunde", "MwSt %", "MwSt Betrag", "Betrag", "Status",
            "Sammel-ID",
        ])
        for b in bookings:
            tax_amount = ""
            if b.tax_rate and b.tax_rate > 0 and b.status != "Storniert":
                tax_amount = str(round(abs(b.amount) * b.tax_rate / (100 + b.tax_rate), 2)).replace(".", ",")
            writer.writerow([
                b.date.strftime("%d.%m.%Y"),
                b.real_account.name if b.real_account else "",
                b.account.name,
                "Einnahme" if b.amount >= 0 else "Ausgabe",
                b.description,
                b.reference or "",
                b.project.name if b.project else "",
                b.customer.name if b.customer else "",
                str(int(b.tax_rate)).replace(".", ",") if b.tax_rate else "",
                tax_amount,
                str(b.amount).replace(".", ","),
                b.status or "",
                f"#{b.group_id}" if b.group_id else "",
            ])
        return output.getvalue()

    return Response(
        generate(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=buchungen_{year}.csv"},
    )


# ---------------------------------------------------------------------------
# Umsatzsteuervoranmeldung / Umsatzsteuererklärung
# ---------------------------------------------------------------------------
# Berechnungslogik liegt in ``app.accounting.services``. Die Routen rufen den
# Service nur noch zum Aufbereiten der Templates auf.


@bp.route("/ust")
@login_required
def ust():
    year = request.args.get("year", date.today().year, type=int)
    quartal = request.args.get("quartal", 0, type=int)
    # Nur für umsatzsteuerpflichtige Jahre verfügbar
    vat_years = [fy.year for fy in FiscalYear.query.filter_by(is_vat_liable=True)
                 .order_by(FiscalYear.year.desc()).all()]
    if not vat_years:
        flash(
            "Es ist kein umsatzsteuerpflichtiges Buchungsjahr angelegt. "
            "Die Umsatzsteuer-Voranmeldung ist nicht verfügbar.",
            "warning",
        )
        return redirect(url_for("accounting.fiscal_years"))
    if not acc_svc.is_year_vat_liable(year):
        flash(
            f"Das Buchungsjahr {year} ist nicht umsatzsteuerpflichtig. "
            f"Die Umsatzsteuer-Voranmeldung ist nur für pflichtige Jahre verfügbar.",
            "warning",
        )
        year = vat_years[0]
    totals = acc_svc.ust_totals(year, quartal)
    return render_template(
        "accounting/ust.html",
        year=year, quartal=quartal,
        vat_years=vat_years,
        date_from=totals["date_from"], date_to=totals["date_to"],
        ust_rows=totals["ust_rows"], vst_rows=totals["vst_rows"],
        total_ust=totals["total_ust"], total_vst=totals["total_vst"],
        zahllast=totals["zahllast"],
        ust_brutto=totals["ust_brutto"], ust_netto=totals["ust_netto"],
        vst_brutto=totals["vst_brutto"], vst_netto=totals["vst_netto"],
    )


@bp.route("/ust/export")
@login_required
def export_ust_csv():
    year = request.args.get("year", date.today().year, type=int)
    quartal = request.args.get("quartal", 0, type=int)
    if not acc_svc.is_year_vat_liable(year):
        flash(f"Das Buchungsjahr {year} ist nicht umsatzsteuerpflichtig.", "warning")
        return redirect(url_for("accounting.ust"))
    totals = acc_svc.ust_totals(year, quartal)
    date_from = totals["date_from"]
    date_to = totals["date_to"]
    ust_rows = totals["ust_rows"]
    vst_rows = totals["vst_rows"]
    total_ust = totals["total_ust"]
    total_vst = totals["total_vst"]
    zahllast = totals["zahllast"]

    label = f"Q{quartal}/{year}" if quartal else str(year)

    def fmt(d):
        return str(d.quantize(Decimal("0.01"))).replace(".", ",")

    def generate():
        output = io.StringIO()
        output.write("\ufeff")  # UTF-8 BOM
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["Zeitraum", label])
        writer.writerow(["Von", date_from.strftime("%d.%m.%Y")])
        writer.writerow(["Bis", date_to.strftime("%d.%m.%Y")])
        writer.writerow([])
        writer.writerow(["Abschnitt", "Steuersatz %", "Bruttobetrag", "Nettobetrag", "Steuerbetrag"])
        for rate, v in ust_rows:
            writer.writerow(["Umsatzsteuer", rate, fmt(v["brutto"]), fmt(v["netto"]), fmt(v["steuer"])])
        for rate, v in vst_rows:
            writer.writerow(["Vorsteuer", rate, fmt(v["brutto"]), fmt(v["netto"]), fmt(v["steuer"])])
        writer.writerow([])
        writer.writerow(["Umsatzsteuer gesamt", "", "", "", fmt(total_ust)])
        writer.writerow(["Vorsteuer gesamt", "", "", "", fmt(total_vst)])
        writer.writerow(["Zahllast", "", "", "", fmt(zahllast)])
        return output.getvalue()

    filename = f"ust_{label.replace('/', '_')}.csv"
    return Response(
        generate(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Kundenauswertung
# ---------------------------------------------------------------------------

import re as _re


def _hn_sort_key(hn):
    """Hausnummer natürlich sortieren: '2a' < '10' < '10a'."""
    if not hn:
        return (999999, "")
    m = _re.match(r"^(\d+)(.*)", hn.strip())
    if m:
        return (int(m.group(1)), m.group(2).strip().lower())
    return (999999, hn.lower())


def _build_kunden_rows(q_name, q_kunden_nr, q_zaehler, q_objekt_adresse,
                       q_rechnungsadresse, modus, sort):
    """
    Gibt eine Liste von (kunde, [(ownership, prop, [(meter, reading)])]) zurück.
    """
    import sqlalchemy as _sa

    kunden_q = Customer.query.filter_by(is_customer=True, active=True)

    if q_name:
        kunden_q = kunden_q.filter(Customer.name.ilike(f"%{q_name}%"))

    if q_kunden_nr:
        try:
            nr = int(q_kunden_nr)
            kunden_q = kunden_q.filter(Customer.customer_number == nr)
        except ValueError:
            kunden_q = kunden_q.filter(_sa.false())

    if q_rechnungsadresse:
        s = f"%{q_rechnungsadresse}%"
        kunden_q = kunden_q.filter(db.or_(
            Customer.strasse.ilike(s),
            Customer.hausnummer.ilike(s),
            Customer.plz.ilike(s),
            Customer.ort.ilike(s),
        ))

    if sort == "name":
        kunden_q = kunden_q.order_by(Customer.name.asc())
    else:
        kunden_q = kunden_q.order_by(
            _sa.case((Customer.customer_number.is_(None), 1), else_=0).asc(),
            Customer.customer_number.asc(),
            Customer.name.asc(),
        )

    kunden = kunden_q.all()

    result = []
    for kunde in kunden:
        ownerships = (
            PropertyOwnership.query
            .filter_by(customer_id=kunde.id, valid_to=None)
            .join(Property, PropertyOwnership.property_id == Property.id)
            .filter(Property.active == True)  # noqa: E712
            .all()
        )

        if q_objekt_adresse:
            s = q_objekt_adresse.lower()
            ownerships = [
                o for o in ownerships
                if s in (o.property.address_display() or "").lower()
            ]
            # Objekt-Filter aktiv → Kunde ohne passendes Objekt ausblenden
            if not ownerships:
                continue

        if modus == "mit_objekt" and not ownerships:
            continue

        ownership_data = []
        for ownership in ownerships:
            prop = ownership.property
            meters = prop.meters.filter(WaterMeter.active == True).all()  # noqa: E712

            if q_zaehler:
                s = q_zaehler.lower()
                meters = [m for m in meters if s in m.meter_number.lower()]
                # Zähler-Filter aktiv → Objekt ohne passenden Zähler überspringen
                if not meters:
                    continue

            meters_with_reading = [(m, m.last_reading()) for m in meters]
            ownership_data.append((ownership, prop, meters_with_reading))

        # Zähler-Filter aktiv → Kunde ohne passendes Ergebnis ausblenden
        if q_zaehler and not ownership_data:
            continue

        if modus == "aktiver_zaehler":
            if not any(mwr for _, _, mwr in ownership_data):
                continue

        result.append((kunde, ownership_data))

    if sort == "objektstrasse":
        def _obj_key(item):
            _, ownerships_list = item
            if not ownerships_list:
                return ("zzz", 999999, "")
            p = ownerships_list[0][1]
            return ((p.strasse or "zzz").lower(), *_hn_sort_key(p.hausnummer))
        result.sort(key=_obj_key)
        for _, ownerships_list in result:
            ownerships_list.sort(
                key=lambda t: ((t[1].strasse or "zzz").lower(), *_hn_sort_key(t[1].hausnummer))
            )

    return result


def _flatten_rows(result):
    """
    Wandelt die strukturierten Daten in eine flache Liste von Dicts um,
    mit rowspan-Infos für Kunde und Objekt.
    """
    flat = []
    for kunde, ownerships_list in result:
        kunde_total = max(sum(max(len(mwr), 1) for _, _, mwr in ownerships_list), 1)
        first_kunde = True

        if not ownerships_list:
            flat.append(dict(
                kunde=kunde, kunde_rowspan=1,
                ownership=None, prop=None, prop_rowspan=1,
                meter=None, reading=None,
            ))
            continue

        for ownership, prop, meters_with_reading in ownerships_list:
            prop_total = max(len(meters_with_reading), 1)
            first_prop = True

            if not meters_with_reading:
                flat.append(dict(
                    kunde=kunde if first_kunde else None,
                    kunde_rowspan=kunde_total if first_kunde else 0,
                    ownership=ownership, prop=prop, prop_rowspan=prop_total,
                    meter=None, reading=None,
                ))
                first_kunde = False
                continue

            for meter, reading in meters_with_reading:
                flat.append(dict(
                    kunde=kunde if first_kunde else None,
                    kunde_rowspan=kunde_total if first_kunde else 0,
                    ownership=ownership if first_prop else None,
                    prop=prop if first_prop else None,
                    prop_rowspan=prop_total if first_prop else 0,
                    meter=meter, reading=reading,
                ))
                first_kunde = False
                first_prop = False

    return flat


@bp.route("/kundenauswertung")
@login_required
def kundenauswertung():
    q_name = request.args.get("q_name", "").strip()
    q_kunden_nr = request.args.get("q_kunden_nr", "").strip()
    q_zaehler = request.args.get("q_zaehler", "").strip()
    q_objekt_adresse = request.args.get("q_objekt_adresse", "").strip()
    q_rechnungsadresse = request.args.get("q_rechnungsadresse", "").strip()
    modus = request.args.get("modus", "alle")
    sort = request.args.get("sort", "kundennummer")

    result = _build_kunden_rows(
        q_name, q_kunden_nr, q_zaehler, q_objekt_adresse,
        q_rechnungsadresse, modus, sort,
    )
    flat = _flatten_rows(result)

    return render_template(
        "accounting/kunden.html",
        flat_rows=flat,
        total_kunden=len(result),
        modus=modus, sort=sort,
        q_name=q_name, q_kunden_nr=q_kunden_nr, q_zaehler=q_zaehler,
        q_objekt_adresse=q_objekt_adresse, q_rechnungsadresse=q_rechnungsadresse,
    )


@bp.route("/kundenauswertung/export")
@login_required
def kundenauswertung_export():
    q_name = request.args.get("q_name", "").strip()
    q_kunden_nr = request.args.get("q_kunden_nr", "").strip()
    q_zaehler = request.args.get("q_zaehler", "").strip()
    q_objekt_adresse = request.args.get("q_objekt_adresse", "").strip()
    q_rechnungsadresse = request.args.get("q_rechnungsadresse", "").strip()
    modus = request.args.get("modus", "alle")
    sort = request.args.get("sort", "kundennummer")
    fmt = request.args.get("format", "csv")

    result = _build_kunden_rows(
        q_name, q_kunden_nr, q_zaehler, q_objekt_adresse,
        q_rechnungsadresse, modus, sort,
    )

    # Flache Zeilen ohne rowspan (Export: Daten wiederholen)
    export_rows = []
    for kunde, ownerships_list in result:
        if not ownerships_list:
            export_rows.append((kunde, None, None, None, None))
            continue
        for ownership, prop, meters_with_reading in ownerships_list:
            if not meters_with_reading:
                export_rows.append((kunde, ownership, prop, None, None))
                continue
            for meter, reading in meters_with_reading:
                export_rows.append((kunde, ownership, prop, meter, reading))

    def _d(v):
        return v.strftime("%d.%m.%Y") if v else ""

    def _n(v, dec=3):
        if v is None:
            return ""
        return str(Decimal(str(v)).quantize(Decimal("0." + "0" * dec))).replace(".", ",")

    headers = [
        # Kunde
        "Kd.-Nr.", "Ext. Kennung", "Name",
        "Rechnungsadr. Straße+Nr", "Rechnungsadr. PLZ/Ort",
        "E-Mail", "Telefon", "Mitglied seit", "Rechnung p. E-Mail",
        # Objekt
        "Obj.-Nr.", "Obj.-Typ", "Obj. Straße+Nr", "Obj. PLZ/Ort",
        "Eigentümer seit",
        # Zähler
        "Zähler-Nr.", "Zähler-Typ", "Standort",
        "Einbaudatum", "Ausbaudatum", "Eichjahr", "Anfangswert (m³)",
        # Zählerstand
        "Letzter Stand Datum", "Letzter Stand (m³)", "Verbrauch (m³)",
    ]

    def _row(k, ow, p, m, r):
        return [
            k.customer_number or "",
            k.externe_kennung or "",
            k.name,
            " ".join(filter(None, [k.strasse, k.hausnummer])),
            " ".join(filter(None, [k.plz, k.ort])),
            k.email or "",
            k.phone or "",
            _d(k.member_since),
            "Ja" if k.rechnung_per_email else "Nein",
            p.object_number if p else "",
            p.object_type if p else "",
            " ".join(filter(None, [p.strasse, p.hausnummer])) if p else "",
            " ".join(filter(None, [p.plz, p.ort])) if p else "",
            _d(ow.valid_from) if ow else "",
            m.meter_number if m else "",
            m.type_label() if m else "",
            (m.location or "") if m else "",
            _d(m.installed_from) if m else "",
            _d(m.installed_to) if m else "",
            str(m.eichjahr) if m and m.eichjahr else "",
            _n(m.initial_value) if m else "",
            _d(r.reading_date) if r else "",
            _n(r.value) if r else "",
            _n(r.consumption) if r else "",
        ]

    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kundenauswertung"

        # Blockköpfe
        blocks = [
            ("Kunde", 9, "1F497D"),
            ("Objekt", 5, "2F6E3B"),
            ("Zähler", 7, "7B3F00"),
            ("Letzter Zählerstand", 3, "4A4A4A"),
        ]
        col = 1
        for label, span, color in blocks:
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + span - 1,
            )
            c = ws.cell(row=1, column=col, value=label)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="center")
            col += span

        # Spaltenköpfe
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D9D9D9")

        # Daten
        for ri, (k, ow, p, m, r) in enumerate(export_rows, 3):
            for ci, val in enumerate(_row(k, ow, p, m, r), 1):
                ws.cell(row=ri, column=ci, value=val)

        # Spaltenbreiten
        col_widths = [8, 12, 25, 22, 16, 22, 14, 14, 10,
                      10, 10, 22, 16, 14,
                      16, 10, 16, 12, 12, 8, 14,
                      14, 14, 14]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(ci)
            ].width = w

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=kundenauswertung.xlsx"},
        )

    # CSV-Export
    def generate_csv():
        out = io.StringIO()
        out.write("﻿")  # UTF-8 BOM für Excel
        w = csv.writer(out, delimiter=";")
        w.writerow(headers)
        for k, ow, p, m, r in export_rows:
            w.writerow(_row(k, ow, p, m, r))
        return out.getvalue()

    return Response(
        generate_csv(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=kundenauswertung.csv"},
    )


# ---------------------------------------------------------------------------
# Reale Bankkonten
# ---------------------------------------------------------------------------

@bp.route("/real-accounts")
@login_required
def real_accounts():
    current_year = date.today().year
    accounts = RealAccount.query.order_by(RealAccount.name).all()

    account_data = []
    for ra in accounts:
        # Alle Jahre mit Buchungen oder gespeicherten Abschlüssen ermitteln
        booking_years = db.session.query(
            extract("year", Booking.date).label("y")
        ).filter(Booking.real_account_id == ra.id).distinct()
        transfer_years_in = db.session.query(
            extract("year", Transfer.date).label("y")
        ).filter(Transfer.to_real_account_id == ra.id).distinct()
        transfer_years_out = db.session.query(
            extract("year", Transfer.date).label("y")
        ).filter(Transfer.from_real_account_id == ra.id).distinct()
        closed_years = {yb.year for yb in ra.year_balances.all()}

        all_years = set()
        for row in booking_years:
            all_years.add(int(row.y))
        for row in transfer_years_in:
            all_years.add(int(row.y))
        for row in transfer_years_out:
            all_years.add(int(row.y))
        all_years |= closed_years
        all_years.add(current_year)

        # Jahres-History aufbauen
        history = []
        for y in sorted(all_years):
            jan1 = _jan1_balance(ra, y)
            income, expense, year_total = _year_movements(ra.id, y)
            closing = jan1 + year_total
            is_closed = y in closed_years
            history.append({
                "year": y,
                "jan1": jan1,
                "income": income,
                "expense": expense,
                "year_total": year_total,
                "closing": closing,
                "is_closed": is_closed,
                "is_current": y == current_year,
            })

        account_data.append({"ra": ra, "history": history})

    return render_template(
        "accounting/real_accounts.html",
        account_data=account_data,
        current_year=current_year,
    )


@bp.route("/real-accounts/new", methods=["GET", "POST"])
@login_required
def real_account_new():
    if request.method == "POST":
        opening_raw = request.form.get("opening_balance", "0").replace(",", ".")
        set_default = "is_default" in request.form
        if set_default:
            RealAccount.query.filter_by(is_default=True).update({"is_default": False})
        ra = RealAccount(
            name=request.form["name"].strip(),
            description=request.form.get("description", "").strip(),
            iban=request.form.get("iban", "").strip(),
            opening_balance=Decimal(opening_raw),
            icon=request.form.get("icon", "fa-university").strip() or "fa-university",
            is_default=set_default,
        )
        db.session.add(ra)
        db.session.commit()
        flash("Bankkonto angelegt.", "success")
        return redirect(url_for("accounting.real_accounts"))
    return render_template("accounting/real_account_form.html", real_account=None)


@bp.route("/real-accounts/<int:ra_id>/edit", methods=["GET", "POST"])
@login_required
def real_account_edit(ra_id):
    ra = db.get_or_404(RealAccount, ra_id)
    if request.method == "POST":
        opening_raw = request.form.get("opening_balance", "0").replace(",", ".")
        set_default = "is_default" in request.form
        if set_default:
            RealAccount.query.filter(RealAccount.id != ra.id, RealAccount.is_default == True).update({"is_default": False})
        ra.name = request.form["name"].strip()
        ra.description = request.form.get("description", "").strip()
        ra.iban = request.form.get("iban", "").strip()
        ra.opening_balance = Decimal(opening_raw)
        ra.active = "active" in request.form
        ra.icon = request.form.get("icon", "fa-university").strip() or "fa-university"
        ra.is_default = set_default
        db.session.commit()
        flash("Bankkonto aktualisiert.", "success")
        return redirect(url_for("accounting.real_accounts"))
    return render_template("accounting/real_account_form.html", real_account=ra)


# ---------------------------------------------------------------------------
# Umbuchungen
# ---------------------------------------------------------------------------

@bp.route("/transfers")
@login_required
def transfers():
    year = request.args.get("year", date.today().year, type=int)
    transfers_list = (
        Transfer.query
        .filter(extract("year", Transfer.date) == year)
        .order_by(Transfer.date.desc(), Transfer.id.desc())
        .all()
    )
    real_accounts = RealAccount.query.filter_by(active=True).order_by(RealAccount.name).all()
    years = db.session.query(extract("year", Transfer.date).label("y")).distinct().order_by("y").all()
    all_years = sorted({t.y for t in years} | {date.today().year}, reverse=True)
    return render_template(
        "accounting/transfers.html",
        transfers=transfers_list,
        year=year,
        all_years=all_years,
        real_accounts=real_accounts,
    )


@bp.route("/transfers/new", methods=["GET", "POST"])
@login_required
def transfer_new():
    real_accounts = RealAccount.query.filter_by(active=True).order_by(RealAccount.name).all()
    if request.method == "POST":
        transfer_date = date.fromisoformat(request.form["date"])
        fy_error = acc_svc.open_fiscal_year_error(transfer_date)
        if fy_error:
            flash(f"{fy_error} Umbuchung nicht möglich.", "danger")
            return render_template("accounting/transfer_form.html", real_accounts=real_accounts, today=date.today().isoformat())

        amount_raw = request.form["amount"].replace(",", ".")
        try:
            amount = Decimal(amount_raw)
        except InvalidOperation:
            flash("Ungültiger Betrag.", "danger")
            return render_template("accounting/transfer_form.html", real_accounts=real_accounts, today=date.today().isoformat())

        if amount <= 0:
            flash("Der Betrag muss größer als 0 sein.", "danger")
            return render_template("accounting/transfer_form.html", real_accounts=real_accounts, today=date.today().isoformat())

        from_id = int(request.form["from_real_account_id"])
        to_id = int(request.form["to_real_account_id"])
        if from_id == to_id:
            flash("Ausgangs- und Zielkonto dürfen nicht gleich sein.", "danger")
            return render_template("accounting/transfer_form.html", real_accounts=real_accounts, today=date.today().isoformat())

        t = Transfer(
            date=transfer_date,
            amount=amount,
            description=request.form["description"].strip(),
            from_real_account_id=from_id,
            to_real_account_id=to_id,
            created_by_id=current_user.id,
        )
        db.session.add(t)
        db.session.commit()
        flash("Umbuchung gespeichert.", "success")
        return redirect(url_for("accounting.transfers"))

    return render_template("accounting/transfer_form.html", real_accounts=real_accounts, today=date.today().isoformat())


@bp.route("/transfers/<int:transfer_id>/delete", methods=["POST"])
@login_required
def transfer_delete(transfer_id):
    t = db.get_or_404(Transfer, transfer_id)
    locked = _locked_fiscal_year(t.date)
    if locked:
        flash(f"Das Buchungsjahr {locked.year} ist abgeschlossen. Löschen nicht möglich.", "danger")
        return redirect(url_for("accounting.transfers"))
    db.session.delete(t)
    db.session.commit()
    flash("Umbuchung gelöscht.", "success")
    return redirect(url_for("accounting.transfers"))


# ---------------------------------------------------------------------------
# Buchungsjahre
# ---------------------------------------------------------------------------

@bp.route("/fiscal-years")
@login_required
def fiscal_years():
    years = FiscalYear.query.order_by(FiscalYear.year.desc()).all()
    return render_template("accounting/fiscal_years.html", fiscal_years=years)


@bp.route("/fiscal-years/new", methods=["GET", "POST"])
@login_required
def fiscal_year_new():
    is_modal = bool(request.headers.get("X-From-Modal"))
    today = date.today()
    default_year = today.year
    default_start = date(default_year, 1, 1).isoformat()
    default_end = date(default_year, 12, 31).isoformat()

    def _body():
        return render_template(
            "accounting/_fiscal_year_form_body.html",
            edit_mode=False, default_year=default_year,
            default_start=default_start, default_end=default_end,
        )

    if request.method == "POST":
        year = int(request.form["year"])
        if FiscalYear.query.get(year):
            flash(f"Buchungsjahr {year} existiert bereits.", "warning")
            return _body() if is_modal else redirect(url_for("accounting.fiscal_year_new"))
        fy = FiscalYear(
            year=year,
            start_date=date.fromisoformat(request.form["start_date"]),
            end_date=date.fromisoformat(request.form["end_date"]),
            is_vat_liable=bool(request.form.get("is_vat_liable")),
        )
        db.session.add(fy)
        db.session.commit()
        flash(f"Buchungsjahr {year} angelegt.", "success")
        if is_modal:
            return _fiscal_year_modal_saved(fy.year)
        return redirect(url_for("accounting.fiscal_years"))

    if is_modal:
        return _body()
    return render_template(
        "accounting/fiscal_year_form.html",
        default_year=default_year,
        default_start=default_start,
        default_end=default_end,
        edit_mode=False,
    )


@bp.route("/fiscal-years/<int:year>/edit", methods=["GET", "POST"])
@login_required
def fiscal_year_edit(year):
    fy = db.get_or_404(FiscalYear, year)
    is_modal = bool(request.headers.get("X-From-Modal"))

    def _body():
        return render_template(
            "accounting/_fiscal_year_form_body.html", fiscal_year=fy, edit_mode=True,
        )

    if request.method == "POST":
        if fy.closed:
            flash(
                f"Buchungsjahr {year} ist abgeschlossen und kann nicht bearbeitet werden.",
                "warning",
            )
            return _body() if is_modal else redirect(url_for("accounting.fiscal_years"))
        try:
            fy.start_date = date.fromisoformat(request.form["start_date"])
            fy.end_date = date.fromisoformat(request.form["end_date"])
            fy.is_vat_liable = bool(request.form.get("is_vat_liable"))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Speichern: {e}", "danger")
            return _body() if is_modal else redirect(
                url_for("accounting.fiscal_year_edit", year=year))
        flash(f"Buchungsjahr {year} aktualisiert.", "success")
        if is_modal:
            return _fiscal_year_modal_saved(fy.year)
        return redirect(url_for("accounting.fiscal_years"))

    if is_modal:
        return _body()
    return render_template(
        "accounting/fiscal_year_form.html",
        fiscal_year=fy,
        edit_mode=True,
    )


def _fiscal_year_modal_saved(year):
    """204 + HX-Trigger fuer das Buchungsjahr-Modal (schliessen + neu laden)."""
    resp = make_response("", 204)
    resp.headers["HX-Trigger"] = json.dumps({
        "closeFiscalYearModal": True,
        "fiscalYearSaved": {"year": year},
    })
    return resp


@bp.route("/fiscal-years/<int:year>/close", methods=["GET", "POST"])
@login_required
def fiscal_year_close(year):
    from datetime import datetime as _dt
    fy = db.get_or_404(FiscalYear, year)
    if fy.closed:
        flash(f"Buchungsjahr {year} ist bereits abgeschlossen.", "warning")
        return redirect(url_for("accounting.fiscal_years"))

    real_accs = RealAccount.query.order_by(RealAccount.name).all()

    if request.method == "POST":
        try:
            # Jahresabschlussstand pro Bankkonto speichern
            for ra in real_accs:
                dec31 = _year_end_balance(ra, year)
                existing = RealAccountYearBalance.query.filter_by(
                    real_account_id=ra.id, year=year
                ).first()
                if existing:
                    existing.closing_balance = dec31
                else:
                    db.session.add(RealAccountYearBalance(
                        real_account_id=ra.id,
                        year=year,
                        closing_balance=dec31,
                    ))
            fy.closed = True
            fy.closed_at = _dt.utcnow()
            fy.closed_by_id = current_user.id
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Jahresabschluss – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("accounting.fiscal_years"))
        flash(f"Buchungsjahr {year} wurde abgeschlossen.", "success")
        return redirect(url_for("accounting.fiscal_years"))

    # GET – Zusammenfassung über zentralen Service ermitteln
    summary = acc_svc.fiscal_year_close_summary(year)

    return render_template(
        "accounting/fiscal_year_close_confirm.html",
        fiscal_year=fy,
        summary=summary,
    )


@bp.route("/fiscal-years/<int:year>/reopen", methods=["GET", "POST"])
@login_required
def fiscal_year_reopen(year):
    fy = db.get_or_404(FiscalYear, year)
    if not fy.closed:
        flash(f"Buchungsjahr {year} ist nicht abgeschlossen.", "warning")
        return redirect(url_for("accounting.fiscal_years"))
    if request.method == "POST":
        reason = request.form.get("reason", "").strip()
        if not reason:
            flash("Bitte einen Grund für die Wiederöffnung angeben.", "danger")
            return render_template("accounting/fiscal_year_reopen_form.html", fiscal_year=fy)
        try:
            log = FiscalYearReopenLog(
                fiscal_year_id=fy.year,
                reopened_by_id=current_user.id,
                reason=reason,
            )
            db.session.add(log)
            fy.closed = False
            fy.closed_at = None
            fy.closed_by_id = None
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Wiederöffnen – alle Änderungen wurden zurückgesetzt: {e}", "danger")
            return redirect(url_for("accounting.fiscal_years"))
        flash(f"Buchungsjahr {year} wurde wieder geöffnet.", "success")
        return redirect(url_for("accounting.fiscal_years"))
    return render_template("accounting/fiscal_year_reopen_form.html", fiscal_year=fy)
